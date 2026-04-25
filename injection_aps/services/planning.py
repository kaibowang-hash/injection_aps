from __future__ import annotations

import json
from collections import defaultdict
from datetime import date as date_cls
from datetime import datetime, timedelta
from typing import Any

import frappe
from frappe import _
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter
from frappe.utils import add_days, cint, flt, get_datetime, getdate, now_datetime, today
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file

from injection_aps.services.permissions import (
	APS_APPROVE_ROLES,
	APS_PLAN_ROLES,
	APS_RELEASE_ROLES,
)


DEMAND_SOURCE_PRIORITY = {
	"Urgent Order": 1000,
	"Customer Delivery Schedule": 800,
	"Sales Order Backlog": 600,
	"Safety Stock": 400,
	"Trial Production": 300,
	"Complaint Replenishment": 300,
}

RUN_OPEN_STATUSES = ("Draft", "Planned", "Approved", "Work Order Proposed", "Shift Proposed")
LOCKED_SEGMENT_STATUSES = ("Approved", "Work Order Proposed", "Shift Proposed", "Applied")
MANUAL_ADJUSTMENT_BLOCKED_SEGMENT_STATUSES = ("Applied", "Completed")
BLOCKING_WORKSTATION_RISK = "Non FDA"
MAX_REBUILD_WARNINGS = 20
ITEM_NAME_PREFIX_FALLBACKS = ("临时物料:",)
SCHEDULABLE_ITEM_GROUPS = ("Plastic Part", "Sub-assemblies")
BLOCKING_MOLD_STATUSES = (
	"Under Maintenance",
	"Under External Maintenance",
	"Scrapped",
	"Outsourced",
	"Pending Asset Link",
)
APS_ALLOWED_MACHINE_STATUSES = ("Available", "Running", "Setup")
FROZEN_SCHEDULING_STATUSES = ("Material Transfer", "Job Card", "Manufacture")
ACTIVE_SCHEDULING_STATUSES = ("", "Schedule Confirmed", "Material Transfer", "Job Card", "Manufacture")
ANCHOR_STRENGTH_HARD = 100
ANCHOR_STRENGTH_RELEASED = 70
ANCHOR_STRENGTH_LOCKED = 50
ANCHOR_STRENGTH_SOFT = 20
CAPACITY_SOURCE_LABELS = {
	"mold_cycle": "Mold Cycle × Output Per Cycle",
	"machine_hourly_fallback": "APS Machine Hourly Capacity",
	"machine_daily_fallback": "APS Machine Daily Capacity",
	"fallback_cycle": "Fallback Cycle",
	"default_hourly_fallback": "Fallback Hourly Capacity",
}

ACTION_REQUIRED_ROLES = {
	"promote_import": APS_PLAN_ROLES,
	"rebuild_demand_pool": APS_PLAN_ROLES,
	"run_trial": APS_PLAN_ROLES,
	"approve": APS_APPROVE_ROLES,
	"generate_work_order_proposals": APS_RELEASE_ROLES,
	"generate_shift_schedule_proposals": APS_RELEASE_ROLES,
	"apply_work_order_proposals": APS_RELEASE_ROLES,
	"apply_shift_schedule_proposals": APS_RELEASE_ROLES,
}


class APSItemReferenceError(frappe.ValidationError):
	pass


def _normalize_item_code(value: str | None) -> str:
	return _resolve_item_name(value) or (value or "")


def _build_campaign_key(item_code: str | None, mould_reference: str | None, workstation: str | None) -> str:
	item_code = _normalize_item_code(item_code)
	return "::".join([item_code or "", mould_reference or "", workstation or ""])


def _coerce_plant_floor_list(plant_floors: Any = None, plant_floor: str | None = None) -> list[str]:
	values = []
	if plant_floors:
		parsed = plant_floors
		if isinstance(parsed, str):
			text = parsed.strip()
			if text.startswith("["):
				try:
					parsed = json.loads(text)
				except Exception:
					parsed = [chunk.strip() for chunk in text.replace("\n", ",").split(",") if chunk.strip()]
			else:
				parsed = [chunk.strip() for chunk in text.replace("\n", ",").split(",") if chunk.strip()]
		if not isinstance(parsed, (list, tuple, set)):
			parsed = [parsed]
		for row in parsed:
			value = row.get("plant_floor") if isinstance(row, dict) else row
			if value and str(value).strip() and str(value).strip() not in values:
				values.append(str(value).strip())
	if plant_floor and plant_floor not in values:
		values.append(plant_floor)
	return values


def _normalize_selected_plant_floors(
	company: str | None,
	plant_floors: Any = None,
	plant_floor: str | None = None,
	required: bool = False,
) -> list[str]:
	selected = _coerce_plant_floor_list(plant_floors=plant_floors, plant_floor=plant_floor)
	if not selected:
		if required:
			frappe.throw(_("Select at least one Plant Floor before APS planning."))
		return []
	if not frappe.db.exists("DocType", "Plant Floor"):
		frappe.throw(_("Plant Floor master data is required for APS planning."))
	rows = frappe.get_all(
		"Plant Floor",
		filters={"name": ("in", selected)},
		fields=["name", "company"],
	)
	row_map = {row.name: row for row in rows}
	missing = [row for row in selected if row not in row_map]
	if missing:
		frappe.throw(_("Plant Floor {0} was not found.").format(", ".join(missing)))
	if company:
		invalid = [row for row in selected if (row_map.get(row) or {}).get("company") not in ("", None, company)]
		if invalid:
			frappe.throw(
				_("Plant Floor {0} does not belong to company {1}.").format(", ".join(invalid), company)
			)
	return selected


def _apply_selected_plant_floors_to_run(run_doc, plant_floors: list[str]):
	plant_floors = _coerce_plant_floor_list(plant_floors=plant_floors)
	run_doc.set("selected_plant_floors", [])
	for plant_floor in plant_floors:
		run_doc.append("selected_plant_floors", {"plant_floor": plant_floor})
	run_doc.plant_floor = plant_floors[0] if len(plant_floors) == 1 else None
	run_doc.selected_plant_floor_summary = ", ".join(plant_floors)


def _get_run_selected_plant_floors(run_doc) -> list[str]:
	rows = [row.plant_floor for row in (run_doc.get("selected_plant_floors") or []) if row.plant_floor]
	if not rows and getattr(run_doc, "plant_floor", None):
		rows = [run_doc.plant_floor]
	return _coerce_plant_floor_list(rows)


def _get_primary_result_plant_floor(segments: list[dict[str, Any]], fallback: str | None = None) -> str | None:
	for row in segments or []:
		if row.get("segment_kind") == "Family Co-Product":
			continue
		if row.get("plant_floor"):
			return row.get("plant_floor")
	return fallback


def _serialize_diagnostic_json(diagnostic: Any = None, diagnostic_json: str | None = None) -> str:
	if diagnostic_json:
		return str(diagnostic_json)
	if diagnostic in (None, "", {}):
		return ""
	try:
		return json.dumps(diagnostic, ensure_ascii=False, sort_keys=True)
	except Exception:
		return ""


def _parse_diagnostic_json(value: Any) -> dict[str, Any]:
	if not value:
		return {}
	if isinstance(value, dict):
		return value
	try:
		parsed = json.loads(value)
	except Exception:
		return {}
	return parsed if isinstance(parsed, dict) else {}


def _label_run_status(status: str | None) -> str:
	return {
		"Draft": "Draft",
		"Planned": "Recalculated",
		"Approved": "Approved",
		"Work Order Proposed": "Work Order Proposals Generated",
		"Shift Proposed": "Day/Night Shift Proposals Generated",
		"Applied": "Formally Applied",
	}.get(status or "", status or "")


def _label_approval_state(state: str | None) -> str:
	return {
		"Pending": "Pending Approval",
		"Approved": "Approved",
		"Rejected": "Rejected",
	}.get(state or "", state or "")


def _build_planning_run_context(doc) -> dict[str, Any]:
	has_applied_wo_batch = bool(
		frappe.db.exists(
			"APS Work Order Proposal Batch",
			{"planning_run": doc.name, "status": "Applied"},
		)
	)
	has_applied_shift_batch = bool(
		frappe.db.exists(
			"APS Shift Schedule Proposal Batch",
			{"planning_run": doc.name, "status": "Applied"},
		)
	)
	blocking_reason = ""
	current_step = _label_run_status(doc.status or "Draft")
	next_step = "Confirm Run"
	if doc.status == "Draft":
		next_step = "Recalculate"
	elif doc.approval_state != "Approved":
		next_step = "Confirm Run"
	elif has_applied_shift_batch or doc.status == "Applied":
		next_step = "Monitor Execution Drift"
	elif has_applied_wo_batch:
		next_step = "Shift Proposals"
	elif doc.status == "Approved":
		next_step = "Work Order Proposals"
	elif doc.status == "Work Order Proposed":
		next_step = "Review Work Order Proposals"
	elif doc.status == "Shift Proposed":
		next_step = "Review Day/Night Shift Proposals"

	if cint(doc.exception_count) and doc.status in ("Planned", "Approved", "Work Order Proposed", "Shift Proposed"):
		blocking_reason = "There are still {0} APS exceptions waiting for review.".format(doc.exception_count)

	selected_plant_floors = _get_run_selected_plant_floors(doc)
	return {
		"doctype": "APS Planning Run",
		"docname": doc.name,
		"current_step": current_step,
		"next_step": next_step,
		"blocking_reason": blocking_reason,
		"company": doc.company,
		"selected_plant_floors": selected_plant_floors,
		"selected_plant_floor_summary": ", ".join(selected_plant_floors),
		"horizon_days": cint(doc.horizon_days or 0),
		"status": doc.status,
		"status_label": _label_run_status(doc.status),
		"approval_state": doc.approval_state,
		"approval_state_label": _label_approval_state(doc.approval_state),
		"exception_count": cint(doc.exception_count or 0),
		"planning_date": doc.planning_date,
		"modified": doc.modified,
		"run_route": _build_form_route("APS Planning Run", doc.name),
	}


def get_recent_run_contexts(limit: int = 8) -> list[dict[str, Any]]:
	rows = frappe.get_all(
		"APS Planning Run",
		filters={"status": ("in", RUN_OPEN_STATUSES)},
		fields=[
			"name",
			"company",
			"plant_floor",
			"selected_plant_floor_summary",
			"planning_date",
			"horizon_days",
			"status",
			"approval_state",
			"exception_count",
			"modified",
		],
		order_by="modified desc",
		limit=limit,
	)
	context_rows = []
	for row in rows:
		doc = frappe._dict(row)
		context = _build_planning_run_context(doc)
		context_rows.append(
			{
				"name": row.name,
				"company": row.company,
				"selected_plant_floors": context.get("selected_plant_floors") or _coerce_plant_floor_list(
					plant_floors=(row.selected_plant_floor_summary or "").split(","), plant_floor=row.plant_floor
				),
				"horizon_days": cint(row.horizon_days or 0),
				"status": row.status,
				"status_label": context.get("status_label"),
				"approval_state": row.approval_state,
				"approval_state_label": context.get("approval_state_label"),
				"exception_count": cint(row.exception_count or 0),
				"modified": row.modified,
				"route": f"aps-run-console?run_name={row.name}",
				"gantt_route": f"aps-schedule-gantt?run_name={row.name}",
				"execution_route": f"aps-release-center?run_name={row.name}",
			}
		)
	return context_rows


def preview_customer_delivery_schedule(
	customer: str,
	company: str,
	version_no: str,
	schedule_scope: str | None = None,
	import_strategy: str | None = None,
	file_url: str | None = None,
	rows_json: str | list[dict] | None = None,
	mapping_json: str | dict[str, Any] | None = None,
) -> dict[str, Any]:
	schedule_scope = _normalize_schedule_scope(schedule_scope or version_no)
	import_strategy = _normalize_schedule_import_strategy(import_strategy)
	rows, parse_context = _normalize_schedule_rows(
		file_url=file_url,
		rows_json=rows_json,
		mapping_json=mapping_json,
	)
	diff_rows = compare_schedule_against_active(
		customer=customer,
		company=company,
		schedule_scope=schedule_scope,
		import_strategy=import_strategy,
		rows=rows,
	)
	return {
		"customer": customer,
		"company": company,
		"schedule_scope": schedule_scope,
		"import_strategy": import_strategy,
		"version_no": version_no,
		"row_count": len(diff_rows),
		"summary": _summarize_change_types(diff_rows),
		"rows": diff_rows,
		"parse_context": parse_context,
	}


def import_customer_delivery_schedule(
	customer: str,
	company: str,
	version_no: str,
	schedule_scope: str | None = None,
	import_strategy: str | None = None,
	file_url: str | None = None,
	rows_json: str | list[dict] | None = None,
	mapping_json: str | dict[str, Any] | None = None,
	source_type: str = "Customer Delivery Schedule",
) -> dict[str, Any]:
	preview = preview_customer_delivery_schedule(
		customer=customer,
		company=company,
		version_no=version_no,
		schedule_scope=schedule_scope,
		import_strategy=import_strategy,
		file_url=file_url,
		rows_json=rows_json,
		mapping_json=mapping_json,
	)
	schedule_scope = preview.get("schedule_scope")
	import_strategy = preview.get("import_strategy")

	import_batch = frappe.get_doc(
		{
			"doctype": "APS Schedule Import Batch",
			"customer": customer,
			"company": company,
			"schedule_scope": schedule_scope,
			"version_no": version_no,
			"import_strategy": import_strategy,
			"status": "Imported",
			"imported_rows": len(preview["rows"]),
			"effective_rows": sum(1 for row in preview["rows"] if flt(row.get("qty")) > 0),
			"change_summary": json.dumps(preview["summary"], ensure_ascii=True, sort_keys=True),
			"source_type": source_type,
			"uploaded_file": file_url,
			"parser_mode": preview.get("parse_context", {}).get("parser_mode"),
			"sheet_name": preview.get("parse_context", {}).get("sheet_name"),
			"mapping_json": _serialize_diagnostic_json(preview.get("parse_context", {}).get("mapping")),
		}
	).insert(ignore_permissions=True)

	if import_strategy in {"Replace Scope", "Partial Item Update"}:
		for name in frappe.get_all(
			"Customer Delivery Schedule",
			filters={
				"customer": customer,
				"company": company,
				"schedule_scope": schedule_scope,
				"status": "Active",
			},
			pluck="name",
		):
			frappe.db.set_value("Customer Delivery Schedule", name, "status", "Superseded")

	schedule = frappe.get_doc(
		{
			"doctype": "Customer Delivery Schedule",
			"customer": customer,
			"company": company,
			"schedule_scope": schedule_scope,
			"version_no": version_no,
			"import_strategy": import_strategy,
			"import_batch": import_batch.name,
			"source_type": source_type,
			"status": "Active",
			"schedule_total_qty": sum(flt(row.get("qty")) for row in preview["rows"]),
			"change_summary": json.dumps(preview["summary"], ensure_ascii=True, sort_keys=True),
			"items": [
				{
					"sales_order": row.get("sales_order"),
					"item_code": row.get("item_code"),
					"customer_part_no": row.get("customer_part_no"),
					"schedule_date": row.get("schedule_date"),
					"qty": row.get("qty"),
					"allocated_qty": row.get("allocated_qty") or 0,
					"produced_qty": row.get("produced_qty") or 0,
					"delivered_qty": row.get("delivered_qty") or 0,
					"balance_qty": max(flt(row.get("qty")) - flt(row.get("delivered_qty")), 0),
					"change_type": row.get("change_type"),
					"status": "Open" if flt(row.get("qty")) > flt(row.get("delivered_qty")) else "Covered",
					"remark": row.get("remark"),
					"source_origin": row.get("source_origin") or "imported",
					"source_excel_row": cint(row.get("source_excel_row")),
					"manual_override": cint(row.get("manual_override")),
					"manual_change_reason": row.get("manual_change_reason"),
				}
				for row in preview["rows"]
			],
		}
	).insert(ignore_permissions=True)
	_record_schedule_deltas(
		import_batch=import_batch.name,
		schedule_name=schedule.name,
		customer=customer,
		company=company,
		schedule_scope=schedule_scope,
		diff_rows=preview["rows"],
	)

	return {
		"import_batch": import_batch.name,
		"schedule": schedule.name,
		"summary": preview["summary"],
	}


def compare_schedule_against_active(
	customer: str,
	company: str,
	schedule_scope: str | None,
	import_strategy: str | None,
	rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
	schedule_scope = _normalize_schedule_scope(schedule_scope)
	import_strategy = _normalize_schedule_import_strategy(import_strategy)
	previous_rows = _get_active_schedule_rows(customer=customer, company=company, schedule_scope=schedule_scope)
	current_rows = _build_effective_schedule_rows(
		previous_rows=previous_rows,
		incoming_rows=rows,
		import_strategy=import_strategy,
	)
	previous_exact = {_schedule_row_key(row): row for row in previous_rows}
	current_exact = {_schedule_row_key(row): row for row in current_rows}
	processed_previous = set()
	diff_rows = []

	for key in sorted(set(previous_exact) & set(current_exact)):
		previous = previous_exact[key]
		current = dict(current_exact[key])
		current["previous_qty"] = flt(previous.get("qty"))
		current["previous_schedule_date"] = previous.get("schedule_date")
		current["allocated_qty"] = flt(previous.get("allocated_qty"))
		current["produced_qty"] = flt(previous.get("produced_qty"))
		current["delivered_qty"] = flt(previous.get("delivered_qty"))
		current["change_type"] = _detect_change_type(previous, current)
		current["balance_qty"] = max(flt(current.get("qty")) - flt(current.get("delivered_qty")), 0)
		diff_rows.append(current)
		processed_previous.add(key)

	previous_unmatched = [
		row for key, row in previous_exact.items() if key not in processed_previous and key not in current_exact
	]
	current_unmatched = [row for key, row in current_exact.items() if key not in processed_previous and key not in previous_exact]

	previous_grouped = defaultdict(list)
	current_grouped = defaultdict(list)
	for row in previous_unmatched:
		previous_grouped[_schedule_identity_key(row)].append(row)
	for row in current_unmatched:
		current_grouped[_schedule_identity_key(row)].append(row)

	for key in sorted(set(previous_grouped) | set(current_grouped)):
		previous_group = sorted(previous_grouped.get(key) or [], key=lambda row: getdate(row.get("schedule_date")))
		current_group = sorted(current_grouped.get(key) or [], key=lambda row: getdate(row.get("schedule_date")))
		pairs = min(len(previous_group), len(current_group))

		for idx in range(pairs):
			previous = previous_group[idx]
			current = dict(current_group[idx])
			current["previous_qty"] = flt(previous.get("qty"))
			current["previous_schedule_date"] = previous.get("schedule_date")
			current["allocated_qty"] = flt(previous.get("allocated_qty"))
			current["produced_qty"] = flt(previous.get("produced_qty"))
			current["delivered_qty"] = flt(previous.get("delivered_qty"))
			current["change_type"] = _detect_change_type(previous, current)
			current["balance_qty"] = max(flt(current.get("qty")) - flt(current.get("delivered_qty")), 0)
			diff_rows.append(current)

		for current in current_group[pairs:]:
			row = dict(current)
			row.setdefault("allocated_qty", 0)
			row.setdefault("produced_qty", 0)
			row.setdefault("delivered_qty", 0)
			row["previous_qty"] = 0
			row["previous_schedule_date"] = None
			row["change_type"] = "Added"
			row["balance_qty"] = max(flt(row.get("qty")) - flt(row.get("delivered_qty")), 0)
			diff_rows.append(row)

		for previous in previous_group[pairs:]:
			row = dict(previous)
			row["previous_qty"] = flt(previous.get("qty"))
			row["previous_schedule_date"] = previous.get("schedule_date")
			row["qty"] = 0
			row["change_type"] = "Cancelled"
			row["balance_qty"] = 0
			diff_rows.append(row)

	sorted_rows = sorted(
		diff_rows,
		key=lambda row: (
			getdate(row.get("schedule_date")),
			row.get("sales_order") or "",
			row.get("item_code") or "",
			row.get("customer_part_no") or "",
		),
	)
	for idx, row in enumerate(sorted_rows, start=1):
		row["line_idx"] = idx
	return sorted_rows


def _normalize_schedule_scope(value: str | None) -> str:
	scope = str(value or "").strip()
	return scope or "Default Scope"


def _normalize_schedule_import_strategy(value: str | None) -> str:
	allowed = {"Replace Scope", "Partial Item Update", "Append"}
	strategy = str(value or "").strip()
	return strategy if strategy in allowed else "Replace Scope"


def _build_effective_schedule_rows(
	previous_rows: list[dict[str, Any]],
	incoming_rows: list[dict[str, Any]],
	import_strategy: str,
) -> list[dict[str, Any]]:
	prepared_rows = _prepare_schedule_rows_for_import(incoming_rows)
	if import_strategy == "Partial Item Update":
		merged_rows = {_schedule_row_key(row): dict(row, source_origin=row.get("source_origin") or "retained_existing") for row in previous_rows}
		for row in prepared_rows:
			merged_rows[_schedule_row_key(row)] = dict(row)
		return list(merged_rows.values())
	return prepared_rows


def _prepare_schedule_rows_for_import(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
	prepared = []
	for row in rows or []:
		item_code = row.get("item_code")
		if not item_code:
			continue
		prepared.append(
			{
				"sales_order": row.get("sales_order"),
				"item_code": _resolve_item_name(item_code) or item_code,
				"customer_part_no": row.get("customer_part_no"),
				"schedule_date": getdate(row.get("schedule_date")) if row.get("schedule_date") else getdate(today()),
				"qty": flt(row.get("qty")),
				"remark": row.get("remark"),
				"source_origin": row.get("source_origin") or "imported",
				"source_excel_row": cint(row.get("source_excel_row")),
				"manual_override": cint(row.get("manual_override")),
				"manual_change_reason": row.get("manual_change_reason"),
			}
		)
	return prepared


def _get_active_schedule_rows(customer: str, company: str, schedule_scope: str | None) -> list[dict[str, Any]]:
	schedule_names = frappe.get_all(
		"Customer Delivery Schedule",
		filters={
			"customer": customer,
			"company": company,
			"schedule_scope": schedule_scope,
			"status": "Active",
		},
		pluck="name",
	)
	if not schedule_names:
		return []
	rows = frappe.get_all(
		"Customer Delivery Schedule Item",
		filters={"parent": ("in", schedule_names), "parenttype": "Customer Delivery Schedule"},
		fields=[
			"parent",
			"sales_order",
			"item_code",
			"customer_part_no",
			"schedule_date",
			"qty",
			"allocated_qty",
			"produced_qty",
			"delivered_qty",
			"balance_qty",
			"remark",
			"source_origin",
			"source_excel_row",
			"manual_override",
			"manual_change_reason",
		],
	)
	return _aggregate_schedule_rows(rows)


def _aggregate_schedule_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
	aggregated = {}
	for row in rows or []:
		key = _schedule_row_key(row)
		existing = aggregated.get(key)
		if not existing:
			aggregated[key] = {
				"sales_order": row.get("sales_order"),
				"item_code": row.get("item_code"),
				"customer_part_no": row.get("customer_part_no"),
				"schedule_date": getdate(row.get("schedule_date")),
				"qty": flt(row.get("qty")),
				"allocated_qty": flt(row.get("allocated_qty")),
				"produced_qty": flt(row.get("produced_qty")),
				"delivered_qty": flt(row.get("delivered_qty")),
				"balance_qty": flt(row.get("balance_qty")),
				"remark": row.get("remark"),
				"source_origin": row.get("source_origin") or "imported",
				"source_excel_row": cint(row.get("source_excel_row")),
				"manual_override": cint(row.get("manual_override")),
				"manual_change_reason": row.get("manual_change_reason"),
			}
			continue
		existing["qty"] += flt(row.get("qty"))
		existing["allocated_qty"] += flt(row.get("allocated_qty"))
		existing["produced_qty"] += flt(row.get("produced_qty"))
		existing["delivered_qty"] += flt(row.get("delivered_qty"))
		existing["balance_qty"] += flt(row.get("balance_qty"))
	return list(aggregated.values())


def _record_schedule_deltas(
	import_batch: str,
	schedule_name: str,
	customer: str,
	company: str,
	schedule_scope: str | None,
	diff_rows: list[dict[str, Any]],
):
	if not frappe.db.exists("DocType", "APS Demand Delta"):
		return
	for row in diff_rows or []:
		previous_qty = flt(row.get("previous_qty"))
		current_qty = flt(row.get("qty"))
		delta_qty = current_qty - previous_qty
		change_type = row.get("change_type") or "Unchanged"
		if change_type == "Unchanged" and abs(delta_qty) < 0.0001:
			continue
		frappe.get_doc(
			{
				"doctype": "APS Demand Delta",
				"import_batch": import_batch,
				"schedule_reference": schedule_name,
				"customer": customer,
				"company": company,
				"schedule_scope": _normalize_schedule_scope(schedule_scope),
				"sales_order": row.get("sales_order"),
				"item_code": _normalize_item_code(row.get("item_code")),
				"customer_part_no": row.get("customer_part_no"),
				"previous_schedule_date": row.get("previous_schedule_date"),
				"current_schedule_date": row.get("schedule_date"),
				"previous_qty": previous_qty,
				"current_qty": current_qty,
				"delta_qty": delta_qty,
				"change_type": change_type,
				"remark": row.get("remark"),
			}
		).insert(ignore_permissions=True)


def rebuild_demand_pool(company: str | None = None) -> dict[str, Any]:
	reference_repair = repair_item_references(company=company, include_standard=0, include_aps=1)
	_delete_system_generated_rows("APS Demand Pool", company=company)

	created_names = []
	warnings = []
	warning_keys = set()
	skipped_rows = 0
	active_schedules = frappe.get_all(
		"Customer Delivery Schedule",
		filters=_strip_none({"company": company, "status": "Active"}),
		fields=["name", "customer", "company", "version_no", "source_type"],
	)

	for schedule in active_schedules:
		for row in frappe.get_all(
			"Customer Delivery Schedule Item",
			filters={"parent": schedule.name, "parenttype": "Customer Delivery Schedule"},
			fields=[
				"name",
				"sales_order",
				"item_code",
				"schedule_date",
				"qty",
				"allocated_qty",
				"produced_qty",
				"delivered_qty",
				"balance_qty",
				"change_type",
				"customer_part_no",
			],
		):
			resolved_item_code = _resolve_item_name(row.item_code)
			if not resolved_item_code:
				skipped_rows += 1
				_append_rebuild_warning(
					warnings,
					warning_keys,
					item_reference=row.item_code,
					source_doctype="Customer Delivery Schedule",
					source_name=schedule.name,
					row_name=row.name,
				)
				continue
			if not _is_schedulable_item(resolved_item_code):
				skipped_rows += 1
				_append_item_group_warning(
					warnings,
					warning_keys,
					item_code=resolved_item_code,
					source_doctype="Customer Delivery Schedule",
					source_name=schedule.name,
					row_name=row.name,
					item_group=_get_item_group(resolved_item_code),
				)
				continue
			if resolved_item_code != row.item_code:
				frappe.db.set_value(
					"Customer Delivery Schedule Item",
					row.name,
					"item_code",
					resolved_item_code,
					update_modified=False,
				)
			balance_qty = row.balance_qty if row.balance_qty not in (None, "") else max(flt(row.qty) - flt(row.delivered_qty), 0)
			open_qty = max(flt(balance_qty) - flt(row.allocated_qty), 0)
			if open_qty <= 0:
				continue
			demand = _build_demand_row(
				company=schedule.company,
				customer=schedule.customer,
				item_code=resolved_item_code,
				demand_source=schedule.source_type or "Customer Delivery Schedule",
				demand_date=row.schedule_date,
				qty=open_qty,
				source_doctype="Customer Delivery Schedule",
				source_name=schedule.name,
				sales_order=row.sales_order,
				remark=row.change_type,
				customer_part_no=row.customer_part_no,
			)
			created_names.append(demand.insert(ignore_permissions=True).name)

	backlog_result = _append_sales_order_backlog(company=company, warnings=warnings, warning_keys=warning_keys)
	created_names.extend(backlog_result["rows"])
	skipped_rows += cint(backlog_result.get("skipped_rows"))
	created_names.extend(_append_safety_stock_demands(company=company))

	return {
		"created_rows": len(created_names),
		"rows": created_names,
		"warning_count": len(warnings),
		"warnings": warnings[:MAX_REBUILD_WARNINGS],
		"skipped_rows": skipped_rows,
		"reference_repair": reference_repair,
	}


def rebuild_net_requirements(company: str | None = None) -> dict[str, Any]:
	reference_repair = repair_item_references(company=company, include_standard=0, include_aps=1)
	_delete_system_generated_rows("APS Net Requirement", company=company)

	demand_rows = frappe.get_all(
		"APS Demand Pool",
		filters=_strip_none({"company": company, "status": ("!=", "Cancelled")}),
		fields=[
			"name",
			"company",
			"customer",
			"item_code",
			"demand_date",
			"qty",
			"demand_source",
			"is_urgent",
		],
		order_by="demand_date asc, priority_score desc, modified asc",
	)
	grouped = defaultdict(list)
	warnings = []
	warning_keys = set()
	skipped_rows = 0
	for row in demand_rows:
		resolved_item_code = _resolve_item_name(row.item_code)
		if not resolved_item_code:
			skipped_rows += 1
			_append_rebuild_warning(
				warnings,
				warning_keys,
				item_reference=row.item_code,
				source_doctype="APS Demand Pool",
				source_name=row.name,
			)
			continue
		if not _is_schedulable_item(resolved_item_code):
			skipped_rows += 1
			_append_item_group_warning(
				warnings,
				warning_keys,
				item_code=resolved_item_code,
				source_doctype="APS Demand Pool",
				source_name=row.name,
				item_group=_get_item_group(resolved_item_code),
			)
			continue
		if resolved_item_code != row.item_code:
			frappe.db.set_value("APS Demand Pool", row.name, "item_code", resolved_item_code, update_modified=False)
			row.item_code = resolved_item_code
		grouped[(row.company, row.customer, resolved_item_code, row.demand_date)].append(row)

	has_safety_demand_by_item = {
		_normalize_item_code(row.item_code)
		for row in demand_rows
		if (row.get("demand_source") or "") == "Safety Stock"
	}
	stock_map = _get_available_stock_map(company)
	open_work_order_map = _get_open_work_order_map(company)
	remaining_stock_map = defaultdict(float, {item: flt(qty) for item, qty in stock_map.items()})
	remaining_work_order_map = defaultdict(float, {item: flt(qty) for item, qty in open_work_order_map.items()})
	safety_gap_remaining_map: dict[str, float] = {}
	settings = get_settings_dict()

	created_names = []
	for (row_company, customer, item_code, demand_date), rows in grouped.items():
		demand_qty = sum(flt(row.qty) for row in rows)
		safety_stock_qty = flt(_get_item_mapping_value(item_code, settings["item_safety_stock_field"]))
		max_stock_qty = flt(_get_item_mapping_value(item_code, settings["item_max_stock_field"]))
		minimum_batch_qty = flt(_get_item_mapping_value(item_code, settings["item_min_batch_field"]))
		if item_code not in safety_gap_remaining_map:
			safety_gap_remaining_map[item_code] = (
				0
				if item_code in has_safety_demand_by_item
				else max(safety_stock_qty - flt(stock_map.get(item_code)), 0)
			)
		available_stock_qty = min(demand_qty, flt(remaining_stock_map[item_code]))
		remaining_stock_map[item_code] = max(flt(remaining_stock_map[item_code]) - available_stock_qty, 0)
		open_qty_after_stock = max(demand_qty - available_stock_qty, 0)
		open_work_order_qty = min(open_qty_after_stock, flt(remaining_work_order_map[item_code]))
		remaining_work_order_map[item_code] = max(flt(remaining_work_order_map[item_code]) - open_work_order_qty, 0)
		safety_gap = flt(safety_gap_remaining_map.get(item_code))
		safety_gap_remaining_map[item_code] = 0
		overstock_qty = max(flt(remaining_stock_map[item_code]) - max_stock_qty, 0) if max_stock_qty else 0
		net_qty = max(demand_qty - available_stock_qty - open_work_order_qty + safety_gap, 0)
		planning_qty = max(net_qty, minimum_batch_qty) if net_qty > 0 and minimum_batch_qty > 0 else net_qty
		reason_text = _build_net_requirement_reason(
			demand_qty=demand_qty,
			available_stock_qty=available_stock_qty,
			open_work_order_qty=open_work_order_qty,
			safety_gap=safety_gap,
			overstock_qty=overstock_qty,
			minimum_batch_qty=minimum_batch_qty,
			planning_qty=planning_qty,
		)

		doc = frappe.get_doc(
			{
				"doctype": "APS Net Requirement",
				"company": row_company,
				"customer": customer,
				"item_code": item_code,
				"demand_date": demand_date,
				"demand_qty": demand_qty,
				"available_stock_qty": available_stock_qty,
				"open_work_order_qty": open_work_order_qty,
				"safety_stock_gap_qty": safety_gap,
				"max_stock_qty": max_stock_qty,
				"overstock_qty": overstock_qty,
				"minimum_batch_qty": minimum_batch_qty,
				"planning_qty": planning_qty,
				"net_requirement_qty": net_qty,
				"reason_text": reason_text,
				"is_system_generated": 1,
			}
		).insert(ignore_permissions=True)
		created_names.append(doc.name)

	return {
		"created_rows": len(created_names),
		"rows": created_names,
		"warning_count": len(warnings),
		"warnings": warnings[:MAX_REBUILD_WARNINGS],
		"skipped_rows": skipped_rows,
		"reference_repair": reference_repair,
	}


def run_planning_run(
	run_name: str | None = None,
	company: str | None = None,
	plant_floor: str | None = None,
	plant_floors: list[str] | str | None = None,
	horizon_days: int | None = None,
	item_code: str | None = None,
	customer: str | None = None,
	run_type: str | None = None,
) -> dict[str, Any]:
	settings = get_settings_dict()
	company = company or settings["default_company"]
	horizon_days = cint(horizon_days or settings["planning_horizon_days"] or 14)
	horizon_start = get_datetime(now_datetime())
	horizon_end = get_datetime(add_days(horizon_start, horizon_days))
	item_code = _resolve_item_name(item_code) if item_code else None

	if run_name:
		run_doc = frappe.get_doc("APS Planning Run", run_name)
		run_doc.company = run_doc.company or company
	else:
		run_doc = frappe.get_doc({"doctype": "APS Planning Run"})
		run_doc.company = company
	selected_plant_floors = _normalize_selected_plant_floors(
		company=run_doc.company or company,
		plant_floors=plant_floors or _get_run_selected_plant_floors(run_doc),
		plant_floor=plant_floor or run_doc.plant_floor,
		required=True,
	)
	run_doc.company = run_doc.company or company
	run_doc.planning_date = run_doc.planning_date or today()
	run_doc.horizon_days = horizon_days
	run_doc.horizon_start = horizon_start
	run_doc.horizon_end = horizon_end
	run_doc.run_type = run_type or run_doc.run_type or "Trial"
	run_doc.status = "Draft"
	run_doc.approval_state = "Pending"
	_apply_selected_plant_floors_to_run(run_doc, selected_plant_floors)
	if run_doc.is_new():
		run_doc.insert(ignore_permissions=True)
	else:
		run_doc.save(ignore_permissions=True)

	demand_rebuild = rebuild_demand_pool(company=run_doc.company)
	net_rebuild = rebuild_net_requirements(company=run_doc.company)

	for name in frappe.get_all("APS Schedule Result", filters={"planning_run": run_doc.name}, pluck="name"):
		frappe.delete_doc("APS Schedule Result", name, force=1, ignore_permissions=True)
	for name in frappe.get_all("APS Exception Log", filters={"planning_run": run_doc.name}, pluck="name"):
		frappe.delete_doc("APS Exception Log", name, force=1, ignore_permissions=True)

	net_rows = frappe.get_all(
		"APS Net Requirement",
		filters=_strip_none(
			{
			"company": run_doc.company,
			"customer": customer,
			"item_code": item_code,
			"net_requirement_qty": (">", 0),
			"demand_date": ("between", [getdate(horizon_start), getdate(horizon_end)]),
			}
		),
		fields=[
			"name",
			"customer",
			"item_code",
			"demand_date",
			"demand_qty",
			"planning_qty",
			"minimum_batch_qty",
			"net_requirement_qty",
			"reason_text",
		],
		order_by="demand_date asc, modified asc",
	)

	capability_rows = _get_machine_capability_rows(plant_floors=selected_plant_floors)
	workstation_state = _build_workstation_state_map(capability_rows)
	locked_segments = _get_locked_segments(selected_plant_floors)
	execution_anchor_rows = _get_execution_anchor_rows(selected_plant_floors)
	mold_state = _build_mold_state_map(locked_segments)
	_apply_locked_segments_to_state(workstation_state, locked_segments)
	_apply_anchor_rows_to_state(workstation_state, mold_state, execution_anchor_rows)

	result_names = []
	exception_names = []
	total_scheduled_qty = 0
	total_unscheduled_qty = 0
	family_credit_map: dict[str, float] = defaultdict(float)

	for row in net_rows:
		original_planning_qty = flt(row.planning_qty or row.net_requirement_qty)
		credit_applied = min(original_planning_qty, flt(family_credit_map.get(row.item_code)))
		if credit_applied:
			family_credit_map[row.item_code] = max(flt(family_credit_map.get(row.item_code)) - credit_applied, 0)
		planning_qty = max(original_planning_qty - credit_applied, 0)
		item_context = _get_item_context(row.item_code, settings)
		demand_source = _get_primary_demand_source(row.item_code, row.customer, row.demand_date)
		best = {
			"scheduled_qty": 0,
			"unscheduled_qty": 0,
			"result_status": "Planned",
			"risk_status": "Normal",
			"segments": [],
			"selected_moulds": [],
			"copy_mold_parallel": 0,
			"family_mold_result": 1 if credit_applied else 0,
			"primary_mould_reference": "",
			"schedule_explanation": "",
			"family_side_outputs": [],
			"family_output_summary": "",
			"exceptions": [],
		}
		if planning_qty > 0:
			candidates = _select_machine_candidates(
				item_code=row.item_code,
				item_context=item_context,
				capability_rows=capability_rows,
				plant_floors=selected_plant_floors,
			)

			best = _choose_best_slot(
				item_code=row.item_code,
				item_context=item_context,
				qty=planning_qty,
				demand_date=row.demand_date,
				horizon_start=horizon_start,
				horizon_end=horizon_end,
				workstation_state=workstation_state,
				mold_state=mold_state,
				candidates=candidates,
				settings=settings,
				selected_plant_floors=selected_plant_floors,
			)
		total_scheduled_for_row = credit_applied + flt(best["scheduled_qty"])
		total_unscheduled_for_row = max(original_planning_qty - total_scheduled_for_row, 0)
		family_messages = []
		if credit_applied:
			family_messages.append(
				_("Covered {0} by prior Family Mold co-production.").format(
					frappe.format(credit_applied, {"fieldtype": "Float"})
				)
			)
		if best.get("family_output_summary"):
			family_messages.append(best["family_output_summary"])
		for side_output in best.get("family_side_outputs") or []:
			family_credit_map[side_output.get("item_code")] = flt(family_credit_map.get(side_output.get("item_code"))) + flt(
				side_output.get("qty")
			)

		flow_step = "Recalculation Completed"
		next_step_hint = "Confirm Run"
		blocking_reason = ""
		if best["result_status"] == "Blocked":
			flow_step = "Blocked"
			next_step_hint = "Handle Exceptions"
			blocking_reason = "; ".join(
				row_error.get("message") for row_error in (best.get("exceptions") or []) if row_error.get("is_blocking")
			)
		elif total_unscheduled_for_row > 0:
			flow_step = "Risk Pending Review"
			next_step_hint = "Review Board and Exceptions"
			blocking_reason = "There is still unscheduled quantity: {0}.".format(total_unscheduled_for_row)

		result_plant_floor = _get_primary_result_plant_floor(best.get("segments") or [], run_doc.plant_floor)
		result_doc = frappe.get_doc(
			{
				"doctype": "APS Schedule Result",
				"planning_run": run_doc.name,
				"company": run_doc.company,
				"plant_floor": result_plant_floor,
				"net_requirement": row.name,
				"customer": row.customer,
				"item_code": row.item_code,
				"requested_date": row.demand_date,
				"demand_source": demand_source,
				"planned_qty": original_planning_qty,
				"scheduled_qty": total_scheduled_for_row,
				"unscheduled_qty": total_unscheduled_for_row,
				"status": best["result_status"],
				"risk_status": best["risk_status"],
				"flow_step": flow_step,
				"next_step_hint": next_step_hint,
				"blocking_reason": blocking_reason,
				"copy_mold_parallel": best.get("copy_mold_parallel") or 0,
				"family_mold_result": best.get("family_mold_result") or (1 if credit_applied else 0),
				"primary_mould_reference": best.get("primary_mould_reference"),
				"selected_moulds": "\n".join(best.get("selected_moulds") or []),
				"schedule_explanation": best.get("schedule_explanation"),
				"family_output_summary": "\n".join(family_messages),
				"is_urgent": 1 if item_context["is_urgent"] else 0,
				"is_locked": 0,
				"is_manual": 0,
				"notes": "\n".join(part for part in [row.reason_text, *family_messages] if part),
				"segments": best["segments"],
			}
		).insert(ignore_permissions=True)
		result_names.append(result_doc.name)
		total_scheduled_qty += flt(total_scheduled_for_row)
		total_unscheduled_qty += flt(total_unscheduled_for_row)

		for error in best["exceptions"]:
			exception_doc = _create_exception(
				planning_run=run_doc.name,
				severity=error["severity"],
				exception_type=error["exception_type"],
				message=error["message"],
				item_code=row.item_code,
				customer=row.customer,
				workstation=error.get("workstation"),
				source_doctype="APS Net Requirement",
				source_name=row.name,
				resolution_hint=error.get("resolution_hint"),
				is_blocking=error.get("is_blocking", 1),
				diagnostic=error.get("diagnostic"),
			)
			exception_names.append(exception_doc.name)

	run_doc.db_set(
		{
			"horizon_days": horizon_days,
			"horizon_start": horizon_start,
			"horizon_end": horizon_end,
			"run_type": run_type or run_doc.run_type or "Trial",
			"status": "Planned",
			"approval_state": "Pending",
			"total_net_requirement_qty": sum(flt(row.planning_qty or row.net_requirement_qty) for row in net_rows),
			"total_scheduled_qty": total_scheduled_qty,
			"total_unscheduled_qty": total_unscheduled_qty,
			"exception_count": len(exception_names),
			"result_count": len(result_names),
		}
	)
	overlap_summary = _validate_run_segment_overlaps(run_doc.name, persist_exceptions=True)
	mold_overlap_summary = _validate_run_mold_overlaps(run_doc.name, persist_exceptions=True)
	if overlap_summary["exception_names"] or mold_overlap_summary["exception_names"]:
		run_doc.db_set(
			"exception_count",
			len(exception_names)
			+ len(overlap_summary["exception_names"])
			+ len(mold_overlap_summary["exception_names"]),
		)

	return {
		"run": run_doc.name,
		"results": result_names,
		"exceptions": exception_names + overlap_summary["exception_names"] + mold_overlap_summary["exception_names"],
		"selected_plant_floors": selected_plant_floors,
		"filters": _strip_none({"item_code": item_code, "customer": customer}),
		"preflight_warning_count": cint(demand_rebuild.get("warning_count")) + cint(net_rebuild.get("warning_count")),
		"preflight_warnings": (demand_rebuild.get("warnings") or []) + (net_rebuild.get("warnings") or []),
		"overlap_count": overlap_summary["count"],
		"mold_overlap_count": mold_overlap_summary["count"],
	}


def approve_planning_run(run_name: str) -> dict[str, Any]:
	run_doc = frappe.get_doc("APS Planning Run", run_name)
	mold_gate = validate_run_mold_readiness(run_name, persist_exceptions=True)
	overlap_summary = _validate_run_segment_overlaps(run_name, persist_exceptions=True)
	mold_overlap_summary = _validate_run_mold_overlaps(run_name, persist_exceptions=True)
	blockers = [row["message"] for row in mold_gate["rows"] if row.get("blocking")]
	if overlap_summary["messages"]:
		blockers.extend(overlap_summary["messages"])
	if mold_overlap_summary["messages"]:
		blockers.extend(mold_overlap_summary["messages"])
	if blockers:
		run_doc.db_set(
			{
				"status": "Planned",
				"approval_state": "Pending",
				"exception_count": frappe.db.count("APS Exception Log", {"planning_run": run_name, "status": "Open"}),
			}
		)
		frappe.throw("<br>".join(blockers[:12]))
	result_names = frappe.get_all("APS Schedule Result", filters={"planning_run": run_name}, pluck="name")
	run_doc.db_set(
		{
			"status": "Approved",
			"approval_state": "Approved",
			"approved_by": frappe.session.user,
			"approved_on": now_datetime(),
		}
	)
	for result_name in result_names:
		frappe.db.set_value("APS Schedule Result", result_name, {"status": "Approved", "flow_step": "Plan Approved", "next_step_hint": "Review Work Order Proposals"})
	if result_names:
		for segment_name in frappe.get_all(
			"APS Schedule Segment",
			filters={"parenttype": "APS Schedule Result", "parent": ("in", result_names)},
			pluck="name",
		):
			frappe.db.set_value(
				"APS Schedule Segment",
				segment_name,
				{
					"segment_status": "Approved",
					"is_locked": 1,
					"anchor_strength": ANCHOR_STRENGTH_LOCKED,
					"execution_anchor_source": "APS Approved Segment",
				},
			)
	return {
		"run": run_name,
		"status": "Approved",
		"mold_gate": mold_gate,
		"overlap_count": overlap_summary["count"],
		"mold_overlap_count": mold_overlap_summary["count"],
	}


def sync_planning_run_to_execution(run_name: str) -> dict[str, Any]:
	return generate_work_order_proposals(run_name)


def release_planning_run(run_name: str, release_horizon_days: int | None = None) -> dict[str, Any]:
	return generate_shift_schedule_proposals(run_name=run_name, release_horizon_days=release_horizon_days)


def generate_work_order_proposals(run_name: str) -> dict[str, Any]:
	run_doc = frappe.get_doc("APS Planning Run", run_name)
	if run_doc.approval_state != "Approved":
		frappe.throw(_("Approve the APS Planning Run before generating work order proposals."))
	mold_gate = validate_run_mold_readiness(run_name, persist_exceptions=True)
	if mold_gate["blocking_count"]:
		frappe.throw(_("Fix mold master blockers before generating work order proposals."))

	items = []
	matched_work_orders = set()
	for result in frappe.get_all(
		"APS Schedule Result",
		filters={"planning_run": run_name, "scheduled_qty": (">", 0), "status": ("!=", "Blocked")},
		fields=["name", "customer", "item_code", "requested_date", "scheduled_qty"],
		order_by="requested_date asc, item_code asc",
	):
		primary_segments = _get_primary_segments_for_result(result.name)
		if not primary_segments:
			continue
		preferred_segment = primary_segments[0]
		preferred_campaign_key = preferred_segment.get("campaign_key") or _build_campaign_key(
			result.item_code,
			preferred_segment.get("mould_reference"),
			preferred_segment.get("workstation"),
		)
		existing = _find_existing_work_order_for_result(
			result.name,
			result.item_code,
			company=run_doc.company,
			preferred_workstation=preferred_segment.get("workstation"),
			preferred_campaign_key=preferred_campaign_key,
			excluded_work_orders=sorted(matched_work_orders),
		)
		proposed_qty = flt(result.scheduled_qty)
		prefer_update_existing = bool(
			existing
			and (
				preferred_campaign_key in (existing.get("campaign_keys") or [])
				or preferred_segment.get("workstation") in (existing.get("workstations") or [])
			)
		)
		action = _classify_work_order_action(
			existing,
			proposed_qty,
			prefer_update_existing=prefer_update_existing,
		)
		existing_qty = 0
		existing_name = None
		if existing:
			existing_qty = flt(existing.get("qty"))
			existing_name = existing.get("name")
			matched_work_orders.add(existing_name)
		review_note = "The system generated a reconciliation proposal against the current execution layer."
		if action == "Update Existing":
			review_note = "Prefer reusing existing work order {0} and update it to the new quantity boundary.".format(existing_name or "-")
		elif action == "Create Delta":
			review_note = "Keep existing work order {0} unchanged and create a delta work order for the extra quantity.".format(existing_name or "-")
		elif action == "Close Residual":
			review_note = "Keep completed quantity and close the remaining unexecuted quantity."
		elif action == "Cancel Unstarted":
			review_note = "The unstarted work order will be cancelled."
		elif action == "Keep Existing":
			review_note = "The existing work order remains the stable execution container."
		items.append(
			{
				"result_reference": result.name,
				"item_code": result.item_code,
				"customer": result.customer,
				"required_delivery_date": result.requested_date,
				"action": action,
				"proposed_qty": proposed_qty,
				"existing_work_order": existing_name,
				"existing_qty": existing_qty,
				"target_start_time": primary_segments[0].get("start_time"),
				"target_end_time": primary_segments[-1].get("end_time"),
				"review_status": "Pending",
				"review_note": review_note,
			}
		)

	for snapshot in _get_open_aps_managed_work_orders(run_doc.company):
		if snapshot.get("name") in matched_work_orders:
			continue
		action = _classify_work_order_action(snapshot, 0)
		if action not in ("Cancel Unstarted", "Close Residual"):
			continue
		result_reference = None
		if snapshot.get("custom_aps_result_reference") and frappe.db.exists("APS Schedule Result", snapshot.get("custom_aps_result_reference")):
			result_reference = snapshot.get("custom_aps_result_reference")
		if not result_reference:
			result_reference = frappe.db.get_value(
				"APS Schedule Result",
				{"planning_run": run_name, "item_code": snapshot.get("production_item")},
				"name",
				order_by="requested_date asc",
			)
		if not result_reference:
			_ensure_open_exception(
				planning_run=run_name,
				severity="Warning",
				exception_type="Unmatched Work Order Residual",
				message=_("Open Work Order {0} no longer matches any current APS result.").format(snapshot.get("name")),
				item_code=snapshot.get("production_item"),
				source_doctype="Work Order",
				source_name=snapshot.get("name"),
				resolution_hint=_("Review whether this work order should be closed or cancelled manually."),
				is_blocking=0,
			)
			continue
		items.append(
			{
				"result_reference": result_reference,
				"item_code": snapshot.get("production_item"),
				"customer": frappe.db.get_value("APS Schedule Result", result_reference, "customer"),
				"required_delivery_date": frappe.db.get_value("APS Schedule Result", result_reference, "requested_date"),
				"action": action,
				"proposed_qty": 0,
				"existing_work_order": snapshot.get("name"),
				"existing_qty": flt(snapshot.get("qty")),
				"target_start_time": snapshot.get("planned_start_date"),
				"target_end_time": snapshot.get("planned_end_date"),
				"review_status": "Pending",
				"review_note": "There is no longer matching APS demand. Recover the unexecuted portion of the existing work order.",
			}
		)

	batch = frappe.get_doc(
		{
			"doctype": "APS Work Order Proposal Batch",
			"planning_run": run_name,
			"company": run_doc.company,
			"plant_floor": run_doc.plant_floor,
			"proposal_date": today(),
			"status": "Ready For Review",
			"approval_state": "Pending",
			"proposal_count": len(items),
			"applied_count": 0,
			"notes": _("Generated from APS Planning Run {0}. Review against existing work orders before formal reconciliation.").format(run_name),
			"items": items,
		}
	).insert(ignore_permissions=True)
	_set_run_result_segment_status(
		run_name=run_name,
		run_status="Work Order Proposed",
		result_status="Work Order Proposed",
		segment_status="Work Order Proposed",
		flow_step="Work Order Proposal Review",
		next_step_hint="Review and apply work order proposals",
	)
	return {"run": run_name, "work_order_proposal_batch": batch.name, "proposal_count": len(items)}


def apply_work_order_proposals(batch_name: str) -> dict[str, Any]:
	batch = frappe.get_doc("APS Work Order Proposal Batch", batch_name)
	run_doc = frappe.get_doc("APS Planning Run", batch.planning_run)
	mold_gate = validate_run_mold_readiness(batch.planning_run, persist_exceptions=True)
	if mold_gate["blocking_count"]:
		frappe.throw(_("Fix mold master blockers before applying work order proposals."))
	approved_rows = [row for row in batch.items if row.review_status == "Approved"]
	if not approved_rows:
		frappe.throw(_("No work order proposal rows are marked Approved. Review the batch before formal creation."))

	settings = get_settings_dict()
	applied_work_orders = []
	applied_result_names = set()
	skipped_rows = []
	for row in approved_rows:
		result_doc = (
			frappe.get_doc("APS Schedule Result", row.result_reference)
			if row.result_reference and frappe.db.exists("APS Schedule Result", row.result_reference)
			else None
		)
		primary_segments = _get_primary_segments_for_result(result_doc.name) if result_doc else []
		start_time = row.target_start_time or (primary_segments[0].get("start_time") if primary_segments else None)
		end_time = row.target_end_time or (primary_segments[-1].get("end_time") if primary_segments else start_time)
		try:
			if row.action == "New":
				if not result_doc:
					raise frappe.ValidationError(_("New work orders require a current APS result reference."))
				work_order_name = _create_formal_work_order(
					run_doc=run_doc,
					result=result_doc,
					qty=flt(row.proposed_qty),
					start_time=start_time,
					end_time=end_time,
					settings=settings,
					proposal_batch=batch.name,
				)
				row.target_work_order = work_order_name
				row.review_status = "Applied"
				row.review_note = _("Formal Work Order {0} created by APS reconciliation.").format(work_order_name)
				applied_work_orders.append(work_order_name)
				applied_result_names.add(result_doc.name)
			elif row.action == "Create Delta":
				if not result_doc:
					raise frappe.ValidationError(_("Delta work orders require a current APS result reference."))
				create_qty = max(flt(row.proposed_qty) - flt(row.existing_qty), 0)
				if create_qty <= 0:
					row.review_status = "Skipped"
					row.review_note = _("No additional delta quantity is required.")
					skipped_rows.append(row.result_reference)
					continue
				work_order_name = _create_formal_work_order(
					run_doc=run_doc,
					result=result_doc,
					qty=create_qty,
					start_time=start_time,
					end_time=end_time,
					settings=settings,
					proposal_batch=batch.name,
				)
				row.target_work_order = work_order_name
				row.review_status = "Applied"
				row.review_note = _("Delta Work Order {0} created while preserving existing container {1}.").format(
					work_order_name,
					row.existing_work_order or "-",
				)
				applied_work_orders.append(work_order_name)
				applied_result_names.add(result_doc.name)
			elif row.action == "Keep Existing" and row.existing_work_order:
				_link_existing_work_order_to_result(
					work_order_name=row.existing_work_order,
					run_name=batch.planning_run,
					result_name=result_doc.name if result_doc else (row.result_reference or ""),
					proposal_batch=batch.name,
					required_delivery_date=row.required_delivery_date,
				)
				row.target_work_order = row.existing_work_order
				row.review_status = "Applied"
				row.review_note = _("Existing Work Order retained as the stable execution container.")
				applied_work_orders.append(row.existing_work_order)
				if result_doc:
					applied_result_names.add(result_doc.name)
			elif row.action == "Update Existing" and row.existing_work_order:
				work_order_name = _update_existing_work_order(
					work_order_name=row.existing_work_order,
					run_name=batch.planning_run,
					result_name=result_doc.name if result_doc else (row.result_reference or ""),
					proposal_batch=batch.name,
					qty=flt(row.proposed_qty),
					start_time=start_time,
					end_time=end_time,
					required_delivery_date=row.required_delivery_date,
				)
				row.target_work_order = work_order_name
				row.review_status = "Applied"
				row.review_note = _("Existing Work Order {0} updated in place.").format(work_order_name)
				applied_work_orders.append(work_order_name)
				if result_doc:
					applied_result_names.add(result_doc.name)
			elif row.action == "Cancel Unstarted" and row.existing_work_order:
				_drop_unfrozen_scheduling_rows_for_work_order(
					work_order_name=row.existing_work_order,
					planning_run=batch.planning_run,
					source_doctype="APS Work Order Proposal Batch",
					source_name=batch.name,
				)
				work_order_name = _cancel_unstarted_work_order(row.existing_work_order)
				row.target_work_order = work_order_name
				row.review_status = "Applied"
				row.review_note = _("Unstarted Work Order {0} cancelled and unreleased scheduling rows removed.").format(work_order_name)
				applied_work_orders.append(work_order_name)
			elif row.action == "Close Residual" and row.existing_work_order:
				_drop_unfrozen_scheduling_rows_for_work_order(
					work_order_name=row.existing_work_order,
					planning_run=batch.planning_run,
					source_doctype="APS Work Order Proposal Batch",
					source_name=batch.name,
				)
				work_order_name = _close_residual_work_order(
					work_order_name=row.existing_work_order,
					run_name=batch.planning_run,
					result_name=result_doc.name if result_doc else row.result_reference,
					proposal_batch=batch.name,
					required_delivery_date=row.required_delivery_date,
				)
				row.target_work_order = work_order_name
				row.review_status = "Applied"
				row.review_note = _("Residual quantity on Work Order {0} was closed.").format(work_order_name)
				applied_work_orders.append(work_order_name)
			else:
				row.review_status = "Skipped"
				row.review_note = _("Action {0} could not be reconciled automatically.").format(row.action)
				skipped_rows.append(row.result_reference)
				if result_doc:
					_ensure_open_exception(
						planning_run=batch.planning_run,
						severity="Warning",
						exception_type="Manual Work Order Review",
						message=_("APS skipped automatic action {0} for result {1}.").format(row.action, result_doc.name),
						item_code=result_doc.item_code,
						customer=result_doc.customer,
						source_doctype="APS Work Order Proposal Batch",
						source_name=batch.name,
						resolution_hint=_("Review work order reconciliation manually before formal execution changes."),
						is_blocking=0,
					)
			continue
		except Exception as exc:
			row.review_status = "Skipped"
			row.review_note = str(exc)
			skipped_rows.append(row.result_reference)

	batch.approved_by = frappe.session.user
	batch.approved_on = now_datetime()
	batch.save(ignore_permissions=True)

	if applied_work_orders and applied_result_names:
		_set_run_result_segment_status(
			run_name=batch.planning_run,
			run_status="Work Order Proposed",
			result_status="Work Order Proposed",
			segment_status="Work Order Proposed",
			flow_step="Formal Work Orders Ready",
			next_step_hint="Generate shift schedule proposals",
			result_names=sorted(applied_result_names),
		)
	elif applied_work_orders:
		frappe.db.set_value("APS Planning Run", batch.planning_run, "status", "Work Order Proposed")
	return {
		"run": batch.planning_run,
		"work_order_proposal_batch": batch.name,
		"applied_work_orders": sorted(set(applied_work_orders)),
		"skipped_rows": skipped_rows,
	}


def _append_review_note(existing_notes: str | None, line: str) -> str:
	notes = (existing_notes or "").strip()
	entry = (line or "").strip()
	if not entry:
		return notes
	if not notes:
		return entry
	return f"{notes}\n{entry}"


def reject_work_order_proposals(batch_name: str, reason: str) -> dict[str, Any]:
	batch = frappe.get_doc("APS Work Order Proposal Batch", batch_name)
	reason_text = (reason or "").strip()
	if not reason_text:
		frappe.throw(_("Please enter a rejection reason."))

	rejected_rows = 0
	for row in batch.items:
		if row.review_status in ("Applied", "Skipped", "Rejected"):
			continue
		row.review_status = "Rejected"
		row.review_note = reason_text
		rejected_rows += 1

	if not rejected_rows:
		frappe.throw(_("No pending or approved work-order proposal rows are available to reject."))

	timestamp = now_datetime().strftime("%Y-%m-%d %H:%M:%S")
	batch.notes = _append_review_note(
		batch.notes,
		_("[{0}] {1} rejected remaining work-order proposal rows: {2}").format(
			timestamp,
			frappe.session.user,
			reason_text,
		),
	)
	batch.approved_by = frappe.session.user
	batch.approved_on = now_datetime()
	batch.save(ignore_permissions=True)
	return {
		"run": batch.planning_run,
		"work_order_proposal_batch": batch.name,
		"rejected_rows": rejected_rows,
	}


def generate_shift_schedule_proposals(
	run_name: str | None = None,
	work_order_proposal_batch: str | None = None,
	release_horizon_days: int | None = None,
) -> dict[str, Any]:
	if not work_order_proposal_batch:
		work_order_proposal_batch = frappe.db.get_value(
			"APS Work Order Proposal Batch",
			{"planning_run": run_name, "status": "Applied"},
			"name",
			order_by="modified desc",
		)
	if not work_order_proposal_batch:
		frappe.throw(_("Apply a work order proposal batch before generating shift schedule proposals."))
	wo_batch = frappe.get_doc("APS Work Order Proposal Batch", work_order_proposal_batch)
	run_doc = frappe.get_doc("APS Planning Run", wo_batch.planning_run)
	release_horizon_days = cint(release_horizon_days or get_settings_dict()["release_horizon_days"] or 3)
	release_to = getdate(add_days(today(), release_horizon_days))

	items = []
	for row in wo_batch.items:
		if row.review_status != "Applied":
			continue
		work_order_name = row.target_work_order or row.existing_work_order
		if not work_order_name:
			continue
		current_segments = []
		if row.result_reference and frappe.db.exists("APS Schedule Result", row.result_reference):
			for segment in _get_primary_segments_for_result(row.result_reference):
				if getdate(segment.get("start_time")) > release_to:
					continue
				current_segments.append(segment)

		matched_existing_rows = set()
		for segment in current_segments:
			existing_row = _find_matching_scheduling_row(
				work_order_name=work_order_name,
				segment=segment,
				matched_row_names=matched_existing_rows,
				release_to=release_to,
			)
			if existing_row:
				matched_existing_rows.add(existing_row.get("name"))
			action = _classify_shift_schedule_action(existing_row, segment)
			items.append(
				{
					"result_reference": row.result_reference,
					"segment_reference": segment.get("name"),
					"action": action,
					"item_code": row.item_code,
					"work_order": work_order_name,
					"plant_floor": segment.get("plant_floor"),
					"posting_date": getdate(segment.get("start_time")),
					"shift_type": _determine_shift_type(segment.get("start_time")),
					"workstation": segment.get("workstation"),
					"planned_start_time": segment.get("start_time"),
					"planned_end_time": segment.get("end_time"),
					"planned_qty": segment.get("planned_qty"),
					"existing_scheduling": existing_row.get("work_order_scheduling") if existing_row else None,
					"existing_scheduling_item": existing_row.get("name") if existing_row else None,
					"review_status": "Pending",
					"review_note": _("Formal scheduling will be reconciled in place for this segment."),
				}
			)

		for existing_row in _get_formal_scheduling_reconciliation_rows(work_order_name, release_to=release_to):
			if existing_row.get("name") in matched_existing_rows or existing_row.get("is_frozen"):
				continue
			if row.action in ("Cancel Unstarted", "Close Residual") or existing_row.get("custom_aps_segment_reference"):
				items.append(
					{
						"result_reference": row.result_reference or existing_row.get("custom_aps_result_reference"),
						"segment_reference": existing_row.get("custom_aps_segment_reference") or f"cancel::{existing_row.get('name')}",
						"action": "Cancel Existing",
						"item_code": row.item_code,
						"work_order": work_order_name,
						"plant_floor": existing_row.get("plant_floor"),
						"posting_date": existing_row.get("posting_date"),
						"shift_type": existing_row.get("shift_type"),
						"workstation": existing_row.get("workstation"),
						"planned_start_time": existing_row.get("planned_start_date"),
						"planned_end_time": existing_row.get("planned_end_date"),
						"planned_qty": existing_row.get("scheduling_qty"),
						"existing_scheduling": existing_row.get("work_order_scheduling"),
						"existing_scheduling_item": existing_row.get("name"),
						"review_status": "Pending",
						"review_note": _("Unexecuted formal scheduling row will be cancelled or removed."),
					}
				)

	batch = frappe.get_doc(
		{
			"doctype": "APS Shift Schedule Proposal Batch",
			"planning_run": run_doc.name,
			"company": run_doc.company,
			"plant_floor": run_doc.plant_floor,
			"work_order_proposal_batch": wo_batch.name,
			"proposal_date": today(),
			"status": "Ready For Review",
			"approval_state": "Pending",
			"proposal_count": len(items),
			"notes": _("Generated from APS work order proposal batch {0}. Review against existing day/night shift scheduling before formal reconciliation.").format(wo_batch.name),
			"items": items,
		}
	).insert(ignore_permissions=True)
	proposed_result_names = sorted({row.get("result_reference") for row in items if row.get("result_reference")})
	proposed_segment_names = sorted({row.get("segment_reference") for row in items if row.get("segment_reference")})
	if proposed_result_names or proposed_segment_names:
		_set_run_result_segment_status(
			run_name=run_doc.name,
			run_status="Shift Proposed",
			result_status="Shift Proposed",
			segment_status="Shift Proposed",
			flow_step="Shift Schedule Proposal Review",
			next_step_hint="Review and apply day/night shift proposals",
			result_names=proposed_result_names,
			segment_names=proposed_segment_names,
		)
	else:
		frappe.db.set_value("APS Planning Run", run_doc.name, "status", "Shift Proposed")
	return {
		"run": run_doc.name,
		"shift_schedule_proposal_batch": batch.name,
		"proposal_count": len(items),
		"release_to": release_to,
	}


def apply_shift_schedule_proposals(batch_name: str) -> dict[str, Any]:
	batch = frappe.get_doc("APS Shift Schedule Proposal Batch", batch_name)
	overlap_summary = _validate_run_segment_overlaps(batch.planning_run, persist_exceptions=True)
	mold_overlap_summary = _validate_run_mold_overlaps(batch.planning_run, persist_exceptions=True)
	mold_gate = validate_run_mold_readiness(batch.planning_run, persist_exceptions=True)
	if overlap_summary["count"] or mold_overlap_summary["count"] or mold_gate["blocking_count"]:
		frappe.throw(_("Resolve overlap and mold blockers before applying shift schedule proposals."))
	approved_rows = [row for row in batch.items if row.review_status == "Approved"]
	if not approved_rows:
		frappe.throw(_("No shift proposal rows are marked Approved. Review the batch before formal scheduling."))

	grouped = defaultdict(list)
	for row in approved_rows:
		grouped[(str(row.posting_date), row.shift_type or "", row.workstation or "", row.work_order or "")].append(row)

	scheduling_docs = set()
	applied_rows = 0
	applied_result_names = set()
	applied_segment_names = set()
	for row in approved_rows:
		try:
			scheduling = _upsert_formal_shift_scheduling(batch, row)
			row.target_scheduling = scheduling.get("docname")
			row.review_status = "Applied"
			row.review_note = scheduling.get("message") or _("Applied to formal Work Order Scheduling {0}.").format(scheduling.get("docname"))
			applied_rows += 1
			if row.result_reference:
				applied_result_names.add(row.result_reference)
			if row.segment_reference:
				applied_segment_names.add(row.segment_reference)
			if scheduling.get("docname"):
				scheduling_docs.add(scheduling["docname"])
		except Exception as exc:
			row.review_status = "Skipped"
			row.review_note = str(exc)

	batch.approved_by = frappe.session.user
	batch.approved_on = now_datetime()
	batch.save(ignore_permissions=True)

	release_batch = frappe.get_doc(
		{
			"doctype": "APS Release Batch",
			"planning_run": batch.planning_run,
			"company": batch.company,
			"release_from_date": today(),
			"release_to_date": max((getdate(row.posting_date) for row in batch.items), default=getdate(today())),
			"status": "Released" if applied_rows else "Draft",
			"generated_work_orders": len(
				{
					row.work_order
					for row in batch.items
					if row.review_status == "Applied" and row.work_order
				}
			),
			"work_order_scheduling": sorted(scheduling_docs)[0] if len(scheduling_docs) == 1 else None,
		}
	).insert(ignore_permissions=True)

	if applied_rows and (applied_result_names or applied_segment_names):
		_set_run_result_segment_status(
			run_name=batch.planning_run,
			run_status="Applied",
			result_status="Applied",
			segment_status="Applied",
			flow_step="Formal Schedule Applied",
			next_step_hint="Monitor execution drift and exceptions",
			result_names=sorted(applied_result_names),
			segment_names=sorted(applied_segment_names),
		)
	elif applied_rows:
		frappe.db.set_value("APS Planning Run", batch.planning_run, "status", "Applied")
	return {
		"run": batch.planning_run,
		"shift_schedule_proposal_batch": batch.name,
		"release_batch": release_batch.name,
		"work_order_schedulings": sorted(scheduling_docs),
		"applied_rows": applied_rows,
	}


def reject_shift_schedule_proposals(batch_name: str, reason: str) -> dict[str, Any]:
	batch = frappe.get_doc("APS Shift Schedule Proposal Batch", batch_name)
	reason_text = (reason or "").strip()
	if not reason_text:
		frappe.throw(_("Please enter a rejection reason."))

	rejected_rows = 0
	for row in batch.items:
		if row.review_status in ("Applied", "Skipped", "Rejected"):
			continue
		row.review_status = "Rejected"
		row.review_note = reason_text
		rejected_rows += 1

	if not rejected_rows:
		frappe.throw(_("No pending or approved day/night shift proposal rows are available to reject."))

	timestamp = now_datetime().strftime("%Y-%m-%d %H:%M:%S")
	batch.notes = _append_review_note(
		batch.notes,
		_("[{0}] {1} rejected remaining day/night shift proposal rows: {2}").format(
			timestamp,
			frappe.session.user,
			reason_text,
		),
	)
	batch.approved_by = frappe.session.user
	batch.approved_on = now_datetime()
	batch.save(ignore_permissions=True)
	return {
		"run": batch.planning_run,
		"shift_schedule_proposal_batch": batch.name,
		"rejected_rows": rejected_rows,
	}


def update_schedule_notes(
	result_name: str | None = None,
	segment_name: str | None = None,
	result_note: str | None = None,
	segment_note: str | None = None,
) -> dict[str, Any]:
	if not result_name and not segment_name:
		frappe.throw(_("Provide result_name or segment_name to update notes."))
	if segment_name and not result_name:
		result_name = frappe.db.get_value("APS Schedule Segment", segment_name, "parent")
	result_doc = frappe.get_doc("APS Schedule Result", result_name)
	if result_note is not None:
		result_doc.notes = result_note
	updated_segment_note = None
	if segment_name and segment_note is not None:
		for row in result_doc.segments:
			if row.name == segment_name:
				row.segment_note = segment_note
				updated_segment_note = row.segment_note
				break
	result_doc.save(ignore_permissions=True)
	return {
		"result_name": result_doc.name,
		"result_note": result_doc.notes,
		"segment_name": segment_name,
		"segment_note": updated_segment_note,
		"modified_by": result_doc.modified_by,
		"modified": result_doc.modified,
	}


def sync_execution_feedback_to_aps(run_name: str) -> dict[str, Any]:
	run_doc = frappe.get_doc("APS Planning Run", run_name)
	now_value = now_datetime()
	updated_segments = 0
	status_counts = defaultdict(int)
	for result_name in frappe.get_all("APS Schedule Result", filters={"planning_run": run_name}, pluck="name"):
		result_doc = frappe.get_doc("APS Schedule Result", result_name)
		segment_statuses = []
		result_progress = 0.0
		actual_start_times = []
		actual_end_times = []
		for segment in result_doc.segments:
			execution = _get_segment_execution_snapshot(segment)
			segment.actual_status = execution["actual_status"]
			segment.actual_completed_qty = execution["actual_completed_qty"]
			segment.actual_start_time = execution["actual_start_time"]
			segment.actual_end_time = execution["actual_end_time"]
			segment.delay_minutes = execution["delay_minutes"]
			segment.last_execution_sync_on = now_value
			segment.linked_work_order = execution["linked_work_order"]
			segment.linked_work_order_scheduling = execution["linked_work_order_scheduling"]
			segment.linked_scheduling_item = execution["linked_scheduling_item"]
			segment_statuses.append(execution["actual_status"])
			status_counts[execution["actual_status"]] += 1
			result_progress += flt(execution["actual_completed_qty"])
			if execution["actual_start_time"]:
				actual_start_times.append(execution["actual_start_time"])
			if execution["actual_end_time"]:
				actual_end_times.append(execution["actual_end_time"])
			updated_segments += 1
		result_doc.actual_status = _rollup_result_actual_status(segment_statuses)
		result_doc.actual_progress_qty = result_progress
		result_doc.actual_start_time = min(actual_start_times) if actual_start_times else None
		result_doc.actual_end_time = max(actual_end_times) if actual_end_times else None
		result_doc.delay_minutes = _rollup_delay_minutes(result_doc.segments)
		result_doc.last_execution_sync_on = now_value
		_sync_execution_exceptions(run_name, result_doc)
		result_doc.save(ignore_permissions=True)
	run_doc.db_set("exception_count", frappe.db.count("APS Exception Log", {"planning_run": run_name, "status": "Open"}))
	return {"run": run_name, "updated_segments": updated_segments, "status_counts": dict(status_counts)}


def get_execution_health_for_run(run_name: str, sync: bool = False) -> dict[str, Any]:
	if sync:
		sync_execution_feedback_to_aps(run_name)
	rows = frappe.get_all(
		"APS Schedule Result",
		filters={"planning_run": run_name},
		fields=["actual_status", "actual_progress_qty"],
	)
	status_counts = defaultdict(int)
	for row in rows:
		status_counts[row.actual_status or "Not Started"] += 1
	today_entries = _count_today_manufacture_entries(run_name)
	return {
		"run": run_name,
		"status_counts": dict(status_counts),
		"running_segments": status_counts.get("Running", 0),
		"delayed_segments": status_counts.get("Delayed", 0) + status_counts.get("Slow Progress", 0),
		"no_recent_update_segments": status_counts.get("No Recent Update", 0),
		"today_completed_entries": today_entries,
	}


def _with_role_filtered_actions(context: dict[str, Any]) -> dict[str, Any]:
	actions = []
	user_roles = set(frappe.get_roles())
	is_admin = frappe.session.user == "Administrator"
	for action in context.get("actions") or []:
		required_roles = ACTION_REQUIRED_ROLES.get(action.get("action_key"))
		if required_roles and not is_admin and not user_roles.intersection(required_roles):
			action = dict(action)
			action["enabled"] = 0
			action["disabled_reason"] = _("Your role can view this APS step but cannot run this action.")
		actions.append(action)
	context["actions"] = actions
	return context


def analyze_change_request_impact(change_request: str) -> dict[str, Any]:
	doc = frappe.get_doc("APS Change Request", change_request)
	impact = analyze_insert_order_impact(
		company=doc.company,
		plant_floor=doc.plant_floor,
		plant_floors=getattr(doc, "selected_plant_floors", None),
		item_code=doc.item_code,
		qty=doc.qty,
		required_date=doc.required_date,
		customer=doc.customer,
	)
	doc.status = "Analyzed"
	doc.impact_summary = json.dumps(
		{
			"scheduled_qty": impact.get("scheduled_qty"),
			"unscheduled_qty": impact.get("unscheduled_qty"),
			"displaced_segments": len(impact.get("displaced_segments") or []),
		},
		ensure_ascii=True,
	)
	doc.save(ignore_permissions=True)
	return impact


def apply_change_request(change_request: str) -> dict[str, Any]:
	doc = frappe.get_doc("APS Change Request", change_request)
	doc.approval_state = "Approved"
	doc.status = "Approved"
	doc.save(ignore_permissions=True)
	return {"change_request": doc.name, "status": doc.status}


def analyze_insert_order_impact(
	company: str,
	plant_floor: str | None = None,
	plant_floors: list[str] | str | None = None,
	item_code: str | None = None,
	qty: float | None = None,
	required_date: str | None = None,
	customer: str | None = None,
) -> dict[str, Any]:
	settings = get_settings_dict()
	item_code = _require_item_name(item_code)
	if not required_date:
		frappe.throw(_("Required Date is required for insert-order impact analysis."))
	qty = flt(qty)
	selected_plant_floors = _normalize_selected_plant_floors(
		company=company,
		plant_floors=plant_floors,
		plant_floor=plant_floor,
		required=True,
	)
	item_context = _get_item_context(item_code, settings)
	available_molds = _get_available_mold_rows(item_code)
	capability_rows = _get_machine_capability_rows(plant_floors=selected_plant_floors)
	workstation_state = _build_workstation_state_map(capability_rows)
	locked_segments = _get_locked_segments(selected_plant_floors)
	mold_state = _build_mold_state_map(locked_segments)
	_apply_locked_segments_to_state(workstation_state, locked_segments)
	candidates = _select_machine_candidates(
		item_code=item_code,
		item_context=item_context,
		capability_rows=capability_rows,
		plant_floors=selected_plant_floors,
	)
	best = _choose_best_slot(
		item_code=item_code,
		item_context=item_context,
		qty=qty,
		demand_date=required_date,
		horizon_start=get_datetime(now_datetime()),
		horizon_end=get_datetime(add_days(required_date, 7)),
		workstation_state=workstation_state,
		mold_state=mold_state,
		candidates=candidates,
		settings=settings,
	)
	impacted = []
	impacted_customers = set()
	changeover_minutes = sum(flt(segment.get("changeover_minutes")) for segment in best.get("segments") or [])
	if best["segments"]:
		first_segment = best["segments"][0]
		overlap_segments = frappe.get_all(
			"APS Schedule Segment",
			filters={
				"workstation": first_segment.get("workstation"),
				"start_time": ("<", first_segment.get("end_time")),
				"end_time": (">", first_segment.get("start_time")),
			},
			fields=["name", "parent", "workstation", "start_time", "end_time", "planned_qty"],
			order_by="start_time asc",
		)
		result_meta = {
			row.name: row
			for row in frappe.get_all(
				"APS Schedule Result",
				filters={"name": ("in", [row.parent for row in overlap_segments])} if overlap_segments else {"name": "__missing__"},
				fields=["name", "customer", "item_code", "requested_date"],
			)
		}
		for row in overlap_segments:
			parent = result_meta.get(row.parent)
			if parent and parent.customer:
				impacted_customers.add(parent.customer)
			impacted.append(
				{
					"workstation": row.get("workstation"),
					"segment_name": row.get("name"),
					"result_name": row.get("parent"),
					"item_code": parent.item_code if parent else None,
					"customer": parent.customer if parent else None,
					"requested_date": parent.requested_date if parent else None,
					"start_time": row.get("start_time"),
					"end_time": row.get("end_time"),
					"planned_qty": row.get("planned_qty"),
				}
			)

	return {
		"item_code": item_code,
		"customer": customer,
		"required_date": required_date,
		"selected_plant_floors": selected_plant_floors,
		"scheduled_qty": best["scheduled_qty"],
		"unscheduled_qty": best["unscheduled_qty"],
		"candidate_workstations": [row.get("workstation") for row in candidates],
		"candidate_molds": [
			{
				"mould_reference": row.get("mold"),
				"mold_name": row.get("mold_name"),
				"is_family_mold": cint(row.get("is_family_mold")),
				"machine_tonnage": row.get("machine_tonnage"),
				"cavity_count": row.get("cavity_count"),
				"cycle_time_seconds": row.get("cycle_time_seconds"),
				"output_qty": row.get("output_qty"),
				"cavity_output_qty": row.get("cavity_output_qty"),
				"effective_output_qty": row.get("effective_output_qty"),
			}
			for row in available_molds
		],
		"parallelization_plan": [
			{
				"workstation": row.get("workstation"),
				"mould_reference": row.get("mould_reference"),
				"planned_qty": row.get("planned_qty"),
				"start_time": row.get("start_time"),
				"end_time": row.get("end_time"),
				"lane_key": row.get("lane_key"),
			}
			for row in (best.get("segments") or [])
			if row.get("segment_kind") != "Family Co-Product"
		],
		"family_side_outputs": best.get("family_side_outputs") or [],
		"impacted_segments": impacted,
		"displaced_segments": impacted,
		"impacted_customers": sorted(impacted_customers),
		"changeover_minutes": changeover_minutes,
		"future_batch_hint": _get_future_demand_hint(item_code=item_code, demand_date=required_date),
		"missing_machine": any(error.get("exception_type") == "Machine Unavailable" for error in best["exceptions"]),
		"missing_mould": 0 if available_molds else 1,
		"schedule_explanation": best.get("schedule_explanation"),
		"family_output_summary": best.get("family_output_summary"),
		"exceptions": best["exceptions"],
	}


def rebuild_exceptions(run_name: str) -> dict[str, Any]:
	for name in frappe.get_all("APS Exception Log", filters={"planning_run": run_name}, pluck="name"):
		frappe.delete_doc("APS Exception Log", name, force=1, ignore_permissions=True)

	recreated = []
	for row in frappe.get_all(
		"APS Schedule Result",
		filters={"planning_run": run_name, "risk_status": ("in", ["Attention", "Critical", "Blocked"])},
		fields=["name", "customer", "item_code", "status", "risk_status", "unscheduled_qty"],
	):
		severity = "Critical" if row.risk_status in ("Critical", "Blocked") else "Warning"
		doc = _create_exception(
			planning_run=run_name,
			severity=severity,
			exception_type="Scheduling Risk",
			message=_("Result {0} has risk state {1} and unscheduled qty {2}.").format(
				row.name, row.risk_status, row.unscheduled_qty
			),
			item_code=row.item_code,
			customer=row.customer,
			source_doctype="APS Schedule Result",
			source_name=row.name,
			resolution_hint=_("Review the planned sequence, capacity and frozen segments."),
			is_blocking=1 if row.risk_status == "Blocked" else 0,
		)
		recreated.append(doc.name)
	return {"run": run_name, "exceptions": recreated}


def get_next_actions_for_context(doctype: str, docname: str) -> dict[str, Any]:
	if doctype == "APS Schedule Import Batch":
		doc = frappe.get_doc(doctype, docname)
		schedule_name = frappe.db.get_value("Customer Delivery Schedule", {"import_batch": doc.name}, "name")
		current_step = "Imported" if doc.status == "Imported" else doc.status or "Draft"
		next_step = "Rebuild Demand Pool / Net Requirement" if doc.status == "Imported" else "Complete Import"
		return _with_role_filtered_actions({
			"doctype": doctype,
			"docname": docname,
			"current_step": current_step,
			"next_step": next_step,
			"blocking_reason": "" if doc.status == "Imported" else "Import has not been completed yet.",
			"actions": [
				{
					"label": "Import and Rebuild",
					"action_key": "promote_import",
					"method": "injection_aps.api.app.promote_schedule_import_to_net_requirement",
					"kwargs": {"import_batch": doc.name},
					"enabled": 1 if doc.status == "Imported" else 0,
				},
				{
					"label": "Open Schedule",
					"action_key": "open_schedule",
					"route": f"Form/Customer Delivery Schedule/{schedule_name}" if schedule_name else "",
					"enabled": 1 if schedule_name else 0,
				},
				{
					"label": "Net Requirements",
					"action_key": "open_net_requirement",
					"route": "aps-net-requirement-workbench",
					"enabled": 1,
				},
			],
		})

	if doctype == "Customer Delivery Schedule":
		doc = frappe.get_doc(doctype, docname)
		return _with_role_filtered_actions({
			"doctype": doctype,
			"docname": docname,
			"current_step": doc.status or "Draft",
			"next_step": "Rebuild Demand and Recalculate" if doc.status == "Active" else "Activate the Current Version First",
			"blocking_reason": "" if doc.status == "Active" else "Only active schedules can drive APS.",
			"actions": [
				{
					"label": "Rebuild Demand",
					"action_key": "rebuild_demand_pool",
					"method": "injection_aps.api.app.rebuild_demand_pool",
					"kwargs": {"company": doc.company},
					"enabled": 1 if doc.status == "Active" else 0,
				},
				{
					"label": "Net Workbench",
					"action_key": "open_net_requirement",
					"route": "aps-net-requirement-workbench",
					"enabled": 1,
				},
				{
					"label": "Version Diff",
					"action_key": "open_schedule",
					"route": f"Form/Customer Delivery Schedule/{doc.name}",
					"enabled": 1,
				},
			],
		})

	if doctype == "APS Planning Run":
		doc = frappe.get_doc(doctype, docname)
		has_applied_wo_batch = bool(
			frappe.db.exists(
				"APS Work Order Proposal Batch",
				{"planning_run": doc.name, "status": "Applied"},
			)
		)
		context = _build_planning_run_context(doc)
		return _with_role_filtered_actions({
			**context,
			"actions": [
				{
					"label": "Recalculate",
					"action_key": "run_trial",
					"method": "injection_aps.api.app.run_planning_run",
					"kwargs": {"run_name": doc.name},
					"enabled": 1,
					"confirm_required": 1,
					"confirm_title": "Confirm Recalculate",
					"confirm_summary": [
						"APS Run: {0}".format(doc.name),
						"Company: {0}".format(doc.company or "-"),
						"Plant Floors: {0}".format(context.get("selected_plant_floor_summary") or "-"),
						"Horizon: {0} days".format(cint(doc.horizon_days or 0)),
						"This action will recalculate APS results from the current demand.",
					],
				},
				{
					"label": "Confirm Run",
					"action_key": "approve",
					"method": "injection_aps.api.app.approve_planning_run",
					"kwargs": {"run_name": doc.name},
					"enabled": 1 if doc.approval_state != "Approved" and doc.status in ("Planned", "Risk", "Draft") else 0,
					"confirm_required": 1,
					"confirm_title": "Confirm APS Run",
					"confirm_summary": [
						"APS Run: {0}".format(doc.name),
						"Current Status: {0}".format(_label_run_status(doc.status)),
						"Exceptions: {0}".format(cint(doc.exception_count or 0)),
						"After confirmation, the run will move into the downstream proposal review flow.",
					],
				},
				{
					"label": "WO Proposals",
					"action_key": "generate_work_order_proposals",
					"method": "injection_aps.api.app.generate_work_order_proposals",
					"kwargs": {"run_name": doc.name},
					"enabled": 1 if doc.approval_state == "Approved" and doc.status in ("Approved",) else 0,
					"confirm_required": 1,
					"confirm_title": "Confirm Generate Work Order Proposals",
					"confirm_summary": [
						"APS Run: {0}".format(doc.name),
						"This action will generate a work-order proposal batch for review.",
					],
				},
				{
					"label": "Shift Proposals",
					"action_key": "generate_shift_schedule_proposals",
					"method": "injection_aps.api.app.generate_shift_schedule_proposals",
					"kwargs": {"run_name": doc.name},
					"enabled": 1 if has_applied_wo_batch or doc.status == "Shift Proposed" else 0,
					"confirm_required": 1,
					"confirm_title": "Confirm Generate Day/Night Shift Proposals",
					"confirm_summary": [
						"APS Run: {0}".format(doc.name),
						"This action will generate a day/night shift proposal batch for review.",
					],
				},
				{
					"label": "Board",
					"action_key": "open_gantt",
					"route": f"aps-schedule-gantt?run_name={doc.name}",
					"enabled": 1,
				},
				{
					"label": "Execution",
					"action_key": "open_release_center",
					"route": f"aps-release-center?run_name={doc.name}",
					"enabled": 1,
				},
			],
		})

	if doctype == "APS Work Order Proposal Batch":
		doc = frappe.get_doc(doctype, docname)
		return _with_role_filtered_actions({
			"doctype": doctype,
			"docname": docname,
			"current_step": doc.status or "Draft",
			"next_step": "Review proposal rows and apply formal work orders",
			"blocking_reason": "" if doc.status in ("Ready For Review", "Reviewed", "Applied") else "Generate proposal rows first.",
			"actions": [
				{
					"label": "Run",
					"action_key": "open_run",
					"route": f"Form/APS Planning Run/{doc.planning_run}" if doc.planning_run else "",
					"enabled": 1 if doc.planning_run else 0,
				},
				{
					"label": "Apply Results",
					"action_key": "apply_work_order_proposals",
					"method": "injection_aps.api.app.apply_work_order_proposals",
					"kwargs": {"batch_name": doc.name},
					"enabled": 1 if doc.status in ("Ready For Review", "Partially Reviewed", "Reviewed") else 0,
					"confirm_required": 1,
					"confirm_title": "Confirm Apply Work Order Results",
					"confirm_summary": [
						"Work Order Proposal Batch: {0}".format(doc.name),
						"Approved rows will formally create or bind work orders.",
					],
				},
				{
					"label": "Execution",
					"action_key": "open_release_center",
					"route": f"aps-release-center?run_name={doc.planning_run}" if doc.planning_run else "aps-release-center",
					"enabled": 1,
				},
			],
		})

	if doctype == "APS Shift Schedule Proposal Batch":
		doc = frappe.get_doc(doctype, docname)
		return _with_role_filtered_actions({
			"doctype": doctype,
			"docname": docname,
			"current_step": doc.status or "Draft",
			"next_step": "Review day/night shift rows and apply formal scheduling",
			"blocking_reason": "" if doc.status in ("Ready For Review", "Reviewed", "Applied") else "Generate shift proposal rows first.",
			"actions": [
				{
					"label": "Run",
					"action_key": "open_run",
					"route": f"Form/APS Planning Run/{doc.planning_run}" if doc.planning_run else "",
					"enabled": 1 if doc.planning_run else 0,
				},
				{
					"label": "Apply Results",
					"action_key": "apply_shift_schedule_proposals",
					"method": "injection_aps.api.app.apply_shift_schedule_proposals",
					"kwargs": {"batch_name": doc.name},
					"enabled": 1 if doc.status in ("Ready For Review", "Partially Reviewed", "Reviewed") else 0,
					"confirm_required": 1,
					"confirm_title": "Confirm Apply Day/Night Shift Results",
					"confirm_summary": [
						"Shift Proposal Batch: {0}".format(doc.name),
						"Approved rows will formally write day/night scheduling rows.",
					],
				},
				{
					"label": "Execution",
					"action_key": "open_release_center",
					"route": f"aps-release-center?run_name={doc.planning_run}" if doc.planning_run else "aps-release-center",
					"enabled": 1,
				},
			],
		})

	if doctype == "APS Release Batch":
		doc = frappe.get_doc(doctype, docname)
		return _with_role_filtered_actions({
			"doctype": doctype,
			"docname": docname,
			"current_step": doc.status or "Draft",
			"next_step": "Monitor execution feedback" if doc.status == "Released" else "Apply formal documents",
			"blocking_reason": "" if doc.status == "Released" else "Formal work orders / shift schedulings have not been applied yet.",
			"actions": [
				{
					"label": "Run",
					"action_key": "open_run",
					"route": f"Form/APS Planning Run/{doc.planning_run}" if doc.planning_run else "",
					"enabled": 1 if doc.planning_run else 0,
				},
				{
					"label": "Execution",
					"action_key": "open_release_center",
					"route": f"aps-release-center?run_name={doc.planning_run}" if doc.planning_run else "aps-release-center",
					"enabled": 1,
				},
				{
					"label": "Open Execution Scheduling",
					"action_key": "open_scheduling",
					"route": f"Form/Work Order Scheduling/{doc.work_order_scheduling}" if doc.work_order_scheduling else "",
					"enabled": 1 if doc.work_order_scheduling else 0,
				},
			],
		})

	frappe.throw(_("Next-action context is not supported for {0}.").format(doctype))


def promote_schedule_import_to_net_requirement(
	import_batch: str | None = None,
	schedule: str | None = None,
	company: str | None = None,
) -> dict[str, Any]:
	if import_batch:
		doc = frappe.get_doc("APS Schedule Import Batch", import_batch)
		company = company or doc.company
	if schedule:
		doc = frappe.get_doc("Customer Delivery Schedule", schedule)
		company = company or doc.company
	demand = rebuild_demand_pool(company=company)
	net = rebuild_net_requirements(company=company)
	return {
		"company": company,
		"demand_pool": demand,
		"net_requirement": net,
		"next_route": "aps-net-requirement-workbench",
	}


def create_trial_run_from_net_requirement_context(
	company: str | None = None,
	plant_floor: str | None = None,
	plant_floors: list[str] | str | None = None,
	item_code: str | None = None,
	customer: str | None = None,
	horizon_days: int | None = None,
) -> dict[str, Any]:
	return run_planning_run(
		company=company,
		plant_floor=plant_floor,
		plant_floors=plant_floors,
		horizon_days=horizon_days,
		item_code=item_code,
		customer=customer,
		run_type="Trial",
	)


def preview_manual_schedule_adjustment(
	segment_name: str,
	target_workstation: str | None = None,
	before_segment_name: str | None = None,
	target_start_time=None,
	target_end_time=None,
	allow_locked: int = 0,
	allow_risk_override: int = 0,
) -> dict[str, Any]:
	segment_rows = frappe.get_all(
		"APS Schedule Segment",
		filters={"name": segment_name},
		fields=[
			"name",
			"parent",
			"parenttype",
			"workstation",
			"plant_floor",
			"start_time",
			"end_time",
			"planned_qty",
			"segment_kind",
			"segment_status",
			"is_locked",
			"is_manual",
			"mould_reference",
			"lane_key",
			"parallel_group",
			"family_group",
		],
		limit=1,
	)
	if not segment_rows:
		frappe.throw(_("APS Schedule Segment {0} was not found.").format(segment_name))
	segment = segment_rows[0]
	if segment.segment_kind == "Family Co-Product":
		frappe.throw(_("Family Co-Product segment cannot be adjusted directly. Move the primary segment instead."))
	if (
		cint(segment.is_locked)
		or segment.segment_status in MANUAL_ADJUSTMENT_BLOCKED_SEGMENT_STATUSES
	) and not cint(allow_locked):
		return {
			"allowed": 0,
			"blocking_reasons": [_("Segment {0} is locked or already applied to formal execution.").format(segment_name)],
		}

	result = frappe.get_doc("APS Schedule Result", segment.parent)
	run_doc = frappe.get_doc("APS Planning Run", result.planning_run)
	selected_plant_floors = _get_run_selected_plant_floors(run_doc)
	settings = get_settings_dict()
	item_context = _get_item_context(result.item_code, settings)
	target_workstation = target_workstation or segment.workstation
	candidates = [
		row
		for row in _select_machine_candidates(
			item_code=result.item_code,
			item_context=item_context,
			capability_rows=_get_machine_capability_rows(selected_plant_floors),
			plant_floors=selected_plant_floors,
		)
		if row.get("workstation") == target_workstation
	]
	if not candidates:
		return {
			"allowed": 0,
			"blocking_reasons": _diagnose_target_workstation_failure(
				item_code=result.item_code,
				target_workstation=target_workstation,
				plant_floors=selected_plant_floors,
				item_context=item_context,
			),
		}

	candidate = next((row for row in candidates if row.get("mould_reference") == segment.mould_reference), candidates[0])
	previous_rows = frappe.get_all(
		"APS Schedule Segment",
		filters={
			"workstation": target_workstation,
			"name": ("!=", segment_name),
			"parenttype": "APS Schedule Result",
			"parent": ("in", frappe.get_all("APS Schedule Result", filters={"planning_run": run_doc.name}, pluck="name") or [""]),
			"segment_kind": ("!=", "Family Co-Product"),
		},
		fields=["name", "workstation", "start_time", "end_time", "color_code", "material_code", "mould_reference"],
		order_by="end_time asc",
	)
	before_segment = None
	if before_segment_name:
		before_rows = frappe.get_all(
			"APS Schedule Segment",
			filters={"name": before_segment_name},
			fields=["name", "start_time", "workstation"],
			limit=1,
		)
		before_segment = before_rows[0] if before_rows else None
	requested_start_time = get_datetime(target_start_time) if target_start_time else None

	base_floor_time = get_datetime("2000-01-01 00:00:00") if target_start_time else get_datetime(now_datetime())
	state = {
		"next_available": base_floor_time,
		"last_color_code": "",
		"last_material_code": "",
		"last_mould_reference": "",
	}
	for row in previous_rows:
		row_end_time = get_datetime(row.end_time)
		if before_segment and row_end_time > get_datetime(before_segment.get("start_time")):
			continue
		if requested_start_time and row_end_time > requested_start_time:
			continue
		if row_end_time >= state["next_available"]:
			state["next_available"] = row_end_time
			state["last_color_code"] = row.get("color_code") or ""
			state["last_material_code"] = row.get("material_code") or ""
			state["last_mould_reference"] = row.get("mould_reference") or ""

	setup_minutes, setup_exceptions, blocked = _estimate_setup_penalty(
		candidate=candidate,
		state=state,
		item_context=item_context,
		settings=settings,
	)
	mold_rows = frappe.db.sql(
		"""
		select
			seg.name,
			seg.workstation,
			seg.start_time,
			seg.end_time,
			res.item_code
		from `tabAPS Schedule Segment` seg
		inner join `tabAPS Schedule Result` res on res.name = seg.parent
		where res.planning_run = %s
			and seg.parenttype = 'APS Schedule Result'
			and ifnull(seg.segment_kind, '') != 'Family Co-Product'
			and ifnull(seg.mould_reference, '') = %s
			and seg.name != %s
		order by seg.start_time asc
		""",
		[run_doc.name, candidate.get("mould_reference"), segment_name],
		as_dict=True,
	)
	mold_next_available = base_floor_time
	for row in mold_rows:
		row_end_time = get_datetime(row.get("end_time"))
		if before_segment and row_end_time > get_datetime(before_segment.get("start_time")):
			continue
		if requested_start_time and row_end_time > requested_start_time:
			continue
		mold_next_available = max(mold_next_available, row_end_time)

	earliest_start_time = max(state["next_available"], mold_next_available) + timedelta(minutes=setup_minutes)
	start_time = earliest_start_time
	if target_start_time:
		start_time = get_datetime(target_start_time)
	hourly_capacity = _estimate_hourly_capacity(candidate=candidate, settings=settings)["hourly_capacity_qty"]
	if target_end_time:
		end_time = get_datetime(target_end_time)
		if end_time <= start_time:
			blocked = True
			setup_exceptions.append(
				{
					"severity": "Critical",
					"exception_type": "Invalid End Time",
					"message": _("Target end time must be later than the computed start time."),
					"workstation": target_workstation,
					"is_blocking": 1,
				}
			)
		duration_hours = max((end_time - start_time).total_seconds() / 3600, 0)
		planned_qty = max(int(duration_hours * max(hourly_capacity, 0)), 0)
	else:
		planned_qty = flt(segment.planned_qty)
		end_time = start_time + timedelta(hours=_estimate_run_hours(planned_qty, candidate, settings))

	blocking_reasons = []
	workstation_overlap_rows = []
	mold_overlap_rows = []
	override_available = 0
	override_reason = ""
	if target_start_time and start_time < earliest_start_time:
		blocked = True
		blocking_reasons.append(
			_("Target start time {0} is earlier than the earliest feasible start {1}.").format(
				frappe.format(start_time, {"fieldtype": "Datetime"}),
				frappe.format(earliest_start_time, {"fieldtype": "Datetime"}),
			)
		)
	if _has_fda_conflict(item_context, candidate):
		override_reason = _("Target workstation {0} violates FDA restriction.").format(target_workstation)
		if cint(allow_risk_override):
			setup_exceptions.append(
				{
					"severity": "Warning",
					"exception_type": "FDA Override",
					"message": _("Manual override accepted FDA risk on workstation {0}.").format(target_workstation),
					"workstation": target_workstation,
					"resolution_hint": _("Confirm contamination controls and approval before execution."),
					"is_blocking": 0,
				}
			)
		else:
			blocked = True
			override_available = 1
			blocking_reasons.append(override_reason)
	if before_segment and end_time > get_datetime(before_segment.get("start_time")):
		blocked = True
		blocking_reasons.append(_("Moved segment would overlap the target sequence anchor {0}.").format(before_segment_name))
	for row in previous_rows:
		if before_segment and row.get("name") == before_segment_name:
			continue
		if get_datetime(row.get("start_time")) < end_time and get_datetime(row.get("end_time")) > start_time:
			blocked = True
			workstation_overlap_rows.append(row)
			blocking_reasons.append(_("Target timing would overlap workstation segment {0}.").format(row.get("name")))
	for row in mold_rows:
		if get_datetime(row.get("start_time")) < end_time and get_datetime(row.get("end_time")) > start_time:
			blocked = True
			mold_overlap_rows.append(row)
			blocking_reasons.append(
				_("Mold {0} would still overlap segment {1} on workstation {2}.").format(
					candidate.get("mould_reference"),
					row.get("name"),
					row.get("workstation") or "-",
				)
			)
	if planned_qty <= 0:
		blocked = True
		blocking_reasons.append(_("Target timing produces zero quantity. Extend the segment window before saving."))
	if blocked:
		for row in setup_exceptions:
			if row.get("is_blocking"):
				blocking_reasons.append(row.get("message"))
	override_available = 1 if override_available and len(blocking_reasons) == 1 else 0
	blocking_summary = ""
	blocking_context_rows = []
	resolution_suggestions = []
	latest_safe_start_time = None
	if mold_overlap_rows:
		first_conflict = sorted(mold_overlap_rows, key=lambda row: get_datetime(row.get("start_time")))[0]
		duration = get_datetime(end_time) - get_datetime(start_time)
		latest_safe_start_time = get_datetime(first_conflict.get("start_time")) - duration
		blocking_summary = (
			"Mold {0} already has continuous scheduling after the selected time. This segment cannot be moved later by itself."
		).format(candidate.get("mould_reference") or "-")
		blocking_context_rows = [
			{"label": "Current Segment", "value": segment_name},
			{"label": "Target Workstation", "value": target_workstation or "-"},
			{"label": "Target Start", "value": _format_manual_adjustment_datetime(start_time)},
			{"label": "Recalculated End", "value": _format_manual_adjustment_datetime(end_time)},
		]
		if latest_safe_start_time and latest_safe_start_time >= get_datetime(earliest_start_time):
			blocking_context_rows.append(
				{
					"label": "Latest Conflict-Free Start",
					"value": _format_manual_adjustment_datetime(latest_safe_start_time),
				}
			)
		blocking_context_rows.extend(
			[
				{
					"label": "Conflicting Segment",
					"value": "{name} / {item_code} / {workstation} / {start} - {end}".format(
						name=row.get("name") or "-",
						item_code=row.get("item_code") or "-",
						workstation=row.get("workstation") or "-",
						start=_format_manual_adjustment_datetime(row.get("start_time")),
						end=_format_manual_adjustment_datetime(row.get("end_time")),
					),
				}
				for row in mold_overlap_rows
			]
		)
		resolution_suggestions = [
			"Keep the current start time, or choose an earlier time window.",
			"If the whole chain must move later, shift the downstream segments that use the same mold together.",
			"If only part of the quantity should move later, split the current segment first and then move the remaining quantity.",
		]
	elif workstation_overlap_rows:
		blocking_summary = "The target workstation already has scheduling in this time window. The current segment cannot be inserted directly."
		blocking_context_rows = [
			{"label": "Current Segment", "value": segment_name},
			{"label": "Target Workstation", "value": target_workstation or "-"},
			{"label": "Target Start", "value": _format_manual_adjustment_datetime(start_time)},
			{"label": "Recalculated End", "value": _format_manual_adjustment_datetime(end_time)},
		]
		blocking_context_rows.extend(
			[
				{
					"label": "Workstation Conflict",
					"value": "{name} / {start} - {end}".format(
						name=row.get("name") or "-",
						start=_format_manual_adjustment_datetime(row.get("start_time")),
						end=_format_manual_adjustment_datetime(row.get("end_time")),
					),
				}
				for row in workstation_overlap_rows
			]
		)
		resolution_suggestions = [
			"Try a free window on the target workstation.",
			"If you only want to change sequence, place the current segment before the target segment in an available slot.",
		]

	return {
		"allowed": 0 if blocked else 1,
		"segment_name": segment_name,
		"result_name": result.name,
		"planning_run": run_doc.name,
		"target_workstation": target_workstation,
		"target_mould_reference": candidate.get("mould_reference"),
		"target_plant_floor": candidate.get("plant_floor"),
		"lane_key": candidate.get("lane_key"),
		"requested_start_time": get_datetime(target_start_time) if target_start_time else None,
		"earliest_start_time": earliest_start_time,
		"start_time": start_time,
		"end_time": end_time,
		"setup_minutes": setup_minutes,
		"planned_qty": planned_qty,
		"hourly_capacity_qty": hourly_capacity,
		"blocking_reasons": list(dict.fromkeys(blocking_reasons)),
		"blocking_title": "Manual Move Blocked",
		"blocking_summary": blocking_summary,
		"blocking_context_rows": blocking_context_rows,
		"resolution_suggestions": resolution_suggestions,
		"latest_safe_start_time": latest_safe_start_time,
		"preview_exceptions": setup_exceptions,
		"override_available": override_available,
		"override_reason": override_reason,
		"schedule_explanation": _("Manual move to {0} with mold {1}.").format(
			target_workstation,
			candidate.get("mould_reference"),
		),
	}


def _diagnose_target_workstation_failure(
	item_code: str,
	target_workstation: str,
	plant_floors: list[str] | str | None,
	item_context: dict[str, Any] | None = None,
) -> list[str]:
	item_context = item_context or {}
	capability_rows = [
		row
		for row in _get_machine_capability_rows(plant_floors)
		if row.get("workstation") == target_workstation
	]
	if not capability_rows:
		return [
			_("Workstation {0} is not enabled in APS Machine Capability for the selected plant floor scope {1}.").format(
				target_workstation,
				", ".join(_coerce_plant_floor_list(plant_floors=plant_floors)) or _("Unknown"),
			)
		]

	capability = capability_rows[0]
	reasons = []
	if capability.get("machine_status") not in APS_ALLOWED_MACHINE_STATUSES:
		reasons.append(
			_("Workstation {0} is currently {1}.").format(
				target_workstation,
				capability.get("machine_status"),
			)
		)

	mold_rows = _get_available_mold_rows(item_code)
	if not mold_rows:
		reasons.append(_("No active mold is available for {0}.").format(item_code))
		return reasons

	tonnage_candidates = []
	required_tonnages = []
	for mold_row in mold_rows:
		required_tonnage = flt(mold_row.get("machine_tonnage"))
		if required_tonnage > 0:
			required_tonnages.append(required_tonnage)
		if not capability.get("machine_tonnage") or required_tonnage <= 0:
			tonnage_candidates.append(mold_row)
			continue
		if flt(capability.get("machine_tonnage")) >= required_tonnage:
			tonnage_candidates.append(mold_row)

	if not tonnage_candidates:
		minimum_required_tonnage = min(required_tonnages) if required_tonnages else 0
		reasons.append(
			_("Workstation {0} tonnage {1}T does not meet the minimum mold tonnage {2}T required for {3}.").format(
				target_workstation,
				frappe.format(capability.get("machine_tonnage") or 0, {"fieldtype": "Float"}),
				frappe.format(minimum_required_tonnage or 0, {"fieldtype": "Float"}),
				item_code,
			)
		)
		return reasons

	workstation_rules = frappe.get_all(
		"APS Mould-Machine Rule",
		filters=_strip_none({"item_code": item_code, "workstation": target_workstation, "is_active": 1}),
		fields=["workstation", "priority", "preferred", "mould_reference", "min_tonnage", "max_tonnage"],
		order_by="preferred desc, priority asc",
	)
	if workstation_rules:
		matched_rules = [
			_match_rule_for_candidate(workstation_rules, capability, mold_row)
			for mold_row in tonnage_candidates
		]
		if not any(matched_rules):
			reasons.append(
				_("APS Mould-Machine Rule does not allow item {0} on workstation {1}.").format(
					item_code,
					target_workstation,
				)
			)
			return reasons

	if item_context and any(_has_fda_conflict(item_context, {**capability, "risk_category": capability.get("risk_category")}) for _ in tonnage_candidates):
		reasons.append(
			_("Workstation {0} violates FDA restriction for {1}. Use manual risk override only if approved.").format(
				target_workstation,
				item_code,
			)
		)

	return reasons or [
		_("Workstation {0} is not a valid lane for {1}.").format(target_workstation, item_code)
	]


def _format_manual_adjustment_datetime(value) -> str:
	if not value:
		return "-"
	return frappe.format(get_datetime(value), {"fieldtype": "Datetime"})


def apply_manual_schedule_adjustment(
	segment_name: str,
	target_workstation: str | None = None,
	before_segment_name: str | None = None,
	target_start_time=None,
	target_end_time=None,
	manual_note: str | None = None,
	allow_locked: int = 0,
	allow_risk_override: int = 0,
) -> dict[str, Any]:
	preview = preview_manual_schedule_adjustment(
		segment_name=segment_name,
		target_workstation=target_workstation,
		before_segment_name=before_segment_name,
		target_start_time=target_start_time,
		target_end_time=target_end_time,
		allow_locked=allow_locked,
		allow_risk_override=allow_risk_override,
	)
	if not preview.get("allowed"):
		frappe.throw("\n".join(preview.get("blocking_reasons") or [_("Manual adjustment is blocked.")]))

	rows = frappe.get_all(
		"APS Schedule Segment",
		filters={"name": segment_name},
		fields=["name", "parent", "segment_kind", "family_group"],
		limit=1,
	)
	if not rows:
		frappe.throw(_("APS Schedule Segment {0} was not found.").format(segment_name))
	segment = rows[0]
	filters = {"parent": segment.parent}
	if segment.get("family_group"):
		filters["family_group"] = segment.get("family_group")
	else:
		filters["name"] = segment.name
	child_segments = frappe.get_all(
		"APS Schedule Segment",
		filters=filters,
		fields=["name", "segment_kind", "planned_qty"],
	)
	result_doc = frappe.get_doc("APS Schedule Result", segment.parent)
	run_doc = frappe.get_doc("APS Planning Run", result_doc.planning_run)
	primary_base_qty = 0.0
	for row in child_segments:
		if row.segment_kind != "Family Co-Product":
			primary_base_qty = flt(row.planned_qty)
			break
	for row in child_segments:
		risk_flags = []
		if cint(allow_risk_override):
			risk_flags.append("FDA Override")
		planned_qty = preview.get("planned_qty")
		if row.segment_kind == "Family Co-Product":
			ratio = 0
			if primary_base_qty > 0:
				ratio = flt(row.planned_qty) / primary_base_qty
			planned_qty = flt(preview.get("planned_qty")) * ratio if ratio else flt(row.planned_qty)
		values = {
			"workstation": preview["target_workstation"],
			"plant_floor": preview.get("target_plant_floor"),
			"start_time": preview["start_time"],
			"end_time": preview["end_time"],
			"planned_qty": planned_qty,
			"mould_reference": preview["target_mould_reference"],
			"lane_key": preview["lane_key"],
			"campaign_key": _build_campaign_key(result_doc.item_code, preview["target_mould_reference"], preview["target_workstation"]),
			"anchor_strength": ANCHOR_STRENGTH_SOFT,
			"execution_anchor_source": "Manual Adjustment",
			"is_manual": 1,
			"manual_change_note": manual_note or preview.get("schedule_explanation"),
			"risk_flags": "\n".join(risk_flags),
		}
		if row.segment_kind != "Family Co-Product":
			values["segment_kind"] = "Manual"
		frappe.db.set_value("APS Schedule Segment", row.name, values)

	result_doc.db_set(
		{
			"is_manual": 1,
			"plant_floor": preview.get("target_plant_floor"),
			"risk_status": "Attention" if cint(allow_risk_override) else result_doc.risk_status,
			"flow_step": "Manual Adjustment Pending Confirmation",
			"next_step_hint": "Confirm Run",
			"blocking_reason": _("Manual FDA override was applied.") if cint(allow_risk_override) else "",
			"primary_mould_reference": preview["target_mould_reference"],
			"selected_moulds": preview["target_mould_reference"],
			"schedule_explanation": preview.get("schedule_explanation"),
		}
	)
	_refresh_result_after_manual_adjustment(result_doc.name)
	if run_doc.status in ("Approved", "Work Order Proposed", "Shift Proposed", "Applied"):
		run_doc.db_set({"status": "Planned", "approval_state": "Pending"})
	if cint(allow_risk_override):
		_create_exception(
			planning_run=run_doc.name,
			severity="Warning",
			exception_type="FDA Override",
			message=_("Manual adjustment placed {0} on {1} with FDA override.").format(result_doc.item_code, preview["target_workstation"]),
			item_code=result_doc.item_code,
			customer=result_doc.customer,
			workstation=preview["target_workstation"],
			source_doctype="APS Schedule Result",
			source_name=result_doc.name,
			resolution_hint=_("Review override approval before syncing or releasing."),
			is_blocking=0,
		)

	return {
		"segment_name": segment_name,
		"result_name": result_doc.name,
		"planning_run": run_doc.name,
		"next_actions": get_next_actions_for_context("APS Planning Run", run_doc.name),
	}


def _refresh_result_after_manual_adjustment(result_name: str):
	primary_segments = _get_primary_segments_for_result(result_name)
	if not primary_segments:
		return
	total_scheduled_qty = sum(flt(row.get("planned_qty")) for row in primary_segments)
	selected_moulds = list(
		dict.fromkeys(row.get("mould_reference") for row in primary_segments if row.get("mould_reference"))
	)
	result_doc = frappe.get_doc("APS Schedule Result", result_name)
	unscheduled_qty = max(flt(result_doc.planned_qty) - total_scheduled_qty, 0)
	frappe.db.set_value(
		"APS Schedule Result",
		result_name,
		{
			"scheduled_qty": total_scheduled_qty,
			"unscheduled_qty": unscheduled_qty,
			"primary_mould_reference": selected_moulds[0] if selected_moulds else "",
			"selected_moulds": "\n".join(selected_moulds),
			"plant_floor": _get_primary_result_plant_floor(primary_segments, result_doc.plant_floor),
			"status": "Risk" if unscheduled_qty > 0 else result_doc.status,
			"risk_status": "Attention" if unscheduled_qty > 0 else result_doc.risk_status,
		},
	)


def _build_segment_capacity_snapshot(
	segment_row: dict[str, Any],
	mold_rows: list[dict[str, Any]],
	capability_map: dict[str, dict[str, Any]],
	settings: dict[str, Any],
) -> dict[str, Any]:
	mold_row = next((row for row in mold_rows if row.get("mold") == segment_row.get("mould_reference")), {}) or {}
	capability = capability_map.get(segment_row.get("workstation")) or {}
	candidate = dict(capability)
	candidate["cycle_time_seconds"] = flt(mold_row.get("cycle_time_seconds"))
	candidate["effective_output_qty"] = flt(mold_row.get("effective_output_qty"))
	candidate["output_qty"] = flt(mold_row.get("output_qty"))
	candidate["cavity_output_qty"] = flt(mold_row.get("cavity_output_qty"))
	return _build_capacity_display(candidate, settings)


def get_schedule_result_detail(result_name: str) -> dict[str, Any]:
	result = frappe.get_doc("APS Schedule Result", result_name)
	settings = get_settings_dict()
	segments = frappe.get_all(
		"APS Schedule Segment",
		filters={"parent": result_name, "parenttype": "APS Schedule Result"},
		fields=[
			"name",
			"workstation",
			"start_time",
			"end_time",
			"planned_qty",
			"sequence_no",
			"lane_key",
			"parallel_group",
			"family_group",
			"segment_kind",
			"primary_item_code",
			"co_product_item_code",
			"mould_reference",
			"segment_status",
			"is_locked",
			"is_manual",
			"schedule_explanation",
			"risk_flags",
			"segment_note",
			"manual_change_note",
			"linked_work_order",
			"linked_work_order_scheduling",
			"linked_scheduling_item",
			"actual_status",
			"actual_completed_qty",
			"actual_start_time",
			"actual_end_time",
			"delay_minutes",
		],
		order_by="sequence_no asc, idx asc",
	)
	stock_entry_map = _get_latest_stock_entry_by_work_order(
		[row.get("linked_work_order") for row in segments if row.get("linked_work_order")]
	)
	for row in segments:
		row["work_order_route"] = _build_form_route("Work Order", row.get("linked_work_order"))
		row["work_order_scheduling_route"] = _build_form_route("Work Order Scheduling", row.get("linked_work_order_scheduling"))
		row["scheduling_item_route"] = _build_form_route("Scheduling Item", row.get("linked_scheduling_item"))
		stock_entry = stock_entry_map.get(row.get("linked_work_order"))
		row["latest_stock_entry"] = stock_entry.get("name") if stock_entry else None
		row["latest_stock_entry_route"] = _build_form_route("Stock Entry", stock_entry.get("name") if stock_entry else None)
	item_detail = _get_item_detail_snapshot(result.item_code, result.customer, settings)
	source_rows = _get_result_source_rows(result)
	exception_rows = _get_result_exception_rows(result)
	mold_rows = _get_result_mold_rows(result, segments)
	workstations = list(dict.fromkeys(row.get("workstation") for row in segments if row.get("workstation")))
	capability_map = (
		{
			row.workstation: row
			for row in frappe.get_all(
				"APS Machine Capability",
				filters={"workstation": ("in", workstations)},
				fields=["workstation", "hourly_capacity_qty", "daily_capacity_qty", "machine_status"],
			)
		}
		if workstations
		else {}
	)
	for row in segments:
		row.update(_build_segment_capacity_snapshot(row, mold_rows, capability_map, settings))
	return {
		"result": result.as_dict(),
		"segments": segments,
		"item_detail": item_detail,
		"source_rows": source_rows,
		"exception_rows": exception_rows,
		"mold_rows": mold_rows,
		"routes": {
			"planning_run": _build_form_route("APS Planning Run", result.planning_run),
			"net_requirement": _build_form_route("APS Net Requirement", result.net_requirement),
			"result": _build_form_route("APS Schedule Result", result.name),
		},
		"next_actions": get_next_actions_for_context("APS Planning Run", result.planning_run),
	}


def get_exception_resolution_context(exception_name: str) -> dict[str, Any]:
	doc = frappe.get_doc("APS Exception Log", exception_name)
	return _build_exception_resolution_context(doc)


def detach_standard_references(dry_run: bool = True) -> dict[str, Any]:
	rows = []
	for doctype, fieldnames in {
		"Work Order": [
			"custom_aps_run",
			"custom_aps_source",
			"custom_aps_required_delivery_date",
			"custom_aps_is_urgent",
			"custom_aps_release_status",
			"custom_aps_locked_for_reschedule",
			"custom_aps_schedule_reference",
			"custom_aps_result_reference",
			"custom_aps_proposal_batch",
		],
		"Work Order Scheduling": [
			"custom_aps_run",
			"custom_aps_freeze_state",
			"custom_aps_approval_state",
		],
		"Scheduling Item": [
			"custom_aps_run",
			"custom_aps_result_reference",
			"custom_aps_segment_reference",
			"custom_aps_shift_proposal",
		],
		"Delivery Plan": [
			"custom_aps_version",
			"custom_aps_source",
		],
	}.items():
		if not frappe.db.exists("DocType", doctype):
			continue
		names = _get_records_with_any_field_set(doctype, fieldnames)
		rows.append({"doctype": doctype, "count": len(names), "names": names[:20]})
		if not dry_run and names:
			for name in names:
				values = {fieldname: None for fieldname in fieldnames if frappe.get_meta(doctype).has_field(fieldname)}
				frappe.db.set_value(doctype, name, values)
	return {"dry_run": cint(dry_run), "rows": rows}


def get_settings_dict() -> dict[str, Any]:
	settings = frappe.get_cached_doc("APS Settings", "APS Settings")
	return {
		"default_company": settings.default_company,
		"default_plant_floor": settings.default_plant_floor,
		"planning_horizon_days": cint(settings.planning_horizon_days or 14),
		"release_horizon_days": cint(settings.release_horizon_days or 3),
		"freeze_days": cint(settings.freeze_days or 2),
		"minimum_parallel_split_qty": flt(settings.minimum_parallel_split_qty or 500),
		"minimum_run_window_hours": flt(settings.minimum_run_window_hours or 2),
		"default_setup_minutes": flt(settings.default_setup_minutes or 30),
		"default_first_article_minutes": flt(settings.default_first_article_minutes or 45),
		"mold_change_penalty_minutes": flt(settings.mold_change_penalty_minutes or 30),
		"missing_cycle_fallback_seconds": flt(settings.missing_cycle_fallback_seconds or 60),
		"default_hourly_capacity_qty": flt(settings.default_hourly_capacity_qty or 120),
		"item_food_grade_field": settings.item_food_grade_field or "custom_food_grade",
		"item_first_article_field": settings.item_first_article_field or "custom_is_first_article",
		"item_color_field": settings.item_color_field or "color",
		"item_material_field": settings.item_material_field or "material",
		"item_safety_stock_field": settings.item_safety_stock_field or "safety_stock",
		"item_max_stock_field": settings.item_max_stock_field or "max_stock_qty",
		"item_min_batch_field": settings.item_min_batch_field or "min_order_qty",
		"customer_short_name_field": settings.customer_short_name_field or "custom_customer_short_name",
		"workstation_risk_field": settings.workstation_risk_field or "custom_production_risk_category",
		"scheduling_item_risk_field": settings.scheduling_item_risk_field or "custom_workstation_risk_category_",
		"plant_floor_source_warehouse_field": settings.plant_floor_source_warehouse_field or "custom_default_source_warehouse",
		"plant_floor_wip_warehouse_field": settings.plant_floor_wip_warehouse_field or "warehouse",
		"plant_floor_fg_warehouse_field": settings.plant_floor_fg_warehouse_field or "custom_default_finished_goods_warehouse",
		"plant_floor_scrap_warehouse_field": settings.plant_floor_scrap_warehouse_field or "custom_default_scrap_warehouse",
	}


def inspect_customer_delivery_schedule_file(
	file_url: str,
	sheet_name: str | None = None,
	header_row_no: int | None = None,
	max_rows: int = 16,
) -> dict[str, Any]:
	sheet_rows, workbook_context = _read_schedule_workbook_rows(
		file_url=file_url,
		sheet_name=sheet_name,
		max_rows=max_rows,
	)
	guess = _guess_schedule_mapping(sheet_rows, forced_header_row_no=header_row_no)
	return {
		"sheet_names": workbook_context.get("sheet_names") or [],
		"selected_sheet": workbook_context.get("sheet_name"),
		"sample_rows": sheet_rows[:max_rows],
		"column_options": _build_schedule_column_options(sheet_rows, guess.get("header_row_no") or 1),
		"detected_mapping": guess,
	}


def _normalize_schedule_rows(
	file_url: str | None = None,
	rows_json: str | list[dict] | None = None,
	mapping_json: str | dict[str, Any] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
	mapping = _normalize_schedule_mapping(mapping_json)
	parser_mode = mapping.get("parser_mode") or "rows"
	parse_context = {
		"parser_mode": parser_mode,
		"sheet_name": mapping.get("sheet_name"),
		"mapping": mapping,
	}
	if rows_json:
		if isinstance(rows_json, str):
			data_rows = json.loads(rows_json or "[]")
		else:
			data_rows = rows_json or []
		parse_context["parser_mode"] = "rows"
		parse_context["source_mode"] = "rows_json"
		normalized = _normalize_schedule_rows_from_long_rows(data_rows)
		return normalized, parse_context
	if file_url and parser_mode == "matrix":
		rows, matrix_context = _normalize_schedule_rows_from_matrix(file_url=file_url, mapping=mapping)
		parse_context.update(matrix_context)
		return rows, parse_context
	if file_url:
		raw_rows, workbook_context = _read_schedule_workbook_rows(
			file_url=file_url,
			sheet_name=mapping.get("sheet_name"),
		)
		if not raw_rows:
			return [], parse_context
		header_row_no = cint(mapping.get("header_row_no") or 1)
		data_start_row_no = cint(mapping.get("data_start_row_no") or (header_row_no + 1))
		header_idx = max(header_row_no - 1, 0)
		headers = [_normalize_header(cell) for cell in (raw_rows[header_idx] if len(raw_rows) > header_idx else [])]
		data_rows = [
			{headers[idx]: row[idx] for idx in range(min(len(headers), len(row))) if headers[idx]}
			for row in raw_rows[max(data_start_row_no - 1, header_idx + 1) :]
			if any(cell not in (None, "") for cell in row)
		]
		parse_context.update(
			{
				"sheet_name": workbook_context.get("sheet_name"),
				"header_row_no": header_row_no,
				"data_start_row_no": data_start_row_no,
				"source_mode": "file_rows",
			}
		)
	else:
		data_rows = []

	normalized = _normalize_schedule_rows_from_long_rows(data_rows)
	return normalized, parse_context


def _normalize_schedule_rows_from_long_rows(data_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
	normalized = []
	for idx, row in enumerate(data_rows or [], start=1):
		item_code = row.get("item_code") or row.get("item") or row.get("item code")
		if not item_code:
			continue
		item_code = _resolve_item_name(item_code) or item_code
		schedule_date = row.get("schedule_date") or row.get("schedule date") or row.get("delivery_date") or row.get("delivery date")
		normalized.append(
			{
				"sales_order": row.get("sales_order") or row.get("sales order"),
				"item_code": item_code,
				"customer_part_no": row.get("customer_part_no") or row.get("customer part no"),
				"schedule_date": getdate(schedule_date) if schedule_date else getdate(today()),
				"qty": flt(row.get("qty") or row.get("quantity")),
				"remark": row.get("remark") or row.get("remarks"),
				"source_origin": row.get("source_origin") or "imported",
				"source_excel_row": cint(row.get("source_excel_row") or row.get("excel_row") or row.get("row_no") or 0) or idx,
				"manual_override": cint(row.get("manual_override")),
				"manual_change_reason": row.get("manual_change_reason"),
			}
		)
	return normalized


def _normalize_schedule_rows_from_matrix(
	file_url: str,
	mapping: dict[str, Any],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
	raw_rows, workbook_context = _read_schedule_workbook_rows(
		file_url=file_url,
		sheet_name=mapping.get("sheet_name"),
	)
	if not raw_rows:
		return [], workbook_context

	header_row_no = cint(mapping.get("header_row_no") or 1)
	data_start_row_no = cint(mapping.get("data_start_row_no") or (header_row_no + 1))
	header_idx = max(header_row_no - 1, 0)
	headers = raw_rows[header_idx] if len(raw_rows) > header_idx else []
	item_idx = _coerce_excel_column_index(mapping.get("item_reference_column"))
	customer_part_idx = _coerce_excel_column_index(mapping.get("customer_part_no_column"))
	description_idx = _coerce_excel_column_index(mapping.get("description_column"))
	sales_order_idx = _coerce_excel_column_index(mapping.get("sales_order_column"))
	row_type_idx = _coerce_excel_column_index(mapping.get("row_type_column"))
	remark_idx = _coerce_excel_column_index(mapping.get("remark_column"))
	demand_row_type_value = str(mapping.get("demand_row_type_value") or "").strip()
	skip_zero_qty = cint(mapping.get("skip_zero_qty") if mapping.get("skip_zero_qty") is not None else 1)
	date_columns = _resolve_matrix_date_columns(headers, mapping)
	if item_idx is None:
		frappe.throw(_("Please select Item Reference Column before previewing matrix schedule import."))
	if not date_columns:
		frappe.throw(_("No date columns were detected for the selected matrix mapping."))
	normalized = []

	for row_number, row in enumerate(raw_rows[max(data_start_row_no - 1, header_idx + 1) :], start=max(data_start_row_no, header_row_no + 1)):
		if not any(cell not in (None, "") for cell in row):
			continue
		item_reference = _get_row_value(row, item_idx)
		if not item_reference:
			continue
		row_type_value = str(_get_row_value(row, row_type_idx) or "").strip()
		if demand_row_type_value and row_type_value and row_type_value != demand_row_type_value:
			continue
		if demand_row_type_value and row_type_idx is not None and not row_type_value:
			continue
		item_name = _resolve_item_name(item_reference) or str(item_reference).strip()
		description = _get_row_value(row, description_idx)
		customer_part_no = _get_row_value(row, customer_part_idx) or item_reference
		sales_order = _get_row_value(row, sales_order_idx)
		remark = _get_row_value(row, remark_idx)
		for date_column in date_columns:
			column_idx = date_column["index"]
			qty = flt(_get_row_value(row, column_idx))
			if skip_zero_qty and abs(qty) < 0.000001:
				continue
			normalized.append(
				{
					"sales_order": sales_order,
					"item_code": item_name,
					"customer_part_no": customer_part_no,
					"schedule_date": date_column["schedule_date"],
					"qty": qty,
					"remark": _build_matrix_row_remark(
						base_remark=remark,
						description=description,
						row_type=row_type_value,
						cell_ref=f"{get_column_letter(column_idx + 1)}{row_number}",
					),
					"source_excel_row": row_number,
				}
			)

	workbook_context.update(
		{
			"parser_mode": "matrix",
			"header_row_no": header_row_no,
			"data_start_row_no": data_start_row_no,
			"date_column_count": len(date_columns),
			"mapping": mapping,
		}
	)
	return normalized, workbook_context


def _build_matrix_row_remark(
	base_remark: Any = None,
	description: Any = None,
	row_type: Any = None,
	cell_ref: str | None = None,
) -> str:
	parts = []
	if base_remark:
		parts.append(str(base_remark).strip())
	if description:
		parts.append(_("Description: {0}").format(str(description).strip()))
	if row_type:
		parts.append(_("Row Type: {0}").format(str(row_type).strip()))
	if cell_ref:
		parts.append(_("Source Cell: {0}").format(cell_ref))
	return " | ".join(part for part in parts if part)


def _normalize_schedule_mapping(mapping_json: str | dict[str, Any] | None) -> dict[str, Any]:
	if not mapping_json:
		return {"parser_mode": "rows"}
	if isinstance(mapping_json, str):
		try:
			mapping = json.loads(mapping_json)
		except Exception:
			mapping = {}
	else:
		mapping = dict(mapping_json or {})
	mapping["parser_mode"] = mapping.get("parser_mode") or "rows"
	return mapping


def _read_schedule_workbook_rows(
	file_url: str,
	sheet_name: str | None = None,
	max_rows: int | None = None,
) -> tuple[list[list[Any]], dict[str, Any]]:
	file_doc = frappe.get_doc("File", {"file_url": file_url})
	workbook = load_workbook(filename=file_doc.get_full_path(), data_only=True, read_only=True)
	selected_sheet_name = ""
	try:
		sheet_names = list(workbook.sheetnames)
		worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active
		selected_sheet_name = worksheet.title
		rows = []
		for row in worksheet.iter_rows(max_row=max_rows):
			rows.append([cell.value for cell in row])
	finally:
		workbook.close()
	return rows, {"sheet_name": selected_sheet_name, "sheet_names": sheet_names}


def _guess_schedule_mapping(rows: list[list[Any]], forced_header_row_no: int | None = None) -> dict[str, Any]:
	header_row_no = cint(forced_header_row_no or 0) or _guess_schedule_header_row_no(rows)
	column_options = _build_schedule_column_options(rows, header_row_no)
	header_idx = max(header_row_no - 1, 0)
	headers = rows[header_idx] if len(rows) > header_idx else []
	item_reference_column = _guess_column_by_labels(headers, ("mtlpartnum", "item code", "item", "material", "part no", "料号", "物料"))
	customer_part_no_column = item_reference_column
	description_column = _guess_column_by_labels(headers, ("description", "描述", "desc"))
	row_type_column = _guess_column_by_labels(headers, ("类别", "category", "type", "row type"))
	plan_qty_column = _guess_column_by_labels(headers, ("plan qty", "总数量", "交货日期/总数量", "plan quantity"))
	po_qty_column = _guess_column_by_labels(headers, ("po qty", "订单数量", "order qty"))
	date_columns = _resolve_matrix_date_columns(headers, {"date_columns_mode": "auto"})
	data_start_row_no = header_row_no + 1
	row_type_values = []
	if row_type_column:
		row_type_idx = _coerce_excel_column_index(row_type_column)
		for row in rows[data_start_row_no - 1 : min(len(rows), data_start_row_no + 20)]:
			value = str(_get_row_value(row, row_type_idx) or "").strip()
			if value and value not in row_type_values:
				row_type_values.append(value)
	demand_row_type_value = next((value for value in row_type_values if "交货" in value or "delivery" in value.lower()), row_type_values[0] if row_type_values else "")
	return {
		"parser_mode": "matrix" if len(date_columns) >= 2 else "rows",
		"header_row_no": header_row_no,
		"data_start_row_no": data_start_row_no,
		"item_reference_column": item_reference_column or "",
		"customer_part_no_column": customer_part_no_column or "",
		"description_column": description_column or "",
		"row_type_column": row_type_column or "",
		"demand_row_type_value": demand_row_type_value,
		"plan_qty_column": plan_qty_column or "",
		"po_qty_column": po_qty_column or "",
		"date_columns_mode": "auto",
		"date_start_column": date_columns[0]["column"] if date_columns else "",
		"date_end_column": date_columns[-1]["column"] if date_columns else "",
		"date_column_letters": [row["column"] for row in date_columns],
		"column_options": column_options,
		"row_type_values": row_type_values,
	}


def _guess_schedule_header_row_no(rows: list[list[Any]]) -> int:
	best_row_no = 1
	best_score = -1
	for idx, row in enumerate(rows[:10], start=1):
		headers = [_normalize_header(cell) for cell in row]
		date_like = len(_resolve_matrix_date_columns(row, {"date_columns_mode": "auto"}))
		score = sum(
			1
			for value in headers
			if value in {"mtlpartnum", "description", "po qty", "plan qty", "类别", "category", "type"}
		) * 10
		score += date_like * 3
		score += sum(1 for value in headers if value) * 0.2
		if score > best_score:
			best_score = score
			best_row_no = idx
	return best_row_no


def _build_schedule_column_options(rows: list[list[Any]], header_row_no: int) -> list[dict[str, Any]]:
	header_idx = max(header_row_no - 1, 0)
	headers = rows[header_idx] if len(rows) > header_idx else []
	options = []
	for idx, header in enumerate(headers):
		column_letter = get_column_letter(idx + 1)
		label = f"{column_letter} · {str(header or '').strip() or _('Blank Header')}"
		options.append({"value": column_letter, "label": label})
	return options


def _guess_column_by_labels(headers: list[Any], candidates: tuple[str, ...]) -> str | None:
	normalized_headers = [_normalize_header(cell) for cell in headers]
	for candidate in candidates:
		for idx, header in enumerate(normalized_headers):
			if header == candidate:
				return get_column_letter(idx + 1)
	for candidate in candidates:
		for idx, header in enumerate(normalized_headers):
			if candidate and candidate in header:
				return get_column_letter(idx + 1)
	return None


def _resolve_matrix_date_columns(headers: list[Any], mapping: dict[str, Any]) -> list[dict[str, Any]]:
	mode = mapping.get("date_columns_mode") or "auto"
	start_idx = _coerce_excel_column_index(mapping.get("date_start_column"))
	end_idx = _coerce_excel_column_index(mapping.get("date_end_column"))
	candidate_indices = []
	for idx, header in enumerate(headers or []):
		if mode == "range":
			if start_idx is None:
				continue
			if idx < start_idx:
				continue
			if end_idx is not None and idx > end_idx:
				continue
		parsed = _parse_schedule_header_date(header)
		if parsed:
			candidate_indices.append((idx, parsed))
	results = []
	previous_date = None
	for idx, parsed in candidate_indices:
		schedule_date = parsed["date"]
		if previous_date and not parsed["has_year"]:
			while schedule_date <= previous_date:
				schedule_date = schedule_date.replace(year=schedule_date.year + 1)
		results.append(
			{
				"index": idx,
				"column": get_column_letter(idx + 1),
				"header": headers[idx],
				"schedule_date": schedule_date,
			}
		)
		previous_date = schedule_date
	return results


def _parse_schedule_header_date(value: Any) -> dict[str, Any] | None:
	if value in (None, ""):
		return None
	if isinstance(value, datetime):
		return {"date": value.date(), "has_year": True}
	if isinstance(value, date_cls):
		return {"date": value, "has_year": True}
	text = str(value).strip()
	if not text:
		return None
	patterns = [
		("%d-%b-%Y", True),
		("%d-%b-%y", True),
		("%Y-%m-%d", True),
		("%Y/%m/%d", True),
		("%d/%m/%Y", True),
		("%m/%d/%Y", True),
		("%d.%m.%Y", True),
		("%d-%b", False),
		("%d/%m", False),
		("%m/%d", False),
	]
	current_year = getdate(today()).year
	for pattern, has_year in patterns:
		try:
			parsed = datetime.strptime(text, pattern)
		except Exception:
			continue
		if not has_year:
			parsed = parsed.replace(year=current_year)
		return {"date": parsed.date(), "has_year": has_year}
	try:
		parsed_date = getdate(text)
	except Exception:
		return None
	return {"date": parsed_date, "has_year": any(char.isdigit() for char in text if char in "-/.")}


def _coerce_excel_column_index(value: Any) -> int | None:
	if value in (None, ""):
		return None
	if isinstance(value, int):
		return max(value - 1, 0)
	text = str(value).strip()
	if not text:
		return None
	if "·" in text:
		text = text.split("·", 1)[0].strip()
	if " " in text and not text.isdigit():
		text = text.split(" ", 1)[0].strip()
	if text.isdigit():
		return max(int(text) - 1, 0)
	return column_index_from_string(text.upper()) - 1


def _get_row_value(row: list[Any], index: int | None) -> Any:
	if index is None or index < 0 or index >= len(row or []):
		return None
	return row[index]


def _normalize_header(value: Any) -> str:
	return str(value or "").strip().lower().replace("_", " ")


def _resolve_item_name(item_reference: str | None) -> str | None:
	if not item_reference or not frappe.db.exists("DocType", "Item"):
		return None

	reference = str(item_reference).strip()
	if not reference:
		return None

	cache = _get_request_cache("injection_aps_item_resolution_cache")
	if reference in cache:
		return cache[reference] or None

	item_name = None
	if frappe.db.exists("Item", reference):
		item_name = reference
	else:
		item_name = _get_unique_item_name_by_field("item_code", reference)
		if not item_name:
			item_name = _get_unique_item_name_by_field("item_name", reference)
		if not item_name:
			for prefix in ITEM_NAME_PREFIX_FALLBACKS:
				item_name = _get_unique_item_name_by_field("item_name", f"{prefix}{reference}")
				if item_name:
					break

	cache[reference] = item_name or ""
	return item_name or None


def _require_item_name(item_reference: str | None) -> str:
	item_name = _resolve_item_name(item_reference)
	if item_name:
		return item_name
	raise APSItemReferenceError(_("Item reference {0} could not be resolved to an Item record.").format(item_reference or ""))


def _get_request_cache(cache_key: str) -> dict[str, Any]:
	cache = getattr(frappe.local, cache_key, None)
	if cache is None:
		cache = {}
		setattr(frappe.local, cache_key, cache)
	return cache


def _get_unique_item_name_by_field(fieldname: str, value: str) -> str | None:
	if not value:
		return None
	names = frappe.get_all("Item", filters={fieldname: value}, pluck="name", limit=2)
	return names[0] if len(names) == 1 else None


def repair_item_references(
	company: str | None = None,
	include_standard: int = 1,
	include_aps: int = 1,
	commit: bool = False,
) -> dict[str, Any]:
	repaired_rows = []
	unresolved_rows = []
	target_summaries = []

	for target in _get_item_reference_repair_targets(
		company=company,
		include_standard=include_standard,
		include_aps=include_aps,
	):
		rows = frappe.db.sql(target["query"], target.get("params") or [], as_dict=True)
		repaired_count, unresolved_count = _repair_item_reference_rows(
			target=target,
			rows=rows,
			repaired_rows=repaired_rows,
			unresolved_rows=unresolved_rows,
		)
		target_summaries.append(
			{
				"label": target["label"],
				"doctype": target["doctype"],
				"scanned_rows": len(rows),
				"repaired_rows": repaired_count,
				"unresolved_rows": unresolved_count,
			}
		)

	if commit and (repaired_rows or unresolved_rows):
		frappe.db.commit()

	return {
		"repaired_count": sum(row["repaired_rows"] for row in target_summaries),
		"unresolved_count": sum(row["unresolved_rows"] for row in target_summaries),
		"repaired_rows": repaired_rows[:MAX_REBUILD_WARNINGS],
		"unresolved_rows": unresolved_rows[:MAX_REBUILD_WARNINGS],
		"targets": target_summaries,
	}


def _get_item_reference_repair_targets(
	company: str | None = None,
	include_standard: int = 1,
	include_aps: int = 1,
) -> list[dict[str, Any]]:
	targets = []
	company_filter_sql = ""
	company_params: list[Any] = []
	if company:
		company_filter_sql = " and {table_alias}.company = %s"
		company_params = [company]

	if cint(include_aps):
		targets.extend(
			[
				{
					"label": "Customer Delivery Schedule Item",
					"doctype": "Customer Delivery Schedule Item",
					"query": """
						select
							cdsi.name,
							cdsi.item_code,
							cdsi.parent as source_name,
							cds.company
						from `tabCustomer Delivery Schedule Item` cdsi
						inner join `tabCustomer Delivery Schedule` cds on cds.name = cdsi.parent
						left join `tabItem` item on item.name = cdsi.item_code
						where ifnull(cdsi.item_code, '') != ''
							and item.name is null
					"""
					+ company_filter_sql.format(table_alias="cds"),
					"params": list(company_params),
				},
				{
					"label": "APS Demand Pool",
					"doctype": "APS Demand Pool",
					"query": """
						select
							dp.name,
							dp.item_code,
							dp.name as source_name,
							dp.company
						from `tabAPS Demand Pool` dp
						left join `tabItem` item on item.name = dp.item_code
						where ifnull(dp.item_code, '') != ''
							and item.name is null
					"""
					+ company_filter_sql.format(table_alias="dp"),
					"params": list(company_params),
				},
				{
					"label": "APS Net Requirement",
					"doctype": "APS Net Requirement",
					"query": """
						select
							nr.name,
							nr.item_code,
							nr.name as source_name,
							nr.company
						from `tabAPS Net Requirement` nr
						left join `tabItem` item on item.name = nr.item_code
						where ifnull(nr.item_code, '') != ''
							and item.name is null
					"""
					+ company_filter_sql.format(table_alias="nr"),
					"params": list(company_params),
				},
				{
					"label": "APS Schedule Result",
					"doctype": "APS Schedule Result",
					"query": """
						select
							sr.name,
							sr.item_code,
							sr.name as source_name,
							sr.company
						from `tabAPS Schedule Result` sr
						left join `tabItem` item on item.name = sr.item_code
						where ifnull(sr.item_code, '') != ''
							and item.name is null
					"""
					+ company_filter_sql.format(table_alias="sr"),
					"params": list(company_params),
				},
				{
					"label": "APS Exception Log",
					"doctype": "APS Exception Log",
					"query": """
						select
							ex.name,
							ex.item_code,
							ex.name as source_name
						from `tabAPS Exception Log` ex
						left join `tabItem` item on item.name = ex.item_code
						where ifnull(ex.item_code, '') != ''
							and item.name is null
					""",
					"params": [],
				},
			]
		)

	if cint(include_standard):
		query = """
			select
				soi.name,
				soi.item_code,
				soi.parent as source_name,
				so.company
			from `tabSales Order Item` soi
			inner join `tabSales Order` so on so.name = soi.parent
			left join `tabItem` item on item.name = soi.item_code
			where ifnull(soi.item_code, '') != ''
				and item.name is null
		"""
		params = []
		if company:
			query += " and so.company = %s"
			params.append(company)
		targets.append(
			{
				"label": "Sales Order Item",
				"doctype": "Sales Order Item",
				"query": query,
				"params": params,
				"read_only": True,
			}
		)

	return targets


def _repair_item_reference_rows(
	target: dict[str, Any],
	rows: list[dict[str, Any]],
	repaired_rows: list[dict[str, Any]],
	unresolved_rows: list[dict[str, Any]],
) -> tuple[int, int]:
	repaired_count = 0
	unresolved_count = 0

	for row in rows:
		current_reference = row.get("item_code")
		resolved_item_code = _resolve_item_name(current_reference)
		if not resolved_item_code:
			unresolved_count += 1
			if len(unresolved_rows) < MAX_REBUILD_WARNINGS:
				unresolved_rows.append(
					{
						"doctype": target["doctype"],
						"docname": row.get("name"),
						"source_name": row.get("source_name"),
						"item_reference": current_reference,
					}
			)
			continue
		if resolved_item_code == current_reference:
			continue
		if target.get("read_only"):
			unresolved_count += 1
			if len(unresolved_rows) < MAX_REBUILD_WARNINGS:
				unresolved_rows.append(
					{
						"doctype": target["doctype"],
						"docname": row.get("name"),
						"source_name": row.get("source_name"),
						"item_reference": current_reference,
						"resolved_item_code": resolved_item_code,
						"message": _("Standard ERPNext rows are not changed automatically by APS."),
					}
				)
			continue
		frappe.db.set_value(
			target["doctype"],
			row.get("name"),
			"item_code",
			resolved_item_code,
			update_modified=False,
		)
		repaired_count += 1
		if len(repaired_rows) < MAX_REBUILD_WARNINGS:
			repaired_rows.append(
				{
					"doctype": target["doctype"],
					"docname": row.get("name"),
					"source_name": row.get("source_name"),
					"old_item_reference": current_reference,
					"new_item_code": resolved_item_code,
				}
			)

	return repaired_count, unresolved_count


def _append_rebuild_warning(
	warnings: list[dict[str, Any]],
	warning_keys: set[tuple[str, str, str, str]],
	*,
	item_reference: str | None,
	source_doctype: str,
	source_name: str | None = None,
	row_name: str | None = None,
):
	key = (
		source_doctype or "",
		source_name or "",
		row_name or "",
		str(item_reference or ""),
	)
	if key in warning_keys:
		return
	warning_keys.add(key)
	message = _("Skipped {0} {1} because item reference {2} could not be resolved to an Item record.").format(
		source_doctype,
		source_name or row_name or "",
		item_reference or _("(blank)"),
	)
	warnings.append(
		{
			"source_doctype": source_doctype,
			"source_name": source_name,
			"row_name": row_name,
			"item_reference": item_reference,
			"message": message,
		}
	)


def _append_item_group_warning(
	warnings: list[dict[str, Any]],
	warning_keys: set[tuple[str, str, str, str]],
	*,
	item_code: str | None,
	source_doctype: str,
	source_name: str | None = None,
	row_name: str | None = None,
	item_group: str | None = None,
):
	key = (
		source_doctype or "",
		source_name or "",
		row_name or "",
		f"item-group::{item_code or ''}",
	)
	if key in warning_keys:
		return
	warning_keys.add(key)
	warnings.append(
		{
			"source_doctype": source_doctype,
			"source_name": source_name,
			"row_name": row_name,
			"item_reference": item_code,
			"message": _(
				"Skipped {0} {1} because item {2} belongs to item group {3}. APS only schedules {4}."
			).format(
				source_doctype,
				source_name or row_name or "",
				item_code or _("(blank)"),
				item_group or _("(blank)"),
				", ".join(SCHEDULABLE_ITEM_GROUPS),
			),
		}
	)


def _get_item_group(item_code: str | None) -> str:
	item_name = _resolve_item_name(item_code)
	if not item_name or not frappe.db.exists("DocType", "Item"):
		return ""
	cache = _get_request_cache("injection_aps_item_group_cache")
	if item_name not in cache:
		cache[item_name] = frappe.db.get_value("Item", item_name, "item_group") or ""
	return cache[item_name]


def _is_schedulable_item(item_code: str | None) -> bool:
	return _get_item_group(item_code) in SCHEDULABLE_ITEM_GROUPS


def _schedule_row_key(row: dict[str, Any]) -> tuple:
	return (
		row.get("sales_order") or "",
		row.get("item_code") or "",
		str(getdate(row.get("schedule_date"))),
		row.get("customer_part_no") or "",
	)


def _schedule_identity_key(row: dict[str, Any]) -> tuple:
	return (
		row.get("sales_order") or "",
		row.get("item_code") or "",
		row.get("customer_part_no") or "",
	)


def _detect_change_type(previous: dict[str, Any], current: dict[str, Any]) -> str:
	if not previous and flt(current.get("qty")) > 0:
		return "Added"
	if previous and flt(current.get("qty")) <= 0:
		return "Cancelled"
	if previous and getdate(current.get("schedule_date")) < getdate(previous.get("schedule_date")):
		return "Advanced"
	if previous and getdate(current.get("schedule_date")) > getdate(previous.get("schedule_date")):
		return "Delayed"
	if flt(current.get("qty")) > flt(previous.get("qty")):
		return "Increased"
	if flt(current.get("qty")) < flt(previous.get("qty")):
		return "Reduced"
	return "Unchanged"


def _summarize_change_types(rows: list[dict[str, Any]]) -> dict[str, int]:
	summary = defaultdict(int)
	for row in rows:
		summary[row.get("change_type") or "Unknown"] += 1
	return dict(summary)


def _build_demand_row(
	company: str,
	customer: str | None,
	item_code: str,
	demand_source: str,
	demand_date,
	qty: float,
	source_doctype: str,
	source_name: str,
	sales_order: str | None = None,
	remark: str | None = None,
	customer_part_no: str | None = None,
	is_urgent: int = 0,
) -> frappe.model.document.Document:
	settings = get_settings_dict()
	item_code = _require_item_name(item_code)
	if not _is_schedulable_item(item_code):
		raise frappe.ValidationError(
			_("Item {0} belongs to item group {1}. Injection APS only schedules {2}.").format(
				item_code,
				_get_item_group(item_code) or _("(blank)"),
				", ".join(SCHEDULABLE_ITEM_GROUPS),
			)
		)
	item_context = _get_item_context(item_code, settings)
	return frappe.get_doc(
		{
			"doctype": "APS Demand Pool",
			"company": company,
			"customer": customer,
			"sales_order": sales_order,
			"item_code": item_code,
			"customer_part_no": customer_part_no,
			"demand_source": demand_source,
			"demand_date": demand_date,
			"qty": qty,
			"status": "Open",
			"priority_score": _score_demand(
				demand_source=demand_source,
				demand_date=demand_date,
				is_urgent=is_urgent,
			),
			"is_urgent": is_urgent,
			"food_grade": item_context["food_grade"],
			"color_code": item_context["color_code"],
			"material_code": item_context["material_code"],
			"is_first_article": 1 if item_context["is_first_article"] else 0,
			"source_doctype": source_doctype,
			"source_name": source_name,
			"remark": remark,
			"is_system_generated": 1,
		}
	)


def _append_sales_order_backlog(
	company: str | None = None,
	warnings: list[dict[str, Any]] | None = None,
	warning_keys: set[tuple[str, str, str, str]] | None = None,
) -> dict[str, Any]:
	if not frappe.db.exists("DocType", "Sales Order Item"):
		return {"rows": [], "skipped_rows": 0}

	query = """
		select
			soi.name as sales_order_item_name,
			so.company,
			so.customer,
			soi.parent as sales_order,
			soi.item_code,
			soi.delivery_date,
			greatest(ifnull(soi.qty, 0) - ifnull(soi.delivered_qty, 0), 0) as open_qty
		from `tabSales Order Item` soi
		inner join `tabSales Order` so on so.name = soi.parent
		where so.docstatus = 1
			and ifnull(so.status, '') not in ('Closed', 'Completed', 'Cancelled')
			and greatest(ifnull(soi.qty, 0) - ifnull(soi.delivered_qty, 0), 0) > 0
	"""
	params = []
	if company:
		query += " and so.company = %s"
		params.append(company)

	rows = frappe.db.sql(query, params, as_dict=True)
	active_schedule_pairs = set()
	for schedule_row in frappe.db.sql(
		"""
		select cdsi.name, cdsi.parent, cdsi.sales_order, cdsi.item_code, cds.customer, cds.company
		from `tabCustomer Delivery Schedule Item` cdsi
		inner join `tabCustomer Delivery Schedule` cds on cds.name = cdsi.parent
		where cds.status = 'Active'
		""",
		as_dict=True,
	):
		resolved_item_code = _resolve_item_name(schedule_row.item_code)
		if not resolved_item_code:
			if warnings is not None and warning_keys is not None:
				_append_rebuild_warning(
					warnings,
					warning_keys,
					item_reference=schedule_row.item_code,
					source_doctype="Customer Delivery Schedule",
					source_name=schedule_row.parent,
					row_name=schedule_row.name,
				)
			continue
		if resolved_item_code != schedule_row.item_code:
			frappe.db.set_value(
				"Customer Delivery Schedule Item",
				schedule_row.name,
				"item_code",
				resolved_item_code,
				update_modified=False,
			)
		active_schedule_pairs.add((schedule_row.company, schedule_row.customer, resolved_item_code))
	created = []
	skipped_rows = 0
	for row in rows:
		resolved_item_code = _resolve_item_name(row.item_code)
		if not resolved_item_code:
			skipped_rows += 1
			if warnings is not None and warning_keys is not None:
				_append_rebuild_warning(
					warnings,
					warning_keys,
					item_reference=row.item_code,
					source_doctype="Sales Order",
					source_name=row.sales_order,
				)
			continue
		if not _is_schedulable_item(resolved_item_code):
			skipped_rows += 1
			if warnings is not None and warning_keys is not None:
				_append_item_group_warning(
					warnings,
					warning_keys,
					item_code=resolved_item_code,
					source_doctype="Sales Order",
					source_name=row.sales_order,
					item_group=_get_item_group(resolved_item_code),
				)
			continue
		if (row.company, row.customer, resolved_item_code) in active_schedule_pairs:
			continue
		demand = _build_demand_row(
			company=row.company,
			customer=row.customer,
			item_code=resolved_item_code,
			demand_source="Sales Order Backlog",
			demand_date=row.delivery_date or today(),
			qty=row.open_qty,
			source_doctype="Sales Order",
			source_name=row.sales_order,
			sales_order=row.sales_order,
		)
		created.append(demand.insert(ignore_permissions=True).name)
	return {"rows": created, "skipped_rows": skipped_rows}


def _append_safety_stock_demands(company: str | None = None) -> list[str]:
	settings = get_settings_dict()
	fieldname = settings["item_safety_stock_field"]
	if not fieldname or not frappe.db.exists("DocType", "Item"):
		return []
	item_meta = frappe.get_meta("Item")
	if not item_meta.has_field(fieldname):
		return []

	created = []
	stock_map = _get_available_stock_map(company)
	item_rows = frappe.get_all(
		"Item",
		filters={"disabled": 0},
		fields=["name", fieldname],
	)
	for item in item_rows:
		safety_stock = flt(item.get(fieldname))
		if not safety_stock:
			continue
		if not _is_schedulable_item(item.name):
			continue
		shortage = max(safety_stock - flt(stock_map.get(item.name)), 0)
		if shortage <= 0:
			continue
		demand = _build_demand_row(
			company=company or frappe.defaults.get_user_default("Company"),
			customer=None,
			item_code=item.name,
			demand_source="Safety Stock",
			demand_date=today(),
			qty=shortage,
			source_doctype="Item",
			source_name=item.name,
		)
		created.append(demand.insert(ignore_permissions=True).name)
	return created


def _score_demand(demand_source: str, demand_date, is_urgent: int = 0) -> int:
	days_to_due = (getdate(demand_date) - getdate(today())).days
	urgency_bonus = 250 if cint(is_urgent) else 0
	date_bonus = max(60 - max(days_to_due, -30), 0)
	return cint(DEMAND_SOURCE_PRIORITY.get(demand_source, 100) + urgency_bonus + date_bonus)


def _get_available_stock_map(company: str | None) -> dict[str, float]:
	if not frappe.db.exists("DocType", "Bin"):
		return {}
	query = """
		select
			bin.item_code,
			sum(ifnull(bin.actual_qty, 0) - ifnull(bin.reserved_qty, 0)) as available_qty
		from `tabBin` bin
		inner join `tabWarehouse` wh on wh.name = bin.warehouse
		where wh.is_group = 0
	"""
	params = []
	if company:
		query += " and wh.company = %s"
		params.append(company)
	query += " group by bin.item_code"
	return {row.item_code: flt(row.available_qty) for row in frappe.db.sql(query, params, as_dict=True)}


def _get_open_work_order_map(company: str | None) -> dict[str, float]:
	if not frappe.db.exists("DocType", "Work Order"):
		return {}
	query = """
		select
			production_item as item_code,
			sum(greatest(ifnull(qty, 0) - ifnull(produced_qty, 0), 0)) as open_qty
		from `tabWork Order`
		where docstatus = 1
			and ifnull(status, '') not in ('Completed', 'Closed', 'Cancelled')
	"""
	params = []
	if company:
		query += " and company = %s"
		params.append(company)
	query += " group by production_item"
	return {row.item_code: flt(row.open_qty) for row in frappe.db.sql(query, params, as_dict=True)}


def _build_net_requirement_reason(
	demand_qty: float,
	available_stock_qty: float,
	open_work_order_qty: float,
	safety_gap: float,
	overstock_qty: float,
	minimum_batch_qty: float,
	planning_qty: float,
) -> str:
	return _(
		"Demand {0} - allocated stock {1} - allocated open work orders {2} + one-time safety gap {3}; remaining overstock {4}; minimum batch {5}; planning qty {6}."
	).format(
		demand_qty,
		available_stock_qty,
		open_work_order_qty,
		safety_gap,
		overstock_qty,
		minimum_batch_qty,
		planning_qty,
	)


def _get_item_mapping_value(item_code: str, fieldname: str | None):
	if not fieldname or not frappe.db.exists("DocType", "Item") or not frappe.get_meta("Item").has_field(fieldname):
		return None
	item_code = _resolve_item_name(item_code)
	if not item_code:
		return None
	return frappe.db.get_value("Item", item_code, fieldname)


def _get_item_context(item_code: str, settings: dict[str, Any]) -> dict[str, Any]:
	item_code = _require_item_name(item_code)
	meta = frappe.get_meta("Item")
	item_doc = frappe.get_cached_doc("Item", item_code)
	food_grade = item_doc.get(settings["item_food_grade_field"]) if meta.has_field(settings["item_food_grade_field"]) else ""
	color_code = item_doc.get(settings["item_color_field"]) if meta.has_field(settings["item_color_field"]) else ""
	material_code = item_doc.get(settings["item_material_field"]) if meta.has_field(settings["item_material_field"]) else ""
	first_article = item_doc.get(settings["item_first_article_field"]) if meta.has_field(settings["item_first_article_field"]) else 0

	if (not color_code or not material_code) and frappe.db.exists("DocType", "Mold"):
		mold_row = _get_primary_mold_row(item_code)
		if mold_row and (not color_code or not material_code):
			material_row = frappe.db.sql(
				"""
				select material_item, color_spec
				from `tabMold Default Material`
				where parent = %s and parenttype = 'Mold'
				order by idx asc
				limit 1
				""",
				(mold_row.get("mold"),),
				as_dict=True,
			)
			if material_row:
				color_code = color_code or material_row[0].get("color_spec")
				material_code = material_code or material_row[0].get("material_item")

	return {
		"item_name": item_doc.item_name or "",
		"item_group": item_doc.item_group or "",
		"food_grade": food_grade or "",
		"color_code": color_code or "",
		"material_code": material_code or "",
		"is_first_article": cint(first_article),
		"is_urgent": 0,
	}


def _get_available_mold_rows(item_code: str) -> list[dict[str, Any]]:
	if not frappe.db.exists("DocType", "Mold"):
		return []
	item_code = _resolve_item_name(item_code)
	if not item_code:
		return []
	query = """
		select
			m.name as mold,
			m.mold_name,
			m.machine_tonnage,
			m.cavity_count,
			m.standard_cycle_seconds,
			m.status as mold_status,
			m.is_family_mold,
			mp.priority,
			mp.is_default_product,
			mp.item_code,
			mp.output_qty,
			mp.cavity_output_qty
		from `tabMold` m
		inner join `tabMold Product` mp on mp.parent = m.name and mp.parenttype = 'Mold'
		where m.docstatus = 1
			and mp.item_code = %s
			and ifnull(m.status, '') not in ({0})
		order by mp.is_default_product desc, mp.priority asc, m.modified desc
	""".format(", ".join(["%s"] * len(BLOCKING_MOLD_STATUSES)))
	params = [item_code, *BLOCKING_MOLD_STATUSES]
	rows = []
	seen_molds = set()
	for row in frappe.db.sql(query, params, as_dict=True):
		if row.get("mold") in seen_molds:
			continue
		seen_molds.add(row.get("mold"))
		row["cycle_time_seconds"] = flt(row.get("standard_cycle_seconds"))
		row["effective_output_qty"] = _get_effective_mold_output_qty(row)
		rows.append(row)
	return rows


def _get_effective_mold_output_qty(mold_row: dict[str, Any]) -> float:
	if flt(mold_row.get("cavity_output_qty")) > 0:
		return flt(mold_row.get("cavity_output_qty"))
	if flt(mold_row.get("output_qty")) > 0:
		return flt(mold_row.get("output_qty"))
	if 0 < flt(mold_row.get("cavity_count")) <= 128:
		return flt(mold_row.get("cavity_count"))
	return 1


def _get_primary_mold_row(item_code: str) -> dict[str, Any] | None:
	rows = _get_available_mold_rows(item_code)
	return rows[0] if rows else None


def _get_preferred_mold_row(item_code: str) -> dict[str, Any] | None:
	return _get_primary_mold_row(item_code)


def _get_family_output_rows(mold_name: str, primary_item_code: str) -> list[dict[str, Any]]:
	if not mold_name or not frappe.db.exists("DocType", "Mold Product"):
		return []
	primary_item_code = _resolve_item_name(primary_item_code)
	query = """
		select
			mp.item_code,
			mp.output_qty,
			mp.cavity_output_qty
		from `tabMold Product` mp
		inner join `tabItem` item on item.name = mp.item_code
		where mp.parent = %s
			and mp.parenttype = 'Mold'
			and mp.item_code != %s
			and item.item_group in ({0})
		order by mp.priority asc, mp.idx asc
	""".format(", ".join(["%s"] * len(SCHEDULABLE_ITEM_GROUPS)))
	params = [mold_name, primary_item_code, *SCHEDULABLE_ITEM_GROUPS]
	return frappe.db.sql(query, params, as_dict=True)


def _get_primary_demand_source(item_code: str, customer: str | None, demand_date) -> str:
	item_code = _resolve_item_name(item_code) or item_code
	row = frappe.get_all(
		"APS Demand Pool",
		filters={
			"item_code": item_code,
			"customer": customer,
			"demand_date": demand_date,
			"status": ("!=", "Cancelled"),
		},
		fields=["demand_source", "priority_score"],
		order_by="priority_score desc, modified asc",
		limit=1,
	)
	return row[0].get("demand_source") if row else ""


def _build_form_route(doctype: str | None, docname: str | None) -> str:
	if not doctype or not docname:
		return ""
	return f"Form/{doctype}/{docname}"


def _get_item_detail_snapshot(item_code: str, customer: str | None, settings: dict[str, Any]) -> dict[str, Any]:
	item_code = _require_item_name(item_code)
	item_doc = frappe.get_cached_doc("Item", item_code)
	meta = frappe.get_meta("Item")
	customer_reference = ""
	customer_reference_field = ""
	for fieldname in ("customer_code", "default_manufacturer_part_no", "custom_part_information"):
		if meta.has_field(fieldname) and item_doc.get(fieldname):
			customer_reference = item_doc.get(fieldname)
			customer_reference_field = fieldname
			break
	drawing_file = ""
	for fieldname in ("drawing_file", "sec_drawing_file"):
		if meta.has_field(fieldname) and item_doc.get(fieldname):
			drawing_file = item_doc.get(fieldname)
			break
	context = _get_item_context(item_code, settings)
	return {
		"item_code": item_code,
		"item_name": item_doc.item_name or "",
		"item_group": item_doc.item_group or "",
		"stock_uom": item_doc.stock_uom or "",
		"customer": customer or item_doc.get("customer") or "",
		"customer_reference": customer_reference or "",
		"customer_reference_field": customer_reference_field or "",
		"drawing_file": drawing_file or "",
		"food_grade": context.get("food_grade") or "",
		"color_code": context.get("color_code") or "",
		"material_code": context.get("material_code") or "",
		"is_first_article": cint(context.get("is_first_article")),
		"item_route": _build_form_route("Item", item_code),
	}


def _get_result_source_rows(result) -> list[dict[str, Any]]:
	if not result.net_requirement or not frappe.db.exists("DocType", "APS Demand Pool"):
		return []
	rows = frappe.get_all(
		"APS Demand Pool",
		filters={
			"company": result.company,
			"customer": result.customer,
			"item_code": result.item_code,
			"demand_date": result.requested_date,
			"status": ("!=", "Cancelled"),
		},
		fields=[
			"name",
			"demand_source",
			"demand_date",
			"qty",
			"customer_part_no",
			"sales_order",
			"source_doctype",
			"source_name",
			"remark",
		],
		order_by="priority_score desc, modified asc",
	)
	for row in rows:
		row["source_route"] = _build_form_route(row.get("source_doctype"), row.get("source_name"))
		row["sales_order_route"] = _build_form_route("Sales Order", row.get("sales_order"))
		row["demand_pool_route"] = _build_form_route("APS Demand Pool", row.get("name"))
	return rows


def _get_result_exception_rows(result) -> list[dict[str, Any]]:
	rows = frappe.get_all(
		"APS Exception Log",
		filters={"planning_run": result.planning_run, "status": "Open"},
		fields=[
			"name",
			"severity",
			"exception_type",
			"message",
			"is_blocking",
			"source_doctype",
			"source_name",
			"resolution_hint",
			"workstation",
			"diagnostic_json",
		],
		order_by="modified desc",
	)
	relevant = []
	relevant_names = {result.name, result.net_requirement}
	for row in rows:
		if row.get("source_name") in relevant_names:
			diagnostic = _parse_diagnostic_json(row.get("diagnostic_json"))
			row["diagnostic"] = diagnostic
			row["root_cause_text"] = diagnostic.get("root_cause_text") or ""
			row["suggested_actions"] = diagnostic.get("suggested_actions") or []
			row["source_route"] = _build_form_route(row.get("source_doctype"), row.get("source_name"))
			relevant.append(row)
	return relevant


def _build_exception_resolution_context(doc) -> dict[str, Any]:
	diagnostic = _parse_diagnostic_json(doc.get("diagnostic_json"))
	root_cause_text = diagnostic.get("root_cause_text") or doc.get("resolution_hint") or doc.get("message") or ""
	suggested_actions = diagnostic.get("suggested_actions") or ([doc.get("resolution_hint")] if doc.get("resolution_hint") else [])
	related_routes = {
		"source": _build_form_route(doc.get("source_doctype"), doc.get("source_name")),
		"item": _build_form_route("Item", doc.get("item_code")),
		"workstation": _build_form_route("Workstation", doc.get("workstation")),
		"gantt": f"aps-schedule-gantt?run_name={doc.get('planning_run')}" if doc.get("planning_run") else "",
		"execution": f"aps-release-center?run_name={doc.get('planning_run')}" if doc.get("planning_run") else "",
	}
	return {
		"name": doc.name,
		"planning_run": doc.get("planning_run"),
		"severity": doc.get("severity"),
		"exception_type": doc.get("exception_type"),
		"item_code": doc.get("item_code"),
		"customer": doc.get("customer"),
		"workstation": doc.get("workstation"),
		"message": doc.get("message"),
		"resolution_hint": doc.get("resolution_hint"),
		"is_blocking": cint(doc.get("is_blocking")),
		"source_doctype": doc.get("source_doctype"),
		"source_name": doc.get("source_name"),
		"diagnostic": diagnostic,
		"root_cause_codes": diagnostic.get("root_cause_codes") or [],
		"root_cause_text": root_cause_text,
		"suggested_actions": [row for row in suggested_actions if row],
		"related_routes": related_routes,
		"gantt_focus": {
			"run_name": doc.get("planning_run"),
			"item_code": doc.get("item_code"),
			"workstation": doc.get("workstation"),
			"exception_name": doc.name,
		},
	}


def _get_result_mold_rows(result, segments: list[dict[str, Any]]) -> list[dict[str, Any]]:
	mold_names = list(dict.fromkeys(row.get("mould_reference") for row in segments if row.get("mould_reference")))
	if not mold_names or not frappe.db.exists("DocType", "Mold"):
		return []
	rows = frappe.db.sql(
		"""
		select
			m.name as mold,
			m.mold_name,
			m.status as mold_status,
			m.machine_tonnage,
			m.cavity_count,
			m.standard_cycle_seconds,
			m.is_family_mold,
			mp.item_code,
			mp.output_qty,
			mp.cavity_output_qty,
			mp.priority
		from `tabMold` m
		left join `tabMold Product` mp
			on mp.parent = m.name
			and mp.parenttype = 'Mold'
			and mp.item_code = %s
		where m.name in ({0})
		order by m.name asc
		""".format(", ".join(["%s"] * len(mold_names))),
		[result.item_code, *mold_names],
		as_dict=True,
	)
	for row in rows:
		row["cycle_time_seconds"] = flt(row.get("standard_cycle_seconds"))
		row["cavity_count"] = flt(row.get("cavity_output_qty") or row.get("output_qty") or row.get("cavity_count"))
		row["effective_output_qty"] = _get_effective_mold_output_qty(row)
		row["mold_route"] = _build_form_route("Mold", row.get("mold"))
	return rows


def _get_mold_master_rows_for_item(item_code: str) -> list[dict[str, Any]]:
	if not item_code or not frappe.db.exists("DocType", "Mold"):
		return []
	rows = frappe.db.sql(
		"""
		select
			m.name as mold,
			m.mold_name,
			m.status as mold_status,
			m.machine_tonnage,
			m.standard_cycle_seconds,
			m.is_family_mold,
			mp.item_code,
			mp.output_qty,
			mp.cavity_output_qty,
			mp.priority
		from `tabMold` m
		inner join `tabMold Product` mp
			on mp.parent = m.name
			and mp.parenttype = 'Mold'
			and mp.item_code = %s
		order by m.name asc, ifnull(mp.priority, 999) asc
		""",
		[item_code],
		as_dict=True,
	)
	for row in rows:
		row["cycle_time_seconds"] = flt(row.get("standard_cycle_seconds"))
		row["effective_output_qty"] = _get_effective_mold_output_qty(row)
		row["mold_route"] = _build_form_route("Mold", row.get("mold"))
	return rows


def _format_root_cause_text(lines: list[str]) -> str:
	return "\n".join(str(line).strip() for line in lines if str(line or "").strip())


def _build_mold_unavailable_diagnostic(item_code: str, selected_plant_floors: list[str] | None = None) -> dict[str, Any]:
	mold_rows = _get_mold_master_rows_for_item(item_code)
	root_cause_codes = []
	root_cause_lines = []
	suggested_actions = []
	if not mold_rows:
		root_cause_codes.append("MOLD_PRODUCT_MAPPING_MISSING")
		root_cause_lines.append("No Mold Product mapping was found for this item, so APS cannot determine an available mold.")
		suggested_actions.extend(
			[
				"Complete the Mold Product mapping for this item in Mold master data.",
				"Make sure the corresponding mold master is submitted and available.",
			]
		)
	else:
		root_cause_codes.append("NO_ACTIVE_MOLD")
		for row in mold_rows:
			root_cause_lines.append(
				"{0} is currently in status {1}.".format(row.get("mold"), row.get("mold_status") or "Status Not Maintained")
			)
		suggested_actions.extend(
			[
				"Check whether the mold is under maintenance, pending asset link, or scrapped.",
				"Make sure the Mold Product mapping, cycle, and output data are complete.",
			]
		)
	return {
		"selected_plant_floors": selected_plant_floors or [],
		"candidate_molds": [row.get("mold") for row in mold_rows if row.get("mold")],
		"candidate_workstations": [],
		"root_cause_codes": root_cause_codes,
		"root_cause_text": _format_root_cause_text(root_cause_lines),
		"suggested_actions": suggested_actions,
		"mold_rows": mold_rows,
	}


def _build_machine_unavailable_diagnostic(
	item_code: str,
	selected_plant_floors: list[str] | None = None,
	candidate_molds: list[str] | None = None,
) -> dict[str, Any]:
	root_cause_lines = [
		"No workstation in the selected plant floors satisfies tonnage, FDA, status, and mold-mapping constraints.",
	]
	suggested_actions = [
		"Check whether APS Machine Capability has been synchronized and enabled.",
		"Check workstation status, risk category, tonnage, and APS Mould-Machine Rule constraints.",
	]
	return {
		"selected_plant_floors": selected_plant_floors or [],
		"candidate_molds": candidate_molds or [],
		"candidate_workstations": [],
		"root_cause_codes": ["NO_ELIGIBLE_MACHINE_LANE"],
		"root_cause_text": _format_root_cause_text(root_cause_lines),
		"suggested_actions": suggested_actions,
	}


def _build_late_delivery_diagnostic(
	item_code: str,
	qty: float,
	scheduled_qty: float,
	unscheduled_qty: float,
	selected_plant_floors: list[str] | None,
	candidates: list[dict[str, Any]],
	selected_options: list[dict[str, Any]],
	blocking_exceptions: list[dict[str, Any]],
	horizon_end,
) -> dict[str, Any]:
	root_cause_codes = ["HORIZON_LIMIT"]
	root_cause_lines = [
		"The current planning horizon ends at {0}, and {1} is still unscheduled.".format(
			frappe.format(get_datetime(horizon_end), {"fieldtype": "Datetime"}),
			frappe.format(unscheduled_qty, {"fieldtype": "Float"}),
		)
	]
	candidate_molds = list(dict.fromkeys(row.get("mould_reference") for row in candidates if row.get("mould_reference")))
	candidate_workstations = list(dict.fromkeys(row.get("workstation") for row in candidates if row.get("workstation")))
	total_available_qty = sum(max(flt(row.get("available_qty")), 0) for row in selected_options or [])
	if total_available_qty < flt(qty):
		root_cause_codes.append("HORIZON_CAPACITY_INSUFFICIENT")
		root_cause_lines.append(
			"Available capacity inside the current horizon is about {0}, which is lower than the demand {1}.".format(
				frappe.format(total_available_qty, {"fieldtype": "Float"}),
				frappe.format(qty, {"fieldtype": "Float"}),
			)
		)
	if len(candidate_molds) <= 1:
		root_cause_codes.append("COPY_MOLD_LIMITED")
		root_cause_lines.append("The number of molds that can participate in scheduling is limited, so copy molds cannot further increase parallelization.")
	blocking_types = [row.get("exception_type") for row in blocking_exceptions if row.get("exception_type")]
	if blocking_types:
		root_cause_codes.append("CONSTRAINT_BLOCKED")
		root_cause_lines.append("Additional constraints blocked some candidate resources: {0}.".format(" / ".join(sorted(set(blocking_types))[:4])))
	if "FDA Conflict" in blocking_types:
		root_cause_codes.append("FDA_CONFLICT")
	if "Color Transition Blocked" in blocking_types:
		root_cause_codes.append("COLOR_BLOCKED")
	suggested_actions = [
		"Extend the APS horizon or split the demand into the next planning window.",
		"Check whether more copy molds, eligible workstations, or acceptable manual changeover plans are available.",
	]
	return {
		"requested_qty": flt(qty),
		"scheduled_qty": flt(scheduled_qty),
		"unscheduled_qty": flt(unscheduled_qty),
		"selected_plant_floors": selected_plant_floors or [],
		"candidate_molds": candidate_molds,
		"candidate_workstations": candidate_workstations,
		"root_cause_codes": list(dict.fromkeys(root_cause_codes)),
		"root_cause_text": _format_root_cause_text(root_cause_lines),
		"suggested_actions": suggested_actions,
	}


def _get_machine_capability_rows(plant_floors: list[str] | str | None) -> list[dict[str, Any]]:
	selected_plant_floors = _coerce_plant_floor_list(plant_floors=plant_floors)
	if not selected_plant_floors:
		return []
	rows = frappe.get_all(
		"APS Machine Capability",
		filters={"plant_floor": ("in", selected_plant_floors), "is_active": 1},
		fields=[
			"name",
			"workstation",
			"plant_floor",
			"machine_tonnage",
			"risk_category",
			"hourly_capacity_qty",
			"daily_capacity_qty",
			"queue_sequence",
			"machine_status",
			"max_run_hours",
		],
		order_by="queue_sequence asc, workstation asc",
	)
	return rows


def _build_workstation_state_map(capability_rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
	state = {}
	baseline_now = get_datetime(now_datetime())
	for row in capability_rows:
		state[row["workstation"]] = {
			"next_available": baseline_now,
			"last_color_code": "",
			"last_material_code": "",
			"last_mould_reference": "",
			"last_end_time": None,
			"anchor_item_code": "",
			"anchor_strength": 0,
			"anchor_source": "",
			"anchor_campaign_key": "",
			"capability": row,
		}
	return state


def _build_mold_state_map(locked_segments: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
	state = {}
	baseline_now = get_datetime(now_datetime())
	for row in locked_segments:
		mold_name = row.get("mould_reference")
		if not mold_name:
			continue
		current = state.setdefault(
			mold_name,
			{
				"next_available": baseline_now,
				"last_workstation": "",
				"last_end_time": None,
				"anchor_item_code": "",
				"anchor_strength": 0,
				"anchor_source": "",
				"anchor_campaign_key": "",
			},
		)
		end_time = get_datetime(row.get("end_time"))
		if end_time > current["next_available"]:
			current["next_available"] = end_time
			current["last_workstation"] = row.get("workstation") or ""
			current["last_end_time"] = end_time
			current["anchor_item_code"] = _normalize_item_code(row.get("primary_item_code"))
			current["anchor_strength"] = cint(row.get("anchor_strength") or ANCHOR_STRENGTH_LOCKED)
			current["anchor_source"] = row.get("execution_anchor_source") or "APS Locked Segment"
			current["anchor_campaign_key"] = row.get("campaign_key") or _build_campaign_key(
				row.get("primary_item_code"),
				row.get("mould_reference"),
				row.get("workstation"),
			)
	return state


def _get_execution_anchor_strength(status: str | None, row: dict[str, Any]) -> int:
	status = (status or "").strip()
	if status in FROZEN_SCHEDULING_STATUSES or row.get("from_time") or flt(row.get("completed_qty")) > 0:
		return ANCHOR_STRENGTH_HARD
	if status in ("Schedule Confirmed", ""):
		return ANCHOR_STRENGTH_RELEASED
	return ANCHOR_STRENGTH_SOFT


def _get_execution_anchor_rows(plant_floors: list[str] | str | None) -> list[dict[str, Any]]:
	selected_plant_floors = _coerce_plant_floor_list(plant_floors=plant_floors)
	if not selected_plant_floors or not frappe.db.exists("DocType", "Work Order Scheduling"):
		return []
	rows = frappe.db.sql(
		"""
		select
			si.name as scheduling_item,
			si.parent as work_order_scheduling,
			si.work_order,
			si.workstation,
			si.planned_start_date as start_time,
			si.planned_end_date as end_time,
			si.from_time,
			si.to_time,
			si.completed_qty,
			si.custom_aps_segment_reference,
			si.custom_aps_result_reference,
			si.custom_aps_run,
			wos.status as scheduling_status,
			wos.plant_floor,
			wo.production_item as item_code,
			seg.mould_reference,
			seg.color_code,
			seg.material_code,
			seg.campaign_key,
			seg.primary_item_code,
			seg.execution_anchor_source
		from `tabScheduling Item` si
		inner join `tabWork Order Scheduling` wos on wos.name = si.parent
		left join `tabWork Order` wo on wo.name = si.work_order
		left join `tabAPS Schedule Segment` seg on seg.name = si.custom_aps_segment_reference
		where wos.plant_floor in ({0})
			and ifnull(si.workstation, '') != ''
			and ifnull(wos.status, '') in ({1})
		order by
			case when ifnull(wos.status, '') in ('Material Transfer', 'Job Card', 'Manufacture') then 0 else 1 end,
			ifnull(si.from_time, si.planned_start_date) asc,
			ifnull(si.to_time, si.planned_end_date) asc
		""".format(", ".join(["%s"] * len(selected_plant_floors)), ", ".join(["%s"] * len(ACTIVE_SCHEDULING_STATUSES))),
		[*selected_plant_floors, *ACTIVE_SCHEDULING_STATUSES],
		as_dict=True,
	)
	anchor_rows = []
	for row in rows:
		start_time = row.get("from_time") or row.get("start_time")
		end_time = row.get("to_time") or row.get("end_time") or start_time
		if not start_time or not end_time:
			continue
		item_code = _normalize_item_code(row.get("primary_item_code") or row.get("item_code"))
		campaign_key = row.get("campaign_key") or _build_campaign_key(
			item_code,
			row.get("mould_reference"),
			row.get("workstation"),
		)
		anchor_rows.append(
			{
				"name": row.get("scheduling_item"),
				"work_order_scheduling": row.get("work_order_scheduling"),
				"work_order": row.get("work_order"),
				"workstation": row.get("workstation"),
				"plant_floor": row.get("plant_floor"),
				"start_time": start_time,
				"end_time": end_time,
				"planned_qty": flt(row.get("completed_qty") or 0),
				"color_code": row.get("color_code") or "",
				"material_code": row.get("material_code") or "",
				"mould_reference": row.get("mould_reference") or "",
				"primary_item_code": item_code,
				"campaign_key": campaign_key,
				"anchor_strength": _get_execution_anchor_strength(row.get("scheduling_status"), row),
				"execution_anchor_source": row.get("execution_anchor_source")
				or (row.get("scheduling_status") or "Scheduling Item"),
			}
		)
	return anchor_rows


def _apply_anchor_rows_to_state(
	workstation_state: dict[str, dict[str, Any]],
	mold_state: dict[str, dict[str, Any]],
	anchor_rows: list[dict[str, Any]],
):
	for row in anchor_rows:
		workstation = row.get("workstation")
		end_time = get_datetime(row.get("end_time"))
		if workstation in workstation_state:
			state = workstation_state[workstation]
			if end_time >= state["next_available"]:
				state["next_available"] = end_time
				state["last_color_code"] = row.get("color_code") or ""
				state["last_material_code"] = row.get("material_code") or ""
				state["last_mould_reference"] = row.get("mould_reference") or ""
				state["last_end_time"] = end_time
				state["anchor_item_code"] = _normalize_item_code(row.get("primary_item_code"))
				state["anchor_strength"] = cint(row.get("anchor_strength") or 0)
				state["anchor_source"] = row.get("execution_anchor_source") or ""
				state["anchor_campaign_key"] = row.get("campaign_key") or ""
		mold_name = row.get("mould_reference")
		if not mold_name:
			continue
		current = mold_state.setdefault(
			mold_name,
			{
				"next_available": get_datetime(now_datetime()),
				"last_workstation": "",
				"last_end_time": None,
				"anchor_item_code": "",
				"anchor_strength": 0,
				"anchor_source": "",
				"anchor_campaign_key": "",
			},
		)
		if end_time >= current["next_available"]:
			current["next_available"] = end_time
			current["last_workstation"] = workstation or ""
			current["last_end_time"] = end_time
			current["anchor_item_code"] = _normalize_item_code(row.get("primary_item_code"))
			current["anchor_strength"] = cint(row.get("anchor_strength") or 0)
			current["anchor_source"] = row.get("execution_anchor_source") or ""
			current["anchor_campaign_key"] = row.get("campaign_key") or ""


def _get_locked_segments(plant_floors: list[str] | str | None) -> list[dict[str, Any]]:
	if not frappe.db.exists("DocType", "APS Schedule Segment"):
		return []
	selected_plant_floors = _coerce_plant_floor_list(plant_floors=plant_floors)
	if not selected_plant_floors:
		return []
	return frappe.get_all(
		"APS Schedule Segment",
		filters={
			"segment_status": ("in", LOCKED_SEGMENT_STATUSES),
			"is_locked": 1,
			"plant_floor": ("in", selected_plant_floors),
		},
		fields=[
			"name",
			"workstation",
			"start_time",
			"end_time",
			"planned_qty",
			"color_code",
			"material_code",
			"mould_reference",
			"primary_item_code",
			"campaign_key",
			"anchor_strength",
			"execution_anchor_source",
		],
	)


def _apply_locked_segments_to_state(
	workstation_state: dict[str, dict[str, Any]],
	locked_segments: list[dict[str, Any]],
):
	for row in locked_segments:
		state = workstation_state.get(row.get("workstation"))
		if not state:
			continue
		end_time = get_datetime(row.get("end_time"))
		if end_time > state["next_available"]:
			state["next_available"] = end_time
			state["last_color_code"] = row.get("color_code") or ""
			state["last_material_code"] = row.get("material_code") or ""
			state["last_mould_reference"] = row.get("mould_reference") or ""
			state["last_end_time"] = end_time
			state["anchor_item_code"] = _normalize_item_code(row.get("primary_item_code"))
			state["anchor_strength"] = cint(row.get("anchor_strength") or ANCHOR_STRENGTH_LOCKED)
			state["anchor_source"] = row.get("execution_anchor_source") or "APS Locked Segment"
			state["anchor_campaign_key"] = row.get("campaign_key") or _build_campaign_key(
				row.get("primary_item_code"),
				row.get("mould_reference"),
				row.get("workstation"),
			)


def _select_machine_candidates(
	item_code: str,
	item_context: dict[str, Any],
	capability_rows: list[dict[str, Any]],
	plant_floors: list[str] | str | None,
) -> list[dict[str, Any]]:
	mold_rows = _get_available_mold_rows(item_code)
	if not mold_rows:
		return []
	selected_plant_floors = _coerce_plant_floor_list(plant_floors=plant_floors)

	rules = frappe.get_all(
		"APS Mould-Machine Rule",
		filters=_strip_none({"item_code": item_code, "is_active": 1}),
		fields=["workstation", "priority", "preferred", "mould_reference", "min_tonnage", "max_tonnage"],
		order_by="preferred desc, priority asc",
	)
	rule_map = defaultdict(list)
	for row in rules:
		rule_map[row.workstation].append(row)

	candidates = []
	for capability in capability_rows:
		if selected_plant_floors and capability.get("plant_floor") not in selected_plant_floors:
			continue
		if capability.get("machine_status") not in APS_ALLOWED_MACHINE_STATUSES:
			continue
		workstation_rules = rule_map.get(capability.get("workstation")) or []
		for mold_row in mold_rows:
			if capability.get("machine_tonnage") and mold_row.get("machine_tonnage"):
				if flt(capability.get("machine_tonnage")) < flt(mold_row.get("machine_tonnage")):
					continue

			rule = _match_rule_for_candidate(workstation_rules, capability, mold_row)
			if workstation_rules and not rule:
				continue

			candidate = dict(capability)
			candidate["preferred"] = cint(rule.get("preferred")) if rule else 0
			candidate["priority"] = cint(rule.get("priority")) if rule else cint(capability.get("queue_sequence") or 999)
			candidate["mould_reference"] = mold_row.get("mold")
			candidate["mold_name"] = mold_row.get("mold_name")
			candidate["cavity_count"] = flt(
				mold_row.get("cavity_output_qty") or mold_row.get("output_qty") or mold_row.get("cavity_count")
			)
			candidate["cycle_time_seconds"] = flt(mold_row.get("cycle_time_seconds"))
			candidate["output_qty"] = flt(mold_row.get("output_qty"))
			candidate["cavity_output_qty"] = flt(mold_row.get("cavity_output_qty"))
			candidate["effective_output_qty"] = flt(mold_row.get("effective_output_qty"))
			candidate["is_family_mold"] = cint(mold_row.get("is_family_mold"))
			candidate["lane_key"] = f"{mold_row.get('mold')}::{capability.get('workstation')}"
			candidate["mold_priority"] = cint(mold_row.get("priority") or 999)
			candidate["default_product"] = cint(mold_row.get("is_default_product"))
			required_tonnage = flt(mold_row.get("machine_tonnage"))
			candidate_tonnage = flt(capability.get("machine_tonnage"))
			candidate["required_tonnage"] = required_tonnage
			candidate["tonnage_gap"] = max(candidate_tonnage - required_tonnage, 0) if candidate_tonnage and required_tonnage else 999999
			candidates.append(candidate)

	return sorted(
		candidates,
		key=lambda row: (
			-cint(row.get("preferred")),
			flt(row.get("tonnage_gap")) if row.get("tonnage_gap") is not None else 999999,
			cint(row.get("priority") or 999),
			cint(row.get("mold_priority") or 999),
			-cint(row.get("default_product")),
			row.get("mould_reference") or "",
			row.get("workstation") or "",
		),
	)


def _match_rule_for_candidate(
	workstation_rules: list[dict[str, Any]],
	capability: dict[str, Any],
	mold_row: dict[str, Any],
) -> dict[str, Any] | None:
	if not workstation_rules:
		return None

	matches = []
	for rule in workstation_rules:
		if rule.get("mould_reference") and rule.get("mould_reference") != mold_row.get("mold"):
			continue
		if rule.get("min_tonnage") and capability.get("machine_tonnage"):
			if flt(capability.get("machine_tonnage")) < flt(rule.get("min_tonnage")):
				continue
		if rule.get("max_tonnage") and capability.get("machine_tonnage"):
			if flt(capability.get("machine_tonnage")) > flt(rule.get("max_tonnage")):
				continue
		matches.append(rule)

	if not matches:
		return None

	matches.sort(key=lambda row: (-cint(row.get("preferred")), cint(row.get("priority") or 999)))
	return matches[0]


def _choose_best_slot(
	item_code: str,
	item_context: dict[str, Any],
	qty: float,
	demand_date,
	horizon_start,
	horizon_end,
	workstation_state: dict[str, dict[str, Any]],
	mold_state: dict[str, dict[str, Any]],
	candidates: list[dict[str, Any]],
	settings: dict[str, Any],
	selected_plant_floors: list[str] | None = None,
) -> dict[str, Any]:
	if not _get_available_mold_rows(item_code):
		diagnostic = _build_mold_unavailable_diagnostic(item_code, selected_plant_floors=selected_plant_floors)
		return {
			"scheduled_qty": 0,
			"unscheduled_qty": qty,
			"result_status": "Blocked",
			"risk_status": "Blocked",
			"segments": [],
			"selected_moulds": [],
			"schedule_explanation": _("No active mold is available for {0}.").format(item_code),
			"family_side_outputs": [],
			"exceptions": [
				{
					"severity": "Critical",
					"exception_type": "Mold Unavailable",
					"message": _("No active mold is available for {0}.").format(item_code),
					"resolution_hint": _("Check Mold status, Mold Product mapping and submitted mold master data."),
					"is_blocking": 1,
					"diagnostic": diagnostic,
				}
			],
		}

	if not candidates:
		diagnostic = _build_machine_unavailable_diagnostic(
			item_code,
			selected_plant_floors=selected_plant_floors,
			candidate_molds=[row.get("mold") for row in _get_mold_master_rows_for_item(item_code)],
		)
		return {
			"scheduled_qty": 0,
			"unscheduled_qty": qty,
			"result_status": "Blocked",
			"risk_status": "Blocked",
			"segments": [],
			"selected_moulds": [],
			"schedule_explanation": _("No eligible machine lane is available for {0}.").format(item_code),
			"family_side_outputs": [],
			"exceptions": [
				{
					"severity": "Critical",
					"exception_type": "Machine Unavailable",
					"message": _("No eligible APS machine capability rows were found for {0}.").format(item_code),
					"resolution_hint": _("Maintain APS Machine Capability or relax mould-machine constraints."),
					"is_blocking": 1,
					"diagnostic": diagnostic,
				}
			],
		}

	proposals = []
	blocking_exceptions = []
	for candidate in candidates:
		proposal = _build_candidate_proposal(
			item_code=item_code,
			item_context=item_context,
			qty=qty,
			candidate=candidate,
			horizon_start=horizon_start,
			horizon_end=horizon_end,
			workstation_state=workstation_state,
			mold_state=mold_state,
			settings=settings,
		)
		if proposal.get("is_blocked"):
			blocking_exceptions.extend(proposal.get("exceptions") or [])
			continue
		proposals.append(proposal)

	if not proposals:
		return {
			"scheduled_qty": 0,
			"unscheduled_qty": qty,
			"result_status": "Blocked",
			"risk_status": "Blocked",
			"segments": [],
			"selected_moulds": [],
			"schedule_explanation": _("All candidate lanes for {0} were blocked by mold, FDA or setup constraints.").format(item_code),
			"family_side_outputs": [],
			"exceptions": blocking_exceptions,
		}

	unique_options = []
	used_workstations = set()
	used_moulds = set()
	for proposal in sorted(proposals, key=lambda row: row["score"]):
		if proposal["workstation"] in used_workstations:
			continue
		if proposal["mould_reference"] in used_moulds:
			continue
		used_workstations.add(proposal["workstation"])
		used_moulds.add(proposal["mould_reference"])
		unique_options.append(proposal)

	if not unique_options:
		return {
			"scheduled_qty": 0,
			"unscheduled_qty": qty,
			"result_status": "Blocked",
			"risk_status": "Blocked",
			"segments": [],
			"selected_moulds": [],
			"schedule_explanation": _("No unique mold-machine lanes remain for {0}.").format(item_code),
			"family_side_outputs": [],
			"exceptions": blocking_exceptions,
		}

	due_datetime = _get_due_datetime(demand_date)
	primary = unique_options[0]
	use_parallel = (
		len(unique_options) > 1
		and flt(qty) >= flt(settings.get("minimum_parallel_split_qty") or 0)
		and (primary["available_qty"] < flt(qty) or primary["end_time_full_qty"] > due_datetime)
	)
	selected_options = unique_options if use_parallel else [primary]
	if not use_parallel and primary["available_qty"] < flt(qty) and len(unique_options) > 1:
		selected_options = unique_options
		use_parallel = True

	parallel_group = f"PAR-{frappe.generate_hash(length=8)}" if use_parallel else ""
	selected_segments = []
	exceptions = list(blocking_exceptions)
	remaining = flt(qty)
	total_changeover = 0

	for sequence_no, proposal in enumerate(selected_options, start=1):
		if remaining <= 0:
			break
		allocatable_qty = min(remaining, proposal["available_qty"] if use_parallel else max(proposal["available_qty"], remaining))
		if not use_parallel and proposal["available_qty"] < remaining:
			allocatable_qty = proposal["available_qty"]
		if allocatable_qty <= 0:
			continue

		run_hours = max(allocatable_qty / max(proposal["hourly_capacity_qty"], 1), 0)
		end_time = proposal["start_time"] + timedelta(hours=run_hours)
		segment_status = "Planned"
		if end_time > get_datetime(horizon_end):
			end_time = get_datetime(horizon_end)
		segment = {
			"workstation": proposal["workstation"],
			"plant_floor": proposal.get("plant_floor"),
			"start_time": proposal["start_time"],
			"end_time": end_time,
			"planned_qty": allocatable_qty,
			"sequence_no": sequence_no,
			"lane_key": proposal["lane_key"],
			"campaign_key": proposal.get("campaign_key") or _build_campaign_key(item_code, proposal.get("mould_reference"), proposal.get("workstation")),
			"parallel_group": parallel_group,
			"family_group": "",
			"segment_kind": "Primary",
			"primary_item_code": item_code,
			"co_product_item_code": "",
			"setup_minutes": proposal["setup_minutes"],
			"changeover_minutes": proposal["setup_minutes"],
			"mould_reference": proposal["mould_reference"],
			"schedule_explanation": proposal["schedule_explanation"],
			"manual_change_note": "",
			"risk_flags": "\n".join(sorted({row.get("exception_type") for row in proposal["exceptions"] if row.get("exception_type")})),
			"segment_status": segment_status,
			"anchor_strength": proposal.get("anchor_strength") or 0,
			"execution_anchor_source": proposal.get("execution_anchor_source") or "",
			"color_code": item_context.get("color_code"),
			"material_code": item_context.get("material_code"),
			"is_locked": 0,
			"is_manual": 0,
			"_output_qty": proposal["output_qty"],
			"_is_family_mold": proposal["is_family_mold"],
		}
		selected_segments.append(segment)
		exceptions.extend(proposal["exceptions"])
		total_changeover += flt(proposal["setup_minutes"])
		remaining -= allocatable_qty

	if not selected_segments:
		return {
			"scheduled_qty": 0,
			"unscheduled_qty": qty,
			"result_status": "Blocked",
			"risk_status": "Blocked",
			"segments": [],
			"selected_moulds": [],
			"schedule_explanation": _("No segment could be placed inside the current planning horizon for {0}.").format(item_code),
			"family_side_outputs": [],
			"exceptions": exceptions,
		}

	scheduled_qty = sum(flt(segment["planned_qty"]) for segment in selected_segments)
	unscheduled_qty = max(flt(qty) - scheduled_qty, 0)
	risk_status = "Normal"
	result_status = "Planned"
	if any(get_datetime(segment["end_time"]) > due_datetime for segment in selected_segments):
		risk_status = "Attention"
		result_status = "Risk"
	if unscheduled_qty > 0:
		late_delivery_diagnostic = _build_late_delivery_diagnostic(
			item_code=item_code,
			qty=qty,
			scheduled_qty=scheduled_qty,
			unscheduled_qty=unscheduled_qty,
			selected_plant_floors=selected_plant_floors,
			candidates=candidates,
			selected_options=selected_options,
			blocking_exceptions=blocking_exceptions,
			horizon_end=horizon_end,
		)
		risk_status = "Critical"
		result_status = "Risk"
		exceptions.append(
			{
				"severity": "Critical",
				"exception_type": "Late Delivery Risk",
				"message": _("Only {0} of {1} can be scheduled inside the current horizon for {2}.").format(
					scheduled_qty,
					qty,
					item_code,
				),
				"resolution_hint": _("Extend the horizon, release additional copy molds or split the requirement."),
				"is_blocking": 0,
				"diagnostic": late_delivery_diagnostic,
			}
		)

	if use_parallel and len(selected_segments) > 1:
		exceptions.append(
			{
				"severity": "Warning",
				"exception_type": "Copy Mold Parallelized",
				"message": _("APS split {0} across {1} mold-machine lanes to protect delivery.").format(item_code, len(selected_segments)),
				"resolution_hint": _("Review the copy-mold split and lock the sequence if the shop agrees."),
				"is_blocking": 0,
			}
		)

	base_hourly_capacity = max(selected_options[0]["hourly_capacity_qty"], 1)
	minimum_window_qty = base_hourly_capacity * flt(settings.get("minimum_run_window_hours") or 0)
	if (
		scheduled_qty > 0
		and minimum_window_qty
		and scheduled_qty < minimum_window_qty
		and total_changeover >= flt(settings.get("mold_change_penalty_minutes") or 0)
	):
		future_hint = _get_future_demand_hint(item_code=item_code, demand_date=demand_date)
		exceptions.append(
			{
				"severity": "Warning",
				"exception_type": "Low Qty High Changeover Risk",
				"message": _(
					"{0} has a small run quantity {1} against changeover {2} minutes. {3}"
				).format(item_code, scheduled_qty, total_changeover, future_hint or ""),
				"resolution_hint": _("Consider batching this item with the next FC window if delivery promise allows."),
				"is_blocking": 0,
			}
		)

	for segment in selected_segments:
		state = workstation_state.get(segment["workstation"])
		if not state:
			pass
		else:
			state["next_available"] = get_datetime(segment["end_time"])
			state["last_color_code"] = segment.get("color_code") or ""
			state["last_material_code"] = segment.get("material_code") or ""
			state["last_mould_reference"] = segment.get("mould_reference") or ""
			state["last_end_time"] = get_datetime(segment["end_time"])
		mold_name = segment.get("mould_reference")
		if mold_name:
			mold_state[mold_name] = {
				"next_available": get_datetime(segment["end_time"]),
				"last_workstation": segment.get("workstation") or "",
				"last_end_time": get_datetime(segment["end_time"]),
			}

	family_side_outputs, family_segments, family_summary = _build_family_side_outputs(
		item_code=item_code,
		primary_segments=selected_segments,
	)
	clean_segments = []
	for segment in selected_segments:
		clean_segment = dict(segment)
		clean_segment.pop("_output_qty", None)
		clean_segment.pop("_is_family_mold", None)
		clean_segments.append(clean_segment)
	clean_segments.extend(family_segments)
	selected_moulds = list(dict.fromkeys(segment.get("mould_reference") for segment in selected_segments if segment.get("mould_reference")))
	schedule_explanation = _(
		"Scheduled {0} on {1} lane(s); molds {2}; due {3}."
	).format(
		scheduled_qty,
		len(selected_segments),
		", ".join(selected_moulds) or _("(none)"),
		getdate(demand_date),
	)
	return {
		"scheduled_qty": scheduled_qty,
		"unscheduled_qty": unscheduled_qty,
		"result_status": result_status,
		"risk_status": risk_status,
		"segments": clean_segments,
		"selected_moulds": selected_moulds,
		"copy_mold_parallel": 1 if len(selected_moulds) > 1 else 0,
		"family_mold_result": 1 if family_side_outputs else 0,
		"primary_mould_reference": selected_moulds[0] if selected_moulds else "",
		"schedule_explanation": schedule_explanation,
		"family_side_outputs": family_side_outputs,
		"family_output_summary": family_summary,
		"exceptions": exceptions,
	}


def _build_candidate_proposal(
	item_code: str,
	item_context: dict[str, Any],
	qty: float,
	candidate: dict[str, Any],
	horizon_start,
	horizon_end,
	workstation_state: dict[str, dict[str, Any]],
	mold_state: dict[str, dict[str, Any]],
	settings: dict[str, Any],
) -> dict[str, Any]:
	state = workstation_state.get(candidate.get("workstation")) or {}
	mold_row = mold_state.get(candidate.get("mould_reference")) or {}
	normalized_item_code = _normalize_item_code(item_code)
	base_start = max(
		get_datetime(horizon_start),
		get_datetime(state.get("next_available") or horizon_start),
		get_datetime(mold_row.get("next_available") or horizon_start),
	)
	setup_minutes, candidate_exceptions, blocked = _estimate_setup_penalty(
		candidate=candidate,
		state=state,
		item_context=item_context,
		settings=settings,
	)
	if _has_fda_conflict(item_context, candidate):
		candidate_exceptions.append(
			{
				"severity": "Critical",
				"exception_type": "FDA Conflict",
				"message": _("Workstation {0} risk category {1} cannot run FDA requirement for {2}.").format(
					candidate.get("workstation"),
					candidate.get("risk_category") or "",
					item_code,
				),
				"workstation": candidate.get("workstation"),
				"resolution_hint": _("Select an FDA-capable workstation or change the risk mapping."),
				"is_blocking": 1,
			}
		)
		blocked = True

	start_time = base_start + timedelta(minutes=setup_minutes)
	capacity = _estimate_hourly_capacity(candidate=candidate, settings=settings)
	hourly_capacity_qty = capacity["hourly_capacity_qty"]
	if capacity.get("capacity_source") == "fallback_cycle":
		candidate_exceptions.append(
			{
				"severity": "Warning",
				"exception_type": "Missing Mold Cycle",
				"message": _("Mold {0} is missing a valid cycle time or output qty; APS used fallback cycle seconds.").format(
					candidate.get("mould_reference")
				),
				"workstation": candidate.get("workstation"),
				"resolution_hint": _("Maintain Mold.standard_cycle_seconds and Mold Product cavity / output values for more accurate scheduling."),
				"is_blocking": 0,
			}
		)

	available_hours = max((get_datetime(horizon_end) - start_time).total_seconds() / 3600, 0)
	if flt(candidate.get("max_run_hours")) > 0:
		available_hours = min(available_hours, flt(candidate.get("max_run_hours")))
	available_qty = max(available_hours * hourly_capacity_qty, 0)
	end_time_full_qty = start_time + timedelta(hours=_estimate_run_hours(qty=qty, candidate=candidate, settings=settings))
	anchor_strength = 0
	execution_anchor_source = ""
	continuity_rank = 3
	if state.get("anchor_item_code") == normalized_item_code and state.get("last_mould_reference") == candidate.get("mould_reference"):
		anchor_strength = max(anchor_strength, cint(state.get("anchor_strength") or 0))
		execution_anchor_source = state.get("anchor_source") or execution_anchor_source
		continuity_rank = 0 if state.get("anchor_campaign_key") else 1
	elif mold_row.get("anchor_item_code") == normalized_item_code and mold_row.get("last_workstation") == candidate.get("workstation"):
		anchor_strength = max(anchor_strength, cint(mold_row.get("anchor_strength") or 0))
		execution_anchor_source = mold_row.get("anchor_source") or execution_anchor_source
		continuity_rank = 1
	elif state.get("last_mould_reference") == candidate.get("mould_reference"):
		continuity_rank = 2
	campaign_key = (
		state.get("anchor_campaign_key")
		if state.get("anchor_item_code") == normalized_item_code
		and state.get("last_mould_reference") == candidate.get("mould_reference")
		and state.get("anchor_campaign_key")
		else _build_campaign_key(item_code, candidate.get("mould_reference"), candidate.get("workstation"))
	)
	schedule_explanation = _(
		"Mold {0} on {1}; setup {2} min; hourly capacity {3}."
	).format(
		candidate.get("mould_reference"),
		candidate.get("workstation"),
		setup_minutes,
		frappe.format(hourly_capacity_qty, {"fieldtype": "Float"}),
	)
	schedule_explanation += " " + _(
		"Cycle {0}s; cavities/output {1}/{2}."
	).format(
		frappe.format(candidate.get("cycle_time_seconds") or 0, {"fieldtype": "Float"}),
		frappe.format(candidate.get("cavity_count") or 0, {"fieldtype": "Float"}),
		frappe.format(candidate.get("effective_output_qty") or 0, {"fieldtype": "Float"}),
	)
	if anchor_strength:
		schedule_explanation += " " + _(
			"Continuation anchor {0} keeps the mold on the current machine."
		).format(execution_anchor_source or _("Execution"))
	immediately_available = 1 if base_start <= (get_datetime(horizon_start) + timedelta(minutes=1)) else 0
	return {
		"is_blocked": blocked,
		"workstation": candidate.get("workstation"),
		"plant_floor": candidate.get("plant_floor"),
		"mould_reference": candidate.get("mould_reference"),
		"lane_key": candidate.get("lane_key"),
		"is_family_mold": cint(candidate.get("is_family_mold")),
		"output_qty": flt(candidate.get("output_qty")),
		"effective_output_qty": flt(candidate.get("effective_output_qty")),
		"start_time": start_time,
		"setup_minutes": setup_minutes,
		"hourly_capacity_qty": hourly_capacity_qty,
		"available_qty": available_qty,
		"end_time_full_qty": end_time_full_qty,
		"schedule_explanation": schedule_explanation,
		"campaign_key": campaign_key,
		"anchor_strength": anchor_strength,
		"execution_anchor_source": execution_anchor_source,
		"exceptions": candidate_exceptions,
		"score": (
			continuity_rank,
			-anchor_strength,
			0 if immediately_available else 1,
			flt(candidate.get("tonnage_gap")) if candidate.get("tonnage_gap") is not None else 999999,
			end_time_full_qty,
			setup_minutes,
			-cint(candidate.get("preferred")),
			cint(candidate.get("priority") or 999),
			cint(candidate.get("mold_priority") or 999),
		),
	}


def _estimate_setup_penalty(candidate, state, item_context, settings):
	setup_minutes = flt(settings["default_setup_minutes"])
	exceptions = []
	is_blocked = False
	transition_rule = _get_color_transition_rule(state.get("last_color_code"), item_context.get("color_code"))
	if transition_rule:
		setup_minutes = max(setup_minutes, flt(transition_rule.get("setup_minutes") or setup_minutes))
		if cint(transition_rule.get("is_blocking")) or (transition_rule.get("change_level") or "") == "Blocked":
			exceptions.append(
				{
					"severity": "Critical",
					"exception_type": "Color Transition Blocked",
					"message": _("Color transition {0} -> {1} is configured as blocking.").format(
						state.get("last_color_code") or "-",
						item_context.get("color_code") or "-",
					),
					"workstation": candidate.get("workstation"),
					"resolution_hint": _("Choose another workstation or maintain a non-blocking color transition."),
					"is_blocking": 1,
				}
			)
			is_blocked = True
		elif transition_rule.get("penalty_score"):
			exceptions.append(
				{
					"severity": "Warning",
					"exception_type": "Color Transition",
					"message": _("Color transition {0} -> {1} has penalty {2}.").format(
						state.get("last_color_code") or "-",
						item_context.get("color_code") or "-",
						transition_rule.get("penalty_score"),
					),
					"workstation": candidate.get("workstation"),
					"resolution_hint": _("Group similar colors to reduce changeover cost."),
					"is_blocking": 0,
				}
			)

	if state.get("last_material_code") and state.get("last_material_code") != item_context.get("material_code"):
		setup_minutes += 15
		exceptions.append(
			{
				"severity": "Warning",
				"exception_type": "Material Changeover",
				"message": _("Material changeover is required on workstation {0}.").format(candidate.get("workstation")),
				"workstation": candidate.get("workstation"),
				"resolution_hint": _("Group the same material family where possible."),
				"is_blocking": 0,
			}
		)

	if cint(item_context.get("is_first_article")):
		setup_minutes += flt(settings["default_first_article_minutes"])
		exceptions.append(
			{
				"severity": "Warning",
				"exception_type": "First Article Confirmation",
				"message": _("First article confirmation time was added for {0}.").format(candidate.get("workstation")),
				"workstation": candidate.get("workstation"),
				"resolution_hint": _("Keep QA review slots visible in the short horizon."),
				"is_blocking": 0,
			}
		)

	if state.get("last_mould_reference") and candidate.get("mould_reference") and state.get("last_mould_reference") != candidate.get("mould_reference"):
		setup_minutes += flt(settings.get("mold_change_penalty_minutes") or 30)
		exceptions.append(
			{
				"severity": "Warning",
				"exception_type": "Mould Changeover",
				"message": _("Mold changeover is required on workstation {0}.").format(candidate.get("workstation")),
				"workstation": candidate.get("workstation"),
				"resolution_hint": _("Avoid short runs immediately after a mould change when future FC can be batched."),
				"is_blocking": 0,
			}
		)

	return setup_minutes, exceptions, is_blocked


def _estimate_hourly_capacity(candidate: dict[str, Any], settings: dict[str, Any]) -> dict[str, Any]:
	effective_output_qty = max(flt(candidate.get("effective_output_qty")), flt(candidate.get("output_qty")))
	if flt(candidate.get("cycle_time_seconds")) > 0 and effective_output_qty > 0:
		return {
			"hourly_capacity_qty": (3600 / flt(candidate.get("cycle_time_seconds"))) * effective_output_qty,
			"capacity_source": "mold_cycle",
		}

	if flt(candidate.get("hourly_capacity_qty")) > 0:
		return {
			"hourly_capacity_qty": flt(candidate.get("hourly_capacity_qty")),
			"capacity_source": "machine_hourly_fallback",
		}

	if flt(candidate.get("daily_capacity_qty")) > 0:
		return {
			"hourly_capacity_qty": flt(candidate.get("daily_capacity_qty")) / 24,
			"capacity_source": "machine_daily_fallback",
		}

	fallback_cycle_seconds = flt(settings.get("missing_cycle_fallback_seconds") or 0)
	if fallback_cycle_seconds > 0 and effective_output_qty > 0:
		return {
			"hourly_capacity_qty": (3600 / fallback_cycle_seconds) * effective_output_qty,
			"capacity_source": "fallback_cycle",
		}

	return {
		"hourly_capacity_qty": flt(settings["default_hourly_capacity_qty"]),
		"capacity_source": "default_hourly_fallback",
	}


def _build_capacity_display(candidate: dict[str, Any], settings: dict[str, Any]) -> dict[str, Any]:
	capacity = _estimate_hourly_capacity(candidate=candidate, settings=settings)
	hourly_capacity_qty = flt(capacity.get("hourly_capacity_qty"))
	explicit_daily_capacity_qty = flt(candidate.get("daily_capacity_qty"))
	daily_capacity_qty = (
		explicit_daily_capacity_qty
		if capacity.get("capacity_source") == "machine_daily_fallback" and explicit_daily_capacity_qty > 0
		else hourly_capacity_qty * 24
	)
	return {
		"hourly_capacity_qty": hourly_capacity_qty,
		"daily_capacity_qty": flt(daily_capacity_qty),
		"capacity_source": capacity.get("capacity_source") or "",
		"capacity_source_label": CAPACITY_SOURCE_LABELS.get(capacity.get("capacity_source") or "", "Unknown Capacity Basis"),
	}


def _estimate_run_hours(qty: float, candidate: dict[str, Any], settings: dict[str, Any]) -> float:
	hourly_capacity_qty = _estimate_hourly_capacity(candidate=candidate, settings=settings)["hourly_capacity_qty"]
	return max(flt(qty) / max(hourly_capacity_qty, 1), 0.25)


def _has_fda_conflict(item_context: dict[str, Any], candidate: dict[str, Any]) -> bool:
	food_grade_value = item_context.get("food_grade")
	food_grade = str(food_grade_value or "").upper()
	risk_category = (candidate.get("risk_category") or "").strip()
	requires_fda = cint(food_grade_value) or food_grade in ("YES", "TRUE", "1") or "FDA" in food_grade
	return bool(requires_fda) and risk_category == BLOCKING_WORKSTATION_RISK


def _get_due_datetime(demand_date) -> datetime:
	date_value = getdate(demand_date or today())
	return get_datetime(f"{date_value} 23:59:59")


def _get_future_demand_hint(item_code: str, demand_date) -> str:
	next_rows = frappe.get_all(
		"APS Demand Pool",
		filters={
			"item_code": item_code,
			"demand_date": (">", getdate(demand_date)),
			"status": ("!=", "Cancelled"),
		},
		fields=["demand_date", "qty", "demand_source"],
		order_by="demand_date asc",
		limit=1,
	)
	if not next_rows:
		return ""
	row = next_rows[0]
	return _("Next open demand is {0} qty on {1} ({2}).").format(
		row.get("qty"),
		row.get("demand_date"),
		row.get("demand_source") or _("Unknown"),
	)


def _build_family_side_outputs(
	item_code: str,
	primary_segments: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], str]:
	side_outputs = []
	side_segments = []
	for segment in primary_segments:
		if not cint(segment.get("_is_family_mold")) or not segment.get("mould_reference"):
			continue
		primary_output_qty = max(flt(segment.get("_output_qty")), 1)
		cycles = flt(segment.get("planned_qty")) / primary_output_qty
		family_group = f"FAM-{frappe.generate_hash(length=8)}"
		segment["family_group"] = family_group
		for sibling in _get_family_output_rows(segment.get("mould_reference"), item_code):
			sibling_output_qty = flt(sibling.get("cavity_output_qty") or sibling.get("output_qty"))
			side_qty = flt(cycles * sibling_output_qty)
			if side_qty <= 0:
				continue
			side_outputs.append(
				{
					"item_code": sibling.get("item_code"),
					"qty": side_qty,
					"source_item_code": item_code,
					"mould_reference": segment.get("mould_reference"),
					"workstation": segment.get("workstation"),
				}
			)
			side_segments.append(
				{
					"workstation": segment.get("workstation"),
					"plant_floor": segment.get("plant_floor"),
					"start_time": segment.get("start_time"),
					"end_time": segment.get("end_time"),
					"planned_qty": side_qty,
					"sequence_no": segment.get("sequence_no"),
					"lane_key": segment.get("lane_key"),
					"campaign_key": segment.get("campaign_key"),
					"parallel_group": segment.get("parallel_group"),
					"family_group": family_group,
					"segment_kind": "Family Co-Product",
					"primary_item_code": item_code,
					"co_product_item_code": sibling.get("item_code"),
					"setup_minutes": 0,
					"changeover_minutes": 0,
					"mould_reference": segment.get("mould_reference"),
					"schedule_explanation": _("Family Mold co-produces {0} together with {1}.").format(
						sibling.get("item_code"),
						item_code,
					),
					"manual_change_note": "",
					"risk_flags": "Family Co-Production",
					"segment_status": segment.get("segment_status"),
					"anchor_strength": segment.get("anchor_strength") or 0,
					"execution_anchor_source": segment.get("execution_anchor_source") or "",
					"color_code": segment.get("color_code"),
					"material_code": segment.get("material_code"),
					"is_locked": 0,
					"is_manual": 0,
				}
			)

	summary = ""
	if side_outputs:
		summary = "; ".join(
			f"{row['item_code']}={frappe.format(row['qty'], {'fieldtype': 'Float'})}@{row['mould_reference']}"
			for row in side_outputs
		)
	return side_outputs, side_segments, summary


def _get_color_transition_rule(from_color: str | None, to_color: str | None) -> dict[str, Any] | None:
	if not from_color or not to_color:
		return None
	rows = frappe.get_all(
		"APS Color Transition Rule",
		filters={"from_color": from_color, "to_color": to_color, "is_active": 1},
		fields=["change_level", "penalty_score", "setup_minutes", "is_blocking"],
		limit=1,
	)
	return rows[0] if rows else None


def _create_exception(
	planning_run: str,
	severity: str,
	exception_type: str,
	message: str,
	item_code: str | None = None,
	customer: str | None = None,
	workstation: str | None = None,
	source_doctype: str | None = None,
	source_name: str | None = None,
	resolution_hint: str | None = None,
	is_blocking: int = 0,
	diagnostic: dict[str, Any] | None = None,
	diagnostic_json: str | None = None,
):
	return frappe.get_doc(
		{
			"doctype": "APS Exception Log",
			"planning_run": planning_run,
			"severity": severity,
			"exception_type": exception_type,
			"message": message,
			"item_code": item_code,
			"customer": customer,
			"workstation": workstation,
			"source_doctype": source_doctype,
			"source_name": source_name,
			"resolution_hint": resolution_hint,
			"is_blocking": is_blocking,
			"diagnostic_json": _serialize_diagnostic_json(diagnostic=diagnostic, diagnostic_json=diagnostic_json),
			"status": "Open",
		}
	).insert(ignore_permissions=True)


def _ensure_open_exception(
	planning_run: str,
	severity: str,
	exception_type: str,
	message: str,
	item_code: str | None = None,
	customer: str | None = None,
	workstation: str | None = None,
	source_doctype: str | None = None,
	source_name: str | None = None,
	resolution_hint: str | None = None,
	is_blocking: int = 0,
	diagnostic: dict[str, Any] | None = None,
	diagnostic_json: str | None = None,
):
	existing = frappe.db.exists(
		"APS Exception Log",
		{
			"planning_run": planning_run,
			"exception_type": exception_type,
			"source_doctype": source_doctype,
			"source_name": source_name,
			"status": "Open",
		},
	)
	if existing:
		frappe.db.set_value(
			"APS Exception Log",
			existing,
			{
				"severity": severity,
				"message": message,
				"item_code": item_code,
				"customer": customer,
				"workstation": workstation,
				"resolution_hint": resolution_hint,
				"is_blocking": is_blocking,
				"diagnostic_json": _serialize_diagnostic_json(diagnostic=diagnostic, diagnostic_json=diagnostic_json),
			},
		)
		return frappe.get_doc("APS Exception Log", existing)
	return _create_exception(
		planning_run=planning_run,
		severity=severity,
		exception_type=exception_type,
		message=message,
		item_code=item_code,
		customer=customer,
		workstation=workstation,
		source_doctype=source_doctype,
		source_name=source_name,
		resolution_hint=resolution_hint,
		is_blocking=is_blocking,
		diagnostic=diagnostic,
		diagnostic_json=diagnostic_json,
	)


def validate_run_mold_readiness(run_name: str, persist_exceptions: bool = False) -> dict[str, Any]:
	rows = []
	exception_names = []
	for result in frappe.get_all(
		"APS Schedule Result",
		filters={"planning_run": run_name, "scheduled_qty": (">", 0)},
		fields=["name", "item_code", "customer", "primary_mould_reference"],
	):
		primary_segments = _get_primary_segments_for_result(result.name)
		selected_molds = list(
			dict.fromkeys(
				segment.get("mould_reference")
				for segment in primary_segments
				if segment.get("mould_reference")
			)
		)
		available_rows = {row.get("mold"): row for row in _get_available_mold_rows(result.item_code)}
		blockers = []
		if not available_rows:
			blockers.append(("Mold Master Missing", _("Item {0} has no approved Mold / Mold Product master data for APS scheduling.").format(result.item_code)))
		if not primary_segments:
			blockers.append(("Primary Segment Missing", _("Result {0} has no primary APS schedule segment.").format(result.name)))
		if not result.primary_mould_reference or not selected_molds:
			blockers.append(("Mold Reference Empty", _("Result {0} is missing a selected mold reference.").format(result.name)))
		for mold_name in selected_molds or ([result.primary_mould_reference] if result.primary_mould_reference else []):
			mold_row = available_rows.get(mold_name)
			if not mold_row:
				blockers.append(("Mold Product Missing", _("Mold {0} is not available from Mold Product master data for {1}.").format(mold_name, result.item_code)))
				continue
			if (mold_row.get("mold_status") or "") in BLOCKING_MOLD_STATUSES:
				blockers.append(("Mold Status Blocked", _("Mold {0} is currently {1}.").format(mold_name, mold_row.get("mold_status"))))
			if flt(mold_row.get("standard_cycle_seconds")) <= 0 or flt(mold_row.get("effective_output_qty")) <= 0:
				blockers.append(("Mold Cycle Missing", _("Mold {0} is missing cycle time or effective output qty.").format(mold_name)))

		for exception_type, message in blockers:
			rows.append(
				{
					"result_name": result.name,
					"item_code": result.item_code,
					"exception_type": exception_type,
					"message": message,
					"blocking": 1,
				}
			)
			if persist_exceptions:
				exception_doc = _ensure_open_exception(
					planning_run=run_name,
					severity="Critical",
					exception_type=exception_type,
					message=message,
					item_code=result.item_code,
					customer=result.customer,
					source_doctype="APS Schedule Result",
					source_name=result.name,
					resolution_hint=_("Maintain Mold / Mold Product master data before formal APS approval."),
					is_blocking=1,
					diagnostic={
						"root_cause_codes": [exception_type.replace(" ", "_").upper()],
						"root_cause_text": message,
						"suggested_actions": [
							"Complete mold master data before recalculating or approving APS.",
						],
						"candidate_molds": selected_molds or ([result.primary_mould_reference] if result.primary_mould_reference else []),
					},
				)
				exception_names.append(exception_doc.name)
	return {
		"run": run_name,
		"rows": rows,
		"blocking_count": len(rows),
		"exception_names": exception_names,
	}


def _validate_run_segment_overlaps(run_name: str, persist_exceptions: bool = False) -> dict[str, Any]:
	rows = frappe.db.sql(
		"""
		select
			seg.name,
			seg.parent,
			seg.workstation,
			seg.start_time,
			seg.end_time,
			seg.segment_kind,
			res.item_code,
			res.customer
		from `tabAPS Schedule Segment` seg
		inner join `tabAPS Schedule Result` res on res.name = seg.parent
		where res.planning_run = %s
			and seg.parenttype = 'APS Schedule Result'
			and ifnull(seg.segment_kind, '') != 'Family Co-Product'
			and ifnull(seg.workstation, '') != ''
		order by seg.workstation asc, seg.start_time asc, seg.end_time asc
		""",
		[run_name],
		as_dict=True,
	)
	by_workstation = defaultdict(list)
	for row in rows:
		if row.get("start_time") and row.get("end_time"):
			by_workstation[row.get("workstation")].append(row)

	messages = []
	exception_names = []
	for workstation, segments in by_workstation.items():
		segments = sorted(segments, key=lambda row: (get_datetime(row.get("start_time")), get_datetime(row.get("end_time"))))
		for previous, current in zip(segments, segments[1:]):
			if get_datetime(current.get("start_time")) < get_datetime(previous.get("end_time")):
				message = _(
					"Workstation {0} has overlapping primary segments {1} and {2}."
				).format(workstation, previous.get("name"), current.get("name"))
				messages.append(message)
				if persist_exceptions:
					exception_doc = _ensure_open_exception(
						planning_run=run_name,
						severity="Critical",
						exception_type="Primary Segment Overlap",
						message=message,
						item_code=current.get("item_code"),
						customer=current.get("customer"),
						workstation=workstation,
						source_doctype="APS Schedule Segment",
						source_name=current.get("name"),
						resolution_hint=_("Adjust sequence or machine assignment before formal approval."),
						is_blocking=1,
						diagnostic={
							"root_cause_codes": ["WORKSTATION_PRIMARY_OVERLAP"],
							"root_cause_text": "There are two primary schedule segments on the same workstation at the same time: {0} / {1}.".format(previous.get("name"), current.get("name")),
							"suggested_actions": [
								"Adjust sequence or change workstation in the board to ensure only one primary segment exists in the same time window.",
							],
							"candidate_workstations": [workstation],
						},
					)
					exception_names.append(exception_doc.name)
	return {"run": run_name, "count": len(messages), "messages": messages, "exception_names": exception_names}


def _validate_run_mold_overlaps(run_name: str, persist_exceptions: bool = False) -> dict[str, Any]:
	rows = frappe.db.sql(
		"""
		select
			seg.name,
			seg.parent,
			seg.workstation,
			seg.mould_reference,
			seg.start_time,
			seg.end_time,
			seg.segment_kind,
			res.item_code,
			res.customer
		from `tabAPS Schedule Segment` seg
		inner join `tabAPS Schedule Result` res on res.name = seg.parent
		where res.planning_run = %s
			and seg.parenttype = 'APS Schedule Result'
			and ifnull(seg.segment_kind, '') != 'Family Co-Product'
			and ifnull(seg.mould_reference, '') != ''
		order by seg.mould_reference asc, seg.start_time asc, seg.end_time asc
		""",
		[run_name],
		as_dict=True,
	)
	by_mold = defaultdict(list)
	for row in rows:
		if row.get("start_time") and row.get("end_time"):
			by_mold[row.get("mould_reference")].append(row)

	messages = []
	exception_names = []
	for mold_name, segments in by_mold.items():
		segments = sorted(segments, key=lambda row: (get_datetime(row.get("start_time")), get_datetime(row.get("end_time"))))
		for previous, current in zip(segments, segments[1:]):
			if get_datetime(current.get("start_time")) < get_datetime(previous.get("end_time")):
				message = _(
					"Mold {0} overlaps across workstations {1} and {2} for segments {3} and {4}."
				).format(
					mold_name,
					previous.get("workstation") or "-",
					current.get("workstation") or "-",
					previous.get("name"),
					current.get("name"),
				)
				messages.append(message)
				if persist_exceptions:
					exception_doc = _ensure_open_exception(
						planning_run=run_name,
						severity="Critical",
						exception_type="Mold Occupancy Overlap",
						message=message,
						item_code=current.get("item_code"),
						customer=current.get("customer"),
						workstation=current.get("workstation"),
						source_doctype="APS Schedule Segment",
						source_name=current.get("name"),
						resolution_hint=_("Keep one mold on one machine at a time or split with another Mold master."),
						is_blocking=1,
						diagnostic={
							"root_cause_codes": ["MOLD_OCCUPANCY_OVERLAP"],
							"root_cause_text": "The same mold {0} is scheduled on both {1} and {2} at the same time.".format(
								mold_name,
								previous.get("workstation") or "-",
								current.get("workstation") or "-",
							),
							"suggested_actions": [
								"Keep only one primary schedule, or use another independent Mold master as the copy mold.",
								"Open the board, locate the conflicting segments, and reschedule them.",
							],
							"candidate_molds": [mold_name],
							"candidate_workstations": [previous.get("workstation"), current.get("workstation")],
						},
					)
					exception_names.append(exception_doc.name)
	return {"run": run_name, "count": len(messages), "messages": messages, "exception_names": exception_names}


def _get_primary_segments_for_result(result_name: str) -> list[dict[str, Any]]:
	return frappe.get_all(
		"APS Schedule Segment",
		filters={
			"parent": result_name,
			"parenttype": "APS Schedule Result",
			"segment_kind": ("!=", "Family Co-Product"),
		},
		fields=[
			"name",
			"workstation",
			"plant_floor",
			"start_time",
			"end_time",
			"planned_qty",
			"mould_reference",
			"campaign_key",
			"anchor_strength",
			"execution_anchor_source",
			"segment_status",
			"linked_work_order",
			"linked_work_order_scheduling",
			"linked_scheduling_item",
		],
		order_by="sequence_no asc, idx asc",
	)


def _get_work_order_reconciliation_snapshot(work_order_name: str) -> dict[str, Any] | None:
	if not work_order_name or not frappe.db.exists("Work Order", work_order_name):
		return None
	wo = frappe.get_all(
		"Work Order",
		filters={"name": work_order_name},
		fields=[
			"name",
			"company",
			"production_item",
			"qty",
			"produced_qty",
			"material_transferred_for_manufacturing",
			"planned_start_date",
			"planned_end_date",
			"status",
			"custom_aps_result_reference",
			"custom_aps_run",
			"custom_aps_schedule_reference",
			"custom_aps_proposal_batch",
			"custom_aps_locked_for_reschedule",
		],
		limit=1,
	)
	if not wo:
		return None
	wo = dict(wo[0])
	scheduling_rows = []
	if frappe.db.exists("DocType", "Scheduling Item"):
		scheduling_rows = frappe.db.sql(
			"""
			select
				si.name,
				si.parent as work_order_scheduling,
				si.workstation,
				si.scheduling_qty,
				si.planned_start_date,
				si.planned_end_date,
				si.from_time,
				si.to_time,
				si.completed_qty,
				si.custom_aps_segment_reference,
				si.custom_aps_result_reference,
				wos.status as scheduling_status,
				wos.plant_floor,
				wos.posting_date,
				wos.shift_type,
				seg.campaign_key,
				seg.mould_reference
			from `tabScheduling Item` si
			inner join `tabWork Order Scheduling` wos on wos.name = si.parent
			left join `tabAPS Schedule Segment` seg on seg.name = si.custom_aps_segment_reference
			where si.work_order = %s
			order by ifnull(si.from_time, si.planned_start_date) asc, ifnull(si.to_time, si.planned_end_date) asc
			""",
			[work_order_name],
			as_dict=True,
		)
	in_execution = any(
		(row.get("scheduling_status") or "") in FROZEN_SCHEDULING_STATUSES
		or row.get("from_time")
		or flt(row.get("completed_qty")) > 0
		for row in scheduling_rows
	)
	has_execution = (
		in_execution
		or flt(wo.get("produced_qty")) > 0
		or flt(wo.get("material_transferred_for_manufacturing")) > 0
	)
	wo["scheduling_rows"] = scheduling_rows
	wo["campaign_keys"] = [row.get("campaign_key") for row in scheduling_rows if row.get("campaign_key")]
	wo["workstations"] = [row.get("workstation") for row in scheduling_rows if row.get("workstation")]
	wo["has_execution"] = has_execution
	wo["in_execution"] = in_execution
	wo["can_cancel_unstarted"] = not has_execution and (wo.get("status") or "") in ("Submitted", "Not Started")
	wo["can_update_existing"] = (wo.get("status") or "") in ("Submitted", "Not Started")
	return wo


def _find_existing_work_order_for_result(
	result_name: str,
	item_code: str,
	company: str | None = None,
	preferred_workstation: str | None = None,
	preferred_campaign_key: str | None = None,
	excluded_work_orders: list[str] | None = None,
) -> dict[str, Any] | None:
	item_name = _normalize_item_code(item_code) or item_code
	excluded_work_orders = [name for name in (excluded_work_orders or []) if name]
	rows = frappe.get_all(
		"Work Order",
		filters=_strip_none(
			{
				"company": company,
				"production_item": item_name,
				"docstatus": 1,
				"status": ("not in", ["Completed", "Closed", "Cancelled"]),
			}
		),
		fields=["name", "custom_aps_result_reference", "planned_start_date", "creation"],
		order_by="planned_start_date asc, creation asc",
	)
	if not rows:
		return None
	best_snapshot = None
	best_score = None
	for row in rows:
		if row.get("name") in excluded_work_orders:
			continue
		snapshot = _get_work_order_reconciliation_snapshot(row.get("name"))
		if not snapshot:
			continue
		score = (
			0 if snapshot.get("custom_aps_result_reference") == result_name else 1,
			0 if preferred_campaign_key and preferred_campaign_key in (snapshot.get("campaign_keys") or []) else 1,
			0 if preferred_workstation and preferred_workstation in (snapshot.get("workstations") or []) else 1,
			0 if snapshot.get("in_execution") else 1,
			0 if snapshot.get("has_execution") else 1,
			get_datetime(snapshot.get("planned_start_date") or now_datetime()),
			snapshot.get("name"),
		)
		if best_score is None or score < best_score:
			best_score = score
			best_snapshot = snapshot
	return best_snapshot


def _classify_work_order_action(
	existing: dict[str, Any] | None,
	proposed_qty: float,
	prefer_update_existing: bool = False,
) -> str:
	if not existing:
		return "New"
	existing_qty = flt(existing.get("qty"))
	produced_qty = flt(existing.get("produced_qty"))
	if proposed_qty <= 0:
		return "Close Residual" if existing.get("has_execution") else "Cancel Unstarted"
	if abs(existing_qty - proposed_qty) < 0.0001:
		return "Keep Existing"
	if existing.get("has_execution"):
		if proposed_qty > existing_qty + 0.0001:
			return "Update Existing" if prefer_update_existing else "Create Delta"
		if proposed_qty > produced_qty + 0.0001:
			return "Update Existing"
		return "Close Residual"
	return "Update Existing"


def _get_open_aps_managed_work_orders(company: str | None = None) -> list[dict[str, Any]]:
	rows = frappe.get_all(
		"Work Order",
		filters=_strip_none(
			{
				"company": company,
				"docstatus": 1,
				"status": ("not in", ["Completed", "Closed", "Cancelled"]),
			}
		),
		fields=["name", "custom_aps_run", "custom_aps_result_reference", "custom_aps_locked_for_reschedule"],
		order_by="planned_start_date asc, creation asc",
	)
	result = []
	for row in rows:
		if not (row.get("custom_aps_run") or row.get("custom_aps_result_reference") or cint(row.get("custom_aps_locked_for_reschedule"))):
			continue
		snapshot = _get_work_order_reconciliation_snapshot(row.get("name"))
		if snapshot:
			result.append(snapshot)
	return result


def _set_run_result_segment_status(
	run_name: str,
	run_status: str,
	result_status: str,
	segment_status: str,
	flow_step: str | None = None,
	next_step_hint: str | None = None,
	result_names: list[str] | tuple[str, ...] | None = None,
	segment_names: list[str] | tuple[str, ...] | None = None,
):
	frappe.db.set_value("APS Planning Run", run_name, "status", run_status)
	result_names = [name for name in (result_names or []) if name]
	segment_names = [name for name in (segment_names or []) if name]
	result_set_clause = ["status = %s"]
	result_params: list[Any] = [result_status]
	if flow_step is not None:
		result_set_clause.append("flow_step = %s")
		result_params.append(flow_step)
	if next_step_hint is not None:
		result_set_clause.append("next_step_hint = %s")
		result_params.append(next_step_hint)
	result_sql = f"""
		update `tabAPS Schedule Result`
		set {", ".join(result_set_clause)}
		where planning_run = %s
	"""
	result_params.append(run_name)
	if result_names:
		result_sql += " and name in %s"
		result_params.append(tuple(result_names))
	if result_status != "Blocked":
		result_sql += " and ifnull(status, '') != 'Blocked'"
	frappe.db.sql(result_sql, result_params)

	segment_sql = """
		update `tabAPS Schedule Segment`
		set segment_status = %s
		where parenttype = 'APS Schedule Result'
			and parent in (
				select name
				from `tabAPS Schedule Result`
				where planning_run = %s
			)
	"""
	segment_params: list[Any] = [segment_status, run_name]
	if result_names:
		segment_sql += " and parent in %s"
		segment_params.append(tuple(result_names))
	if segment_names:
		segment_sql += " and name in %s"
		segment_params.append(tuple(segment_names))
	if segment_status != "Blocked":
		segment_sql += " and ifnull(segment_status, '') != 'Blocked'"
	frappe.db.sql(segment_sql, segment_params)


def _create_formal_work_order(
	run_doc,
	result,
	qty: float,
	start_time,
	end_time,
	settings: dict[str, Any],
	proposal_batch: str,
) -> str:
	item_code = _resolve_item_name(result.item_code) or result.item_code
	bom_no = frappe.db.get_value("Item", item_code, "default_bom")
	if not bom_no:
		bom_no = frappe.db.get_value("BOM", {"item": item_code, "is_default": 1, "is_active": 1}, "name")
	if not bom_no:
		frappe.throw(_("No BOM was found for {0}.").format(item_code))
	plant_floor_doc = None
	primary_segments = _get_primary_segments_for_result(result.name)
	work_order_floor = _get_primary_result_plant_floor(primary_segments, run_doc.plant_floor)
	if work_order_floor and frappe.db.exists("DocType", "Plant Floor"):
		plant_floor_doc = frappe.get_doc("Plant Floor", work_order_floor)
	warehouse_values = _get_work_order_warehouse_values(
		plant_floor_doc=plant_floor_doc,
		settings=settings,
		item_code=item_code,
		company=run_doc.company,
	)
	work_order = frappe.get_doc(
		{
			"doctype": "Work Order",
			"production_item": item_code,
			"bom_no": bom_no,
			"qty": qty,
			"company": run_doc.company,
			"planned_start_date": start_time,
			"planned_end_date": end_time,
			"wip_warehouse": warehouse_values.get("wip_warehouse"),
			"source_warehouse": warehouse_values.get("source_warehouse"),
			"fg_warehouse": warehouse_values.get("fg_warehouse"),
			"scrap_warehouse": warehouse_values.get("scrap_warehouse"),
			"custom_aps_run": run_doc.name,
			"custom_aps_source": result.demand_source or "APS Planning Run",
			"custom_aps_required_delivery_date": result.requested_date,
			"custom_aps_is_urgent": result.is_urgent,
			"custom_aps_release_status": "Planned",
			"custom_aps_locked_for_reschedule": 1,
			"custom_aps_schedule_reference": result.name,
			"custom_aps_result_reference": result.name,
			"custom_aps_proposal_batch": proposal_batch,
		}
	)
	work_order.insert(ignore_permissions=True)
	work_order.submit()
	return work_order.name


def _save_work_order_with_controller(work_order, ignore_submitted_validation: bool = True):
	if ignore_submitted_validation:
		work_order.flags.ignore_validate_update_after_submit = True
	work_order.save(ignore_permissions=True)


def _link_existing_work_order_to_result(
	work_order_name: str,
	run_name: str,
	result_name: str,
	proposal_batch: str,
	required_delivery_date=None,
):
	values = {
		"custom_aps_run": run_name,
		"custom_aps_release_status": "Planned",
		"custom_aps_locked_for_reschedule": 1,
		"custom_aps_schedule_reference": result_name,
		"custom_aps_result_reference": result_name,
		"custom_aps_proposal_batch": proposal_batch,
	}
	if required_delivery_date:
		values["custom_aps_required_delivery_date"] = required_delivery_date
	work_order = frappe.get_doc("Work Order", work_order_name)
	for fieldname, value in values.items():
		work_order.set(fieldname, value)
	_save_work_order_with_controller(work_order)


def _update_existing_work_order(
	work_order_name: str,
	run_name: str,
	result_name: str,
	proposal_batch: str,
	qty: float,
	start_time,
	end_time,
	required_delivery_date=None,
) -> str:
	snapshot = _get_work_order_reconciliation_snapshot(work_order_name)
	if not snapshot or not snapshot.get("can_update_existing"):
		frappe.throw(_("Work Order {0} can no longer be rewritten by APS.").format(work_order_name))
	old_qty = max(flt(snapshot.get("qty")), 1)
	new_qty = max(flt(qty), 0)
	if new_qty <= 0:
		frappe.throw(_("Work Order {0} cannot be updated to zero qty.").format(work_order_name))
	if new_qty + 0.0001 < flt(snapshot.get("produced_qty")):
		frappe.throw(
			_("Work Order {0} cannot be reduced below already produced qty {1}.").format(
				work_order_name,
				frappe.format(flt(snapshot.get("produced_qty")), {"fieldtype": "Float"}),
			)
		)
	ratio = new_qty / old_qty if old_qty else 1
	values = {
		"qty": new_qty,
		"planned_start_date": start_time,
		"planned_end_date": end_time,
		"custom_aps_run": run_name,
		"custom_aps_release_status": "Planned",
		"custom_aps_locked_for_reschedule": 1,
		"custom_aps_schedule_reference": result_name,
		"custom_aps_result_reference": result_name,
		"custom_aps_proposal_batch": proposal_batch,
	}
	if required_delivery_date:
		values["custom_aps_required_delivery_date"] = required_delivery_date
	work_order = frappe.get_doc("Work Order", work_order_name)
	for fieldname, value in values.items():
		work_order.set(fieldname, value)
	for row in work_order.get("required_items") or []:
		if cint(row.get("is_additional_item")):
			continue
		required_qty = flt(row.get("required_qty")) * ratio
		amount = flt(row.get("rate")) * required_qty if flt(row.get("rate")) else 0
		row.required_qty = required_qty
		row.amount = amount
	for row in work_order.get("operations") or []:
		if flt(row.get("completed_qty")) > 0 or (row.get("status") or "") == "Completed":
			continue
		time_in_mins = flt(row.get("time_in_mins")) * ratio if flt(row.get("time_in_mins")) else 0
		planned_operating_cost = flt(row.get("planned_operating_cost")) * ratio if flt(row.get("planned_operating_cost")) else 0
		row.time_in_mins = time_in_mins
		row.planned_operating_cost = planned_operating_cost
	_save_work_order_with_controller(work_order)
	return work_order_name


def _cancel_unstarted_work_order(work_order_name: str) -> str:
	snapshot = _get_work_order_reconciliation_snapshot(work_order_name)
	if not snapshot or not snapshot.get("can_cancel_unstarted"):
		frappe.throw(_("Work Order {0} can no longer be cancelled as unstarted.").format(work_order_name))
	work_order = frappe.get_doc("Work Order", work_order_name)
	work_order.cancel()
	return work_order.name


def _close_residual_work_order(
	work_order_name: str,
	run_name: str,
	result_name: str | None = None,
	proposal_batch: str | None = None,
	required_delivery_date=None,
) -> str:
	from erpnext.manufacturing.doctype.work_order.work_order import close_work_order

	close_work_order(work_order_name, "Closed")
	values = {
		"custom_aps_run": run_name,
		"custom_aps_release_status": "Locked",
		"custom_aps_locked_for_reschedule": 1,
	}
	if result_name:
		values["custom_aps_schedule_reference"] = result_name
		values["custom_aps_result_reference"] = result_name
	if proposal_batch:
		values["custom_aps_proposal_batch"] = proposal_batch
	if required_delivery_date:
		values["custom_aps_required_delivery_date"] = required_delivery_date
	work_order = frappe.get_doc("Work Order", work_order_name)
	for fieldname, value in values.items():
		work_order.set(fieldname, value)
	_save_work_order_with_controller(work_order)
	return work_order_name


def _is_frozen_scheduling_row(row: dict[str, Any] | None) -> bool:
	row = row or {}
	return (
		(row.get("scheduling_status") or "") in FROZEN_SCHEDULING_STATUSES
		or bool(row.get("from_time"))
		or flt(row.get("completed_qty")) > 0
	)


def _get_formal_scheduling_reconciliation_rows(
	work_order_name: str,
	release_to=None,
) -> list[dict[str, Any]]:
	snapshot = _get_work_order_reconciliation_snapshot(work_order_name)
	rows = [dict(row) for row in (snapshot or {}).get("scheduling_rows") or []]
	limit_date = getdate(release_to) if release_to else None
	result = []
	for row in rows:
		posting_date = getdate(row.get("posting_date")) if row.get("posting_date") else None
		if limit_date and posting_date and posting_date > limit_date:
			continue
		row["is_frozen"] = 1 if _is_frozen_scheduling_row(row) else 0
		row["posting_date"] = posting_date
		result.append(row)
	return result


def _remove_scheduling_item_from_doc(doc, scheduling_item_name: str):
	child = next((row for row in doc.get("scheduling_items") if row.name == scheduling_item_name), None)
	if child:
		doc.remove(child)


def _cleanup_empty_formal_scheduling_doc(docname: str):
	if not docname or not frappe.db.exists("Work Order Scheduling", docname):
		return
	doc = frappe.get_doc("Work Order Scheduling", docname)
	if doc.get("scheduling_items"):
		doc.save(ignore_permissions=True)
		return
	if (doc.status or "") in FROZEN_SCHEDULING_STATUSES:
		doc.save(ignore_permissions=True)
		return
	frappe.delete_doc("Work Order Scheduling", doc.name, force=1, ignore_permissions=True)


def _drop_unfrozen_scheduling_rows_for_work_order(
	work_order_name: str,
	planning_run: str,
	source_doctype: str,
	source_name: str,
) -> dict[str, Any]:
	removed_rows = []
	frozen_rows = []
	for row in _get_formal_scheduling_reconciliation_rows(work_order_name):
		if _is_frozen_scheduling_row(row):
			frozen_rows.append(row)
			continue
		doc = frappe.get_doc("Work Order Scheduling", row.get("work_order_scheduling"))
		_remove_scheduling_item_from_doc(doc, row.get("name"))
		doc.custom_aps_run = planning_run
		doc.custom_aps_freeze_state = "Open"
		doc.custom_aps_approval_state = "Approved"
		doc.save(ignore_permissions=True)
		_cleanup_empty_formal_scheduling_doc(doc.name)
		removed_rows.append(row)
	if frozen_rows:
		_ensure_open_exception(
			planning_run=planning_run,
			severity="Warning",
			exception_type="Residual Scheduling Frozen",
			message=_("Work Order {0} still has frozen formal scheduling rows.").format(work_order_name),
			item_code=frappe.db.get_value("Work Order", work_order_name, "production_item"),
			source_doctype=source_doctype,
			source_name=source_name,
			resolution_hint=_("Frozen scheduling rows were preserved and require manual completion or cancellation handling."),
			is_blocking=0,
		)
	return {"removed_rows": removed_rows, "frozen_rows": frozen_rows}


def _get_or_create_formal_shift_scheduling_doc(
	company: str,
	plant_floor: str,
	posting_date,
	shift_type: str,
	planning_run: str,
	batch_name: str,
):
	existing_name = frappe.db.get_value(
		"Work Order Scheduling",
		{
			"posting_date": getdate(posting_date),
			"company": company,
			"plant_floor": plant_floor,
			"shift_type": shift_type,
		},
		"name",
		order_by="modified desc",
	)
	if existing_name:
		doc = frappe.get_doc("Work Order Scheduling", existing_name)
		if (doc.status or "") in FROZEN_SCHEDULING_STATUSES:
			frappe.throw(_("Work Order Scheduling {0} is already frozen by execution.").format(doc.name))
	else:
		doc = frappe.get_doc(
			{
				"doctype": "Work Order Scheduling",
				"posting_date": getdate(posting_date),
				"company": company,
				"plant_floor": plant_floor,
				"shift_type": shift_type,
				"purpose": "Manufacture",
				"status": "",
			}
		)
	doc.custom_aps_run = planning_run
	doc.custom_aps_freeze_state = "Open"
	doc.custom_aps_approval_state = "Approved"
	doc.remarks = "\n".join(part for part in [doc.remarks, _("APS Shift Proposal {0}").format(batch_name)] if part)
	return doc


def _find_matching_scheduling_row(
	work_order_name: str,
	segment: dict[str, Any],
	matched_row_names: set[str] | None = None,
	release_to=None,
) -> dict[str, Any] | None:
	matched_row_names = matched_row_names or set()
	rows = [
		row
		for row in _get_formal_scheduling_reconciliation_rows(work_order_name, release_to=release_to)
		if row.get("name") not in matched_row_names and not row.get("is_frozen")
	]
	if not rows:
		return None
	target_posting_date = getdate(segment.get("start_time")) if segment.get("start_time") else None
	target_shift_type = _determine_shift_type(segment.get("start_time")) if segment.get("start_time") else ""
	target_campaign_key = segment.get("campaign_key") or _build_campaign_key(
		segment.get("primary_item_code"),
		segment.get("mould_reference"),
		segment.get("workstation"),
	)
	target_start = get_datetime(segment.get("start_time")) if segment.get("start_time") else now_datetime()
	best_row = None
	best_score = None
	for row in rows:
		row_start = get_datetime(row.get("from_time") or row.get("planned_start_date") or target_start)
		score = (
			0 if row.get("custom_aps_segment_reference") == segment.get("name") else 1,
			0 if target_campaign_key and row.get("campaign_key") == target_campaign_key else 1,
			0 if row.get("workstation") == segment.get("workstation") else 1,
			0 if row.get("posting_date") == target_posting_date and (row.get("shift_type") or "") == target_shift_type else 1,
			abs((row_start - target_start).total_seconds()),
			row.get("name"),
		)
		if best_score is None or score < best_score:
			best_score = score
			best_row = row
	return best_row


def _classify_shift_schedule_action(existing_row: dict[str, Any] | None, segment: dict[str, Any]) -> str:
	if not existing_row:
		return "New"
	target_posting_date = getdate(segment.get("start_time")) if segment.get("start_time") else None
	target_shift_type = _determine_shift_type(segment.get("start_time")) if segment.get("start_time") else ""
	target_plant_floor = segment.get("plant_floor")
	target_workstation = segment.get("workstation")
	target_start = get_datetime(segment.get("start_time")) if segment.get("start_time") else None
	target_end = get_datetime(segment.get("end_time")) if segment.get("end_time") else None
	same_target_doc = (
		existing_row.get("posting_date") == target_posting_date
		and (existing_row.get("shift_type") or "") == target_shift_type
		and (existing_row.get("plant_floor") or "") == (target_plant_floor or "")
	)
	same_values = (
		same_target_doc
		and (existing_row.get("workstation") or "") == (target_workstation or "")
		and abs(flt(existing_row.get("scheduling_qty")) - flt(segment.get("planned_qty"))) < 0.0001
		and get_datetime(existing_row.get("planned_start_date") or target_start) == target_start
		and get_datetime(existing_row.get("planned_end_date") or target_end) == target_end
	)
	if same_values:
		return "Keep Existing"
	if same_target_doc:
		return "Update Existing"
	return "Move Existing"


def _determine_shift_type(start_time) -> str:
	start_dt = get_datetime(start_time)
	return "白班" if 8 <= start_dt.hour < 20 else "晚班"


def _upsert_formal_shift_scheduling(batch, row) -> dict[str, Any]:
	if not frappe.db.exists("DocType", "Work Order Scheduling"):
		frappe.throw(_("Work Order Scheduling is not available in this site."))
	plant_floor = row.plant_floor or batch.plant_floor
	if row.action == "Cancel Existing":
		if not row.existing_scheduling or not row.existing_scheduling_item:
			return {"docname": None, "scheduling_item": None, "message": _("No existing formal scheduling row was found to cancel.")}
		doc = frappe.get_doc("Work Order Scheduling", row.existing_scheduling)
		if (doc.status or "") in FROZEN_SCHEDULING_STATUSES:
			_ensure_open_exception(
				planning_run=batch.planning_run,
				severity="Critical",
				exception_type="Shift Scheduling Frozen",
				message=_("Work Order Scheduling {0} is already in execution status {1}.").format(doc.name, doc.status),
				item_code=row.item_code,
				workstation=row.workstation,
				source_doctype="APS Shift Schedule Proposal Batch",
				source_name=batch.name,
				resolution_hint=_("Create residual replan instead of overwriting frozen shift scheduling."),
				is_blocking=1,
			)
			frappe.throw(_("Work Order Scheduling {0} is already frozen by execution.").format(doc.name))
		_remove_scheduling_item_from_doc(doc, row.existing_scheduling_item)
		doc.custom_aps_run = batch.planning_run
		doc.custom_aps_freeze_state = "Open"
		doc.custom_aps_approval_state = "Approved"
		doc.save(ignore_permissions=True)
		_cleanup_empty_formal_scheduling_doc(doc.name)
		if row.segment_reference and frappe.db.exists("APS Schedule Segment", row.segment_reference):
			frappe.db.set_value(
				"APS Schedule Segment",
				row.segment_reference,
				{
					"linked_work_order_scheduling": None,
					"linked_scheduling_item": None,
				},
			)
		return {
			"docname": doc.name if frappe.db.exists("Work Order Scheduling", doc.name) else None,
			"scheduling_item": None,
			"message": _("Existing formal scheduling row was cancelled."),
		}

	target_doc = _get_or_create_formal_shift_scheduling_doc(
		company=batch.company,
		plant_floor=plant_floor,
		posting_date=row.posting_date,
		shift_type=row.shift_type,
		planning_run=batch.planning_run,
		batch_name=batch.name,
	)
	child_row = None
	source_doc = None
	if row.existing_scheduling and frappe.db.exists("Work Order Scheduling", row.existing_scheduling):
		source_doc = frappe.get_doc("Work Order Scheduling", row.existing_scheduling)
		if (source_doc.status or "") in FROZEN_SCHEDULING_STATUSES:
			_ensure_open_exception(
				planning_run=batch.planning_run,
				severity="Critical",
				exception_type="Shift Scheduling Frozen",
				message=_("Work Order Scheduling {0} is already in execution status {1}.").format(source_doc.name, source_doc.status),
				item_code=row.item_code,
				workstation=row.workstation,
				source_doctype="APS Shift Schedule Proposal Batch",
				source_name=batch.name,
				resolution_hint=_("Create residual replan instead of overwriting frozen shift scheduling."),
				is_blocking=1,
			)
			frappe.throw(_("Work Order Scheduling {0} is already frozen by execution.").format(source_doc.name))
		child_row = next((child for child in source_doc.get("scheduling_items") if child.name == row.existing_scheduling_item), None)

	if child_row and source_doc and source_doc.name != target_doc.name:
		_remove_scheduling_item_from_doc(source_doc, child_row.name)
		source_doc.custom_aps_run = batch.planning_run
		source_doc.custom_aps_freeze_state = "Open"
		source_doc.custom_aps_approval_state = "Approved"
		source_doc.save(ignore_permissions=True)
		_cleanup_empty_formal_scheduling_doc(source_doc.name)
		child_row = None

	if not child_row and row.existing_scheduling_item:
		child_row = next((child for child in target_doc.get("scheduling_items") if child.name == row.existing_scheduling_item), None)
	if not child_row:
		child_row = next(
			(
				child
				for child in target_doc.get("scheduling_items")
				if child.get("work_order") == row.work_order
				and child.get("custom_aps_segment_reference") == row.segment_reference
			),
			None,
		)

	if not child_row:
		child_row = target_doc.append(
			"scheduling_items",
			{
				"work_order": row.work_order,
				"scheduling_qty": row.planned_qty,
				"workstation": row.workstation,
				"planned_start_date": row.planned_start_time,
				"planned_end_date": row.planned_end_time,
				"remarks": row.result_reference,
				"custom_aps_run": batch.planning_run,
				"custom_aps_result_reference": row.result_reference,
				"custom_aps_segment_reference": row.segment_reference,
				"custom_aps_shift_proposal": batch.name,
			},
		)
	else:
		child_row.work_order = row.work_order
		child_row.scheduling_qty = row.planned_qty
		child_row.workstation = row.workstation
		child_row.planned_start_date = row.planned_start_time
		child_row.planned_end_date = row.planned_end_time
		child_row.remarks = row.result_reference
		child_row.custom_aps_run = batch.planning_run
		child_row.custom_aps_result_reference = row.result_reference
		child_row.custom_aps_segment_reference = row.segment_reference
		child_row.custom_aps_shift_proposal = batch.name

	target_doc.save(ignore_permissions=True)
	if row.segment_reference and frappe.db.exists("APS Schedule Segment", row.segment_reference):
		frappe.db.set_value(
			"APS Schedule Segment",
			row.segment_reference,
			{
				"linked_work_order": row.work_order,
				"linked_work_order_scheduling": target_doc.name,
				"linked_scheduling_item": child_row.name,
			},
		)
	message = _("Formal scheduling kept in place.") if row.action == "Keep Existing" else _("Formal scheduling reconciled to {0}.").format(target_doc.name)
	return {"docname": target_doc.name, "scheduling_item": child_row.name, "message": message}


def _get_segment_execution_snapshot(segment) -> dict[str, Any]:
	snapshot = {
		"linked_work_order": segment.linked_work_order or None,
		"linked_work_order_scheduling": segment.linked_work_order_scheduling or None,
		"linked_scheduling_item": segment.linked_scheduling_item or None,
		"actual_completed_qty": 0.0,
		"actual_start_time": None,
		"actual_end_time": None,
		"delay_minutes": 0.0,
		"actual_status": "Not Started",
	}
	scheduling_item = None
	if segment.linked_scheduling_item and frappe.db.exists("Scheduling Item", segment.linked_scheduling_item):
		scheduling_item = frappe.get_doc("Scheduling Item", segment.linked_scheduling_item)
	elif frappe.db.exists("DocType", "Scheduling Item"):
		row = frappe.get_all(
			"Scheduling Item",
			filters={"custom_aps_segment_reference": segment.name},
			fields=["name"],
			limit=1,
			order_by="modified desc",
		)
		if row:
			scheduling_item = frappe.get_doc("Scheduling Item", row[0].name)
	if scheduling_item:
		snapshot["linked_scheduling_item"] = scheduling_item.name
		snapshot["linked_work_order"] = scheduling_item.work_order
		snapshot["linked_work_order_scheduling"] = scheduling_item.parent
		snapshot["actual_completed_qty"] = flt(scheduling_item.completed_qty)
		snapshot["actual_start_time"] = scheduling_item.from_time
		snapshot["actual_end_time"] = scheduling_item.to_time

	work_order = None
	if snapshot["linked_work_order"] and frappe.db.exists("Work Order", snapshot["linked_work_order"]):
		work_order = frappe.get_doc("Work Order", snapshot["linked_work_order"])
	if not scheduling_item and work_order:
		snapshot["actual_completed_qty"] = flt(work_order.produced_qty)

	planned_qty = flt(segment.planned_qty)
	now_value = now_datetime()
	start_time = get_datetime(segment.start_time)
	end_time = get_datetime(segment.end_time)
	actual_qty = flt(snapshot["actual_completed_qty"])
	if actual_qty > planned_qty * 1.02:
		snapshot["actual_status"] = "Overproduced"
	elif actual_qty >= planned_qty and planned_qty > 0:
		snapshot["actual_status"] = "Completed"
	elif snapshot["actual_start_time"] or actual_qty > 0:
		snapshot["actual_status"] = "Running"
		elapsed_minutes = max((now_value - start_time).total_seconds() / 60, 0)
		total_minutes = max((end_time - start_time).total_seconds() / 60, 1)
		elapsed_ratio = elapsed_minutes / total_minutes
		completed_ratio = actual_qty / planned_qty if planned_qty else 0
		if now_value > end_time and actual_qty < planned_qty:
			snapshot["actual_status"] = "Delayed"
		elif elapsed_ratio > 0.4 and completed_ratio + 0.15 < elapsed_ratio:
			snapshot["actual_status"] = "Slow Progress"
	elif now_value > end_time:
		snapshot["actual_status"] = "No Recent Update"
	elif work_order and flt(work_order.produced_qty) > 0:
		snapshot["actual_status"] = "Running"

	if snapshot["actual_status"] in ("Delayed", "No Recent Update", "Slow Progress"):
		snapshot["delay_minutes"] = max((now_value - end_time).total_seconds() / 60, 0)
	return snapshot


def _rollup_result_actual_status(segment_statuses: list[str]) -> str:
	if not segment_statuses:
		return "Not Started"
	priority = [
		"Overproduced",
		"Delayed",
		"Slow Progress",
		"No Recent Update",
		"Running",
		"Completed",
		"Not Started",
	]
	for status in priority:
		if status in segment_statuses:
			return status
	return segment_statuses[0]


def _rollup_delay_minutes(segments) -> float:
	return max((flt(segment.delay_minutes) for segment in segments), default=0.0)


def _sync_execution_exceptions(run_name: str, result_doc):
	execution_types = (
		"Slow Progress",
		"Delayed Execution",
		"No Recent Update",
		"Actual Output Mismatch",
	)
	for name in frappe.get_all(
		"APS Exception Log",
		filters={
			"planning_run": run_name,
			"source_doctype": "APS Schedule Result",
			"source_name": result_doc.name,
			"exception_type": ("in", execution_types),
			"status": "Open",
		},
		pluck="name",
	):
		frappe.db.set_value("APS Exception Log", name, "status", "Closed")

	status_map = {
		"Slow Progress": "Slow Progress",
		"Delayed": "Delayed Execution",
		"No Recent Update": "No Recent Update",
		"Overproduced": "Actual Output Mismatch",
	}
	exception_type = status_map.get(result_doc.actual_status)
	if not exception_type:
		return
	_ensure_open_exception(
		planning_run=run_name,
		severity="Critical" if result_doc.actual_status in ("Delayed", "Overproduced") else "Warning",
		exception_type=exception_type,
		message=_("Execution status for result {0} is {1}.").format(result_doc.name, result_doc.actual_status),
		item_code=result_doc.item_code,
		customer=result_doc.customer,
		source_doctype="APS Schedule Result",
		source_name=result_doc.name,
		resolution_hint=_("Generate a manual replan suggestion if the formal shift schedule needs to change."),
		is_blocking=1 if result_doc.actual_status == "Delayed" else 0,
	)


def _count_today_manufacture_entries(run_name: str) -> int:
	if not frappe.db.exists("DocType", "Stock Entry"):
		return 0
	rows = frappe.db.sql(
		"""
		select count(distinct se.name) as entry_count
		from `tabStock Entry` se
		inner join `tabWork Order` wo on wo.name = se.work_order
		where se.docstatus = 1
			and (
				ifnull(se.stock_entry_type, '') = 'Manufacture'
				or ifnull(se.purpose, '') = 'Manufacture'
			)
			and se.posting_date = %s
			and ifnull(wo.custom_aps_run, '') = %s
		""",
		[today(), run_name],
		as_dict=True,
	)
	return cint(rows[0].entry_count) if rows else 0


def _get_latest_stock_entry_by_work_order(work_order_names: list[str]) -> dict[str, dict[str, Any]]:
	work_order_names = [name for name in work_order_names if name]
	if not work_order_names or not frappe.db.exists("DocType", "Stock Entry"):
		return {}
	rows = frappe.db.sql(
		"""
		select
			se.name,
			se.work_order,
			se.posting_date,
			se.posting_time
		from `tabStock Entry` se
		where se.docstatus = 1
			and se.work_order in ({0})
			and (
				ifnull(se.stock_entry_type, '') = 'Manufacture'
				or ifnull(se.purpose, '') = 'Manufacture'
			)
		order by se.posting_date desc, se.posting_time desc, se.modified desc
		""".format(", ".join(["%s"] * len(work_order_names))),
		work_order_names,
		as_dict=True,
	)
	latest = {}
	for row in rows:
		latest.setdefault(row.work_order, row)
	return latest


def _sync_delivery_plan(run_doc) -> str | None:
	if not frappe.db.exists("DocType", "Delivery Plan"):
		return None

	result_rows = frappe.get_all(
		"APS Schedule Result",
		filters={"planning_run": run_doc.name, "scheduled_qty": (">", 0)},
		fields=["customer", "item_code", "requested_date", "scheduled_qty"],
		order_by="requested_date asc, item_code asc",
	)
	if not result_rows:
		return None

	customer = next((row.customer for row in result_rows if row.customer), None)
	if not customer:
		return None

	dp = frappe.get_doc(
		{
			"doctype": "Delivery Plan",
			"customer": customer,
			"company": run_doc.company,
			"delivery_date": getdate(run_doc.horizon_start),
			"arrival_date": getdate(run_doc.horizon_start),
			"remark": _("Generated by Injection APS run {0}").format(run_doc.name),
			"custom_aps_version": run_doc.name,
			"custom_aps_source": "APS Planning Run",
			"item_qties": [
				{
					"item_code": row.item_code,
					"planned_delivery_qty": row.scheduled_qty,
					"staging_qty": row.scheduled_qty,
					"required_arrival_date": row.requested_date,
				}
				for row in result_rows
			],
		}
	).insert(ignore_permissions=True)
	return dp.name


def _sync_existing_work_orders_to_scheduling(run_doc) -> str | None:
	if not frappe.db.exists("DocType", "Work Order Scheduling"):
		return None

	result_rows = frappe.get_all(
		"APS Schedule Result",
		filters={"planning_run": run_doc.name, "scheduled_qty": (">", 0)},
		fields=["name", "item_code"],
	)
	items = []
	for result in result_rows:
		work_orders = frappe.get_all(
			"Work Order",
			filters={
				"production_item": result.item_code,
				"docstatus": 1,
				"status": ("not in", ["Completed", "Closed", "Cancelled"]),
			},
			fields=["name", "qty", "produced_qty"],
			order_by="planned_start_date asc, creation asc",
			limit=1,
		)
		if not work_orders:
			continue
		segments = frappe.get_all(
			"APS Schedule Segment",
			filters={
				"parent": result.name,
				"parenttype": "APS Schedule Result",
				"segment_kind": ("!=", "Family Co-Product"),
			},
			fields=["workstation", "start_time", "end_time", "planned_qty"],
			limit=1,
			order_by="sequence_no asc, idx asc",
		)
		if not segments:
			continue
		segment = segments[0]
		items.append(
			{
				"work_order": work_orders[0].name,
				"scheduling_qty": segment.planned_qty,
				"workstation": segment.workstation,
				"planned_start_date": segment.start_time,
				"planned_end_date": segment.end_time,
				"remarks": result.name,
			}
		)

	if not items:
		return None

	scheduling = frappe.get_doc(
		{
			"doctype": "Work Order Scheduling",
			"posting_date": today(),
			"company": run_doc.company,
			"plant_floor": run_doc.plant_floor,
			"purpose": "Manufacture",
			"status": "",
			"custom_aps_run": run_doc.name,
			"custom_aps_freeze_state": "Open",
			"custom_aps_approval_state": "Approved",
			"scheduling_items": items,
		}
	).insert(ignore_permissions=True)
	return scheduling.name


def _ensure_released_work_order(run_doc, result: dict[str, Any], segment: dict[str, Any], settings: dict[str, Any]) -> str | None:
	item_code = _resolve_item_name(result["item_code"]) or result["item_code"]
	existing = frappe.get_all(
		"Work Order",
		filters={
			"production_item": item_code,
			"custom_aps_run": run_doc.name,
			"docstatus": 1,
			"status": ("not in", ["Completed", "Closed", "Cancelled"]),
		},
		fields=["name"],
		limit=1,
	)
	if existing:
		return existing[0].name

	bom_no = frappe.db.get_value("Item", item_code, "default_bom")
	if not bom_no:
		bom_no = frappe.db.get_value("BOM", {"item": item_code, "is_default": 1, "is_active": 1}, "name")
	if not bom_no:
		_create_exception(
			planning_run=run_doc.name,
			severity="Critical",
			exception_type="Missing BOM",
			message=_("No BOM was found for {0}, so APS could not release a work order.").format(item_code),
			item_code=item_code,
			customer=result.get("customer"),
			source_doctype="APS Schedule Result",
			source_name=result["name"],
			resolution_hint=_("Set a default BOM before attempting release."),
			is_blocking=1,
		)
		return None

	plant_floor_doc = None
	if run_doc.plant_floor and frappe.db.exists("DocType", "Plant Floor"):
		plant_floor_doc = frappe.get_doc("Plant Floor", run_doc.plant_floor)

	work_order = frappe.get_doc(
		{
			"doctype": "Work Order",
			"production_item": item_code,
			"bom_no": bom_no,
			"qty": segment["planned_qty"],
			"company": run_doc.company,
			"planned_start_date": segment["start_time"],
			"planned_end_date": segment["end_time"],
			"wip_warehouse": _get_doc_field_value(
				plant_floor_doc, settings.get("plant_floor_wip_warehouse_field")
			),
			"source_warehouse": _get_doc_field_value(
				plant_floor_doc, settings.get("plant_floor_source_warehouse_field")
			),
			"fg_warehouse": _get_doc_field_value(
				plant_floor_doc, settings.get("plant_floor_fg_warehouse_field")
			),
			"scrap_warehouse": _get_doc_field_value(
				plant_floor_doc, settings.get("plant_floor_scrap_warehouse_field")
			),
			"custom_aps_run": run_doc.name,
			"custom_aps_source": result.get("demand_source") or "APS Planning Run",
			"custom_aps_required_delivery_date": result.get("requested_date"),
			"custom_aps_is_urgent": result.get("is_urgent"),
			"custom_aps_release_status": "Released",
			"custom_aps_locked_for_reschedule": 1,
			"custom_aps_schedule_reference": result["name"],
		}
	)
	work_order.flags.ignore_mandatory = True
	work_order.insert(ignore_permissions=True)
	work_order.submit()
	return work_order.name


def _create_release_work_order_scheduling(run_doc, release_batch: str, scheduling_items: list[dict[str, Any]]) -> str | None:
	if not scheduling_items or not frappe.db.exists("DocType", "Work Order Scheduling"):
		return None
	doc = frappe.get_doc(
		{
			"doctype": "Work Order Scheduling",
			"posting_date": today(),
			"company": run_doc.company,
			"plant_floor": run_doc.plant_floor,
			"purpose": "Manufacture",
			"status": "",
			"remarks": _("APS Release Batch {0}").format(release_batch),
			"custom_aps_run": run_doc.name,
			"custom_aps_freeze_state": "Locked",
			"custom_aps_approval_state": "Approved",
			"scheduling_items": scheduling_items,
		}
	).insert(ignore_permissions=True)
	return doc.name


def _extract_tonnage_from_name(workstation_name: str | None) -> float:
	if not workstation_name:
		return 0

	name = str(workstation_name)
	patterns = [
		r"(\d+(?:\.\d+)?)\s*[Tt]\b",
		r"(\d+(?:\.\d+)?)\s*吨",
	]
	for pattern in patterns:
		matches = re.findall(pattern, name)
		if matches:
			return flt(matches[-1])
	digits = []
	for token in name.replace("/", " ").replace("_", " ").split():
		filtered = "".join(ch for ch in token if ch.isdigit())
		if filtered:
			digits.append(filtered)
	if not digits:
		return 0
	return flt(max(digits, key=len))


def _get_records_with_any_field_set(doctype: str, fieldnames: list[str]) -> list[str]:
	meta = frappe.get_meta(doctype)
	available = [fieldname for fieldname in fieldnames if meta.has_field(fieldname)]
	if not available:
		return []

	conditions = []
	for fieldname in available:
		field = meta.get_field(fieldname)
		if field.fieldtype in ("Check", "Int", "Float", "Currency", "Percent"):
			conditions.append(f"ifnull(`{fieldname}`, 0) != 0")
		else:
			conditions.append(f"ifnull(`{fieldname}`, '') != ''")

	query = f"select name from `tab{doctype}` where {' or '.join(conditions)}"
	return [row.name for row in frappe.db.sql(query, as_dict=True)]


def _get_doc_field_value(doc, fieldname: str | None):
	if not doc or not fieldname:
		return None
	return doc.get(fieldname) if doc.meta.has_field(fieldname) else None


def _first_valid_warehouse(candidates: list[str | None], company: str | None = None) -> str | None:
	if not frappe.db.exists("DocType", "Warehouse"):
		return None
	for warehouse in candidates:
		if not warehouse or not frappe.db.exists("Warehouse", warehouse):
			continue
		warehouse_company = frappe.db.get_value("Warehouse", warehouse, "company")
		if not company or warehouse_company in (None, "", company):
			return warehouse
	return None


def _get_work_order_warehouse_values(
	plant_floor_doc,
	settings: dict[str, Any],
	item_code: str | None,
	company: str | None,
) -> dict[str, str | None]:
	item_default_warehouse = (
		frappe.db.get_value("Item", item_code, "default_warehouse")
		if item_code and frappe.db.exists("DocType", "Item")
		else None
	)
	wip_warehouse = _first_valid_warehouse(
		[
			_get_doc_field_value(plant_floor_doc, settings.get("plant_floor_wip_warehouse_field")),
			item_default_warehouse,
		],
		company=company,
	)
	source_warehouse = _first_valid_warehouse(
		[
			_get_doc_field_value(plant_floor_doc, settings.get("plant_floor_source_warehouse_field")),
			wip_warehouse,
			item_default_warehouse,
		],
		company=company,
	)
	fg_warehouse = _first_valid_warehouse(
		[
			_get_doc_field_value(plant_floor_doc, settings.get("plant_floor_fg_warehouse_field")),
			item_default_warehouse,
			wip_warehouse,
		],
		company=company,
	)
	scrap_warehouse = _first_valid_warehouse(
		[
			_get_doc_field_value(plant_floor_doc, settings.get("plant_floor_scrap_warehouse_field")),
		],
		company=company,
	)
	values = {
		"wip_warehouse": wip_warehouse,
		"source_warehouse": source_warehouse,
		"fg_warehouse": fg_warehouse,
		"scrap_warehouse": scrap_warehouse,
	}
	work_order_meta = frappe.get_meta("Work Order")
	missing = [
		(work_order_meta.get_field(fieldname).label or fieldname)
		for fieldname, value in values.items()
		if work_order_meta.has_field(fieldname)
		and cint(work_order_meta.get_field(fieldname).reqd)
		and not value
	]
	if missing:
		frappe.throw(
			_("Maintain {0} on Plant Floor or Item before creating Work Orders from APS.").format(
				", ".join(missing)
			)
		)
	return values


def _delete_system_generated_rows(doctype: str, company: str | None = None):
	if not frappe.db.exists("DocType", doctype):
		return
	filters = {"is_system_generated": 1}
	if company and frappe.get_meta(doctype).has_field("company"):
		filters["company"] = company
	names = frappe.get_all(doctype, filters=filters, pluck="name")
	if doctype == "APS Net Requirement" and names and frappe.db.exists("DocType", "APS Schedule Result"):
		frappe.db.sql(
			"""
			update `tabAPS Schedule Result`
			set net_requirement = ''
			where net_requirement in %s
			""",
			[tuple(names)],
		)
	for name in names:
		frappe.delete_doc(doctype, name, force=1, ignore_permissions=True)


def _strip_none(values: dict[str, Any]) -> dict[str, Any]:
	return {key: value for key, value in values.items() if value not in (None, "")}
import re
